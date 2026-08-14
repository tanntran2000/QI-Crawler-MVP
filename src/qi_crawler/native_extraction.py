"""Native, auditable PDF/DOCX/XLSX extraction without OCR or interpretation."""

from __future__ import annotations

import json
import zipfile
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy import select

from .db import Database
from .models import Document, DocumentEvidence, DocumentExtraction

EXTRACTOR_VERSION = "native-v1"
SUPPORTED_FORMATS = frozenset({"PDF", "DOCX", "XLSX"})
_WORD_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
NATIVE_OK = "NATIVE_OK"
EMPTY_PAGE = "EMPTY_PAGE"
NEEDS_OCR = "NEEDS_OCR"
TEXT_ENCODING_SUSPECT = "TEXT_ENCODING_SUSPECT"
TABLE_STRUCTURE_UNCERTAIN = "TABLE_STRUCTURE_UNCERTAIN"
NEEDS_REVIEW = "NEEDS_REVIEW"


class NativeExtractionError(RuntimeError):
    """The immutable source document could not be read natively and safely."""


@dataclass(frozen=True)
class NativeEvidence:
    source_locator: str
    content_type: str
    text: str | None = None
    page_number: int | None = None
    sheet_name: str | None = None
    section_heading: str | None = None
    table: object | None = None
    flags: tuple[str, ...] = (NATIVE_OK,)


@dataclass(frozen=True)
class NativeExtractionResult:
    document_id: int
    extraction_id: int
    outcome: str
    status: str
    file_format: str
    evidence_count: int


class NativeHSMTExtractionService:
    """Extract source-traceable native content; never assesses requirements."""

    def __init__(self, database: Database) -> None:
        self.database = database

    def extract_document(self, document_id: int) -> NativeExtractionResult:
        self.database.require_current_schema()
        with self.database.session() as session:
            document = session.get(Document, document_id)
            if document is None:
                raise NativeExtractionError(f"Khong tim thay tai lieu {document_id}.")
            file_format = (document.file_format or "").upper()
            if file_format not in SUPPORTED_FORMATS:
                raise NativeExtractionError(
                    "Chi ho tro trich xuat native PDF, DOCX va XLSX."
                )
            existing = session.scalar(
                select(DocumentExtraction).where(
                    DocumentExtraction.document_id == document_id,
                    DocumentExtraction.extractor_version == EXTRACTOR_VERSION,
                )
            )
            if existing is not None:
                return NativeExtractionResult(
                    document_id=document_id,
                    extraction_id=existing.id,
                    outcome="ALREADY_EXTRACTED",
                    status=existing.status,
                    file_format=file_format,
                    evidence_count=len(existing.evidence),
                )
            source_path = Path(document.stored_path)
            source_sha256 = document.sha256
            document_version = document.version

        if not source_path.is_file():
            raise NativeExtractionError("Khong tim thay file goc trong Document Store.")
        evidence = self._extract(source_path, file_format)

        with self.database.session() as session:
            existing = session.scalar(
                select(DocumentExtraction).where(
                    DocumentExtraction.document_id == document_id,
                    DocumentExtraction.extractor_version == EXTRACTOR_VERSION,
                )
            )
            if existing is not None:
                return NativeExtractionResult(
                    document_id=document_id,
                    extraction_id=existing.id,
                    outcome="ALREADY_EXTRACTED",
                    status=existing.status,
                    file_format=file_format,
                    evidence_count=len(existing.evidence),
                )
            warnings = _warning_metadata(evidence)
            status = NEEDS_REVIEW if warnings else NATIVE_OK
            extraction = DocumentExtraction(
                document_id=document_id,
                extractor_version=EXTRACTOR_VERSION,
                status=status,
                metadata_json=json.dumps(
                    {
                        "file_format": file_format,
                        "source_sha256": source_sha256,
                        "document_version": document_version,
                        "status": status,
                        "warnings": warnings,
                    },
                    ensure_ascii=False,
                ),
            )
            session.add(extraction)
            session.flush()
            for ordinal, item in enumerate(evidence, start=1):
                session.add(
                    DocumentEvidence(
                        extraction_id=extraction.id,
                        ordinal=ordinal,
                        source_locator=item.source_locator,
                        page_number=item.page_number,
                        sheet_name=item.sheet_name,
                        section_heading=item.section_heading,
                        content_type=item.content_type,
                        text=item.text,
                        table_json=(
                            json.dumps(item.table, ensure_ascii=False)
                            if item.table is not None
                            else None
                        ),
                        metadata_json=json.dumps({"flags": item.flags}),
                    )
                )
            session.flush()
            return NativeExtractionResult(
                document_id=document_id,
                extraction_id=extraction.id,
                outcome="EXTRACTED",
                status=status,
                file_format=file_format,
                evidence_count=len(evidence),
            )

    def _extract(self, source_path: Path, file_format: str) -> list[NativeEvidence]:
        try:
            if file_format == "PDF":
                return self._extract_pdf(source_path)
            if file_format == "DOCX":
                return self._extract_docx(source_path)
            return self._extract_xlsx(source_path)
        except NativeExtractionError:
            raise
        except Exception as exc:  # native parser errors are returned to the operator safely
            raise NativeExtractionError(f"Khong the trich xuat native: {exc}") from exc

    @staticmethod
    def _extract_pdf(source_path: Path) -> list[NativeEvidence]:
        reader = PdfReader(str(source_path))
        result: list[NativeEvidence] = []
        for page_number, page in enumerate(reader.pages, start=1):
            text = (page.extract_text() or "").strip()
            if _has_meaningful_text(text):
                flags = _text_flags(text)
            elif _pdf_has_image_content(page):
                flags = (NEEDS_OCR,)
            else:
                flags = (EMPTY_PAGE,)
            result.append(
                NativeEvidence(
                    source_locator=f"page:{page_number}",
                    page_number=page_number,
                    content_type="TEXT",
                    text=text,
                    flags=flags,
                )
            )
        return result

    @staticmethod
    def _extract_docx(source_path: Path) -> list[NativeEvidence]:
        with zipfile.ZipFile(source_path) as archive:
            try:
                document_root = ElementTree.fromstring(archive.read("word/document.xml"))
                styles_root = ElementTree.fromstring(archive.read("word/styles.xml"))
            except KeyError as exc:
                raise NativeExtractionError("DOCX khong co cau truc Word hop le.") from exc
        heading_styles = {
            style.attrib.get(f"{{{_WORD_NS['w']}}}styleId", "")
            for style in styles_root.findall(".//w:style", _WORD_NS)
            if "heading" in " ".join(
                value.lower() for value in style.attrib.values() if value
            )
            or "heading" in " ".join(
                node.attrib.get(f"{{{_WORD_NS['w']}}}val", "").lower()
                for node in style.findall(".//w:name", _WORD_NS)
            )
        }
        result: list[NativeEvidence] = []
        current_heading: str | None = None
        paragraph_number = 0
        table_number = 0
        body = document_root.find("w:body", _WORD_NS)
        if body is None:
            return result
        for child in body:
            tag = child.tag.rsplit("}", 1)[-1]
            if tag == "p":
                paragraph_number += 1
                text = "".join(node.text or "" for node in child.findall(".//w:t", _WORD_NS)).strip()
                if not text:
                    continue
                style = child.find("w:pPr/w:pStyle", _WORD_NS)
                style_id = style.attrib.get(f"{{{_WORD_NS['w']}}}val", "") if style is not None else ""
                if style_id in heading_styles:
                    current_heading = text
                result.append(
                    NativeEvidence(
                        source_locator=f"word/document.xml:paragraph:{paragraph_number}",
                        content_type="TEXT",
                        text=text,
                        section_heading=current_heading,
                        flags=_text_flags(text),
                    )
                )
            elif tag == "tbl":
                table_number += 1
                table, flags = _extract_docx_table(child)
                rows = table["rows"] if isinstance(table, dict) else table
                table_text = "\n".join(" | ".join(row) for row in rows)
                result.append(
                    NativeEvidence(
                        source_locator=f"word/document.xml:table:{table_number}",
                        content_type="TABLE",
                        text=table_text,
                        section_heading=current_heading,
                        table=table,
                        flags=_combine_flags(flags, _text_flags(table_text)),
                    )
                )
        return result

    @staticmethod
    def _extract_xlsx(source_path: Path) -> list[NativeEvidence]:
        workbook = load_workbook(source_path, read_only=False, data_only=False)
        result: list[NativeEvidence] = []
        try:
            for worksheet in workbook.worksheets:
                merged_ranges = tuple(worksheet.merged_cells.ranges)
                for row_number, row in enumerate(worksheet.iter_rows(), start=1):
                    cells = [
                        {"cell": cell.coordinate, "value": _json_value(cell.value)}
                        for cell in row
                        if cell.value is not None
                    ]
                    if not cells:
                        continue
                    row_merges = [
                        str(merged)
                        for merged in merged_ranges
                        if merged.min_row <= row_number <= merged.max_row
                    ]
                    table: object = cells
                    flags = (NATIVE_OK,)
                    row_text = " | ".join(str(cell["value"]) for cell in cells)
                    if row_merges:
                        table = {"cells": cells, "merged_ranges": row_merges}
                        flags = (TABLE_STRUCTURE_UNCERTAIN,)
                    result.append(
                        NativeEvidence(
                            source_locator=f"sheet:{worksheet.title}!{cells[0]['cell']}:{cells[-1]['cell']}",
                            sheet_name=worksheet.title,
                            section_heading=worksheet.title,
                            content_type="TABLE_ROW",
                            text=row_text,
                            table=table,
                            flags=_combine_flags(flags, _text_flags(row_text)),
                        )
                    )
        finally:
            workbook.close()
        return result


def _json_value(value: object) -> object:
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    return str(value)


def _has_meaningful_text(text: str) -> bool:
    return sum(character.isalnum() for character in text) >= 2


def _text_flags(text: str) -> tuple[str, ...]:
    if "\ufffd" in text:
        return (TEXT_ENCODING_SUSPECT,)
    visible = [character for character in text if not character.isspace()]
    controls = sum(ord(character) < 32 for character in visible)
    mojibake = ("Ãƒ", "Ã‚", "â€")
    if (visible and controls / len(visible) > 0.05) or any(token in text for token in mojibake):
        return (TEXT_ENCODING_SUSPECT,)
    return (NATIVE_OK,)


def _pdf_has_image_content(page: object) -> bool:
    try:
        resources = page.get("/Resources")
        if resources is None:
            return False
        xobjects = resources.get("/XObject")
        if xobjects is None:
            return False
        for value in xobjects.values():
            item = value.get_object()
            if str(item.get("/Subtype", "")) == "/Image":
                return True
    except (AttributeError, KeyError, TypeError):
        return False
    return False


def _extract_docx_table(table_node: ElementTree.Element) -> tuple[object, tuple[str, ...]]:
    rows: list[list[str]] = []
    merged_cells: list[dict[str, object]] = []
    for row_number, row in enumerate(table_node.findall("w:tr", _WORD_NS), start=1):
        values: list[str] = []
        for column_number, cell in enumerate(row.findall("w:tc", _WORD_NS), start=1):
            values.append(
                "".join(node.text or "" for node in cell.findall(".//w:t", _WORD_NS)).strip()
            )
            properties = cell.find("w:tcPr", _WORD_NS)
            if properties is None:
                continue
            grid_span = properties.find("w:gridSpan", _WORD_NS)
            vertical_merge = properties.find("w:vMerge", _WORD_NS)
            span = int(grid_span.attrib.get(f"{{{_WORD_NS['w']}}}val", "1")) if grid_span is not None else 1
            if span > 1 or vertical_merge is not None:
                merged_cells.append(
                    {
                        "row": row_number,
                        "column": column_number,
                        "grid_span": span,
                        "vertical_merge": (
                            vertical_merge.attrib.get(f"{{{_WORD_NS['w']}}}val", "continue")
                            if vertical_merge is not None
                            else None
                        ),
                    }
                )
        rows.append(values)
    if merged_cells:
        return (
            {"rows": rows, "merged_cells": merged_cells},
            (TABLE_STRUCTURE_UNCERTAIN,),
        )
    return rows, (NATIVE_OK,)


def _combine_flags(*flag_sets: tuple[str, ...]) -> tuple[str, ...]:
    warnings = tuple(
        dict.fromkeys(flag for flags in flag_sets for flag in flags if flag != NATIVE_OK)
    )
    return warnings or (NATIVE_OK,)


def _warning_metadata(evidence: list[NativeEvidence]) -> list[dict[str, object]]:
    warnings: list[dict[str, object]] = []
    for item in evidence:
        for flag in item.flags:
            if flag == NATIVE_OK:
                continue
            warnings.append(
                {
                    "flag": flag,
                    "source_locator": item.source_locator,
                    "page_number": item.page_number,
                    "sheet_name": item.sheet_name,
                    "section_heading": item.section_heading,
                }
            )
    if not evidence:
        warnings.append({"flag": EMPTY_PAGE, "source_locator": "document"})
    return warnings
