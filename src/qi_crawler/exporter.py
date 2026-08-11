from __future__ import annotations

import csv
import html
import re
import unicodedata
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from .db import Database
from .excel_safety import safe_excel_row
from .export.tbmt_schema import TBMT_COLUMNS
from .models import InventoryItem, Notice
from .stock import StockCheck, check_stock

HEADERS = [
    "id",
    "notice_code",
    "title",
    "buyer",
    "investor",
    "package_price",
    "currency",
    "published_at",
    "closing_at",
    "location",
    "sector",
    "selection_method",
    "notice_version",
    "source_url",
    "source_kind",
    "data_quality_status",
    "requested_item_count",
    "requested_quantity_details",
    "response_table",
    "attachment_count",
    "attachments_downloaded",
    "attachments_failed",
    "first_seen_at",
    "last_seen_at",
]

TBMT_HEADERS = list(TBMT_COLUMNS)

TBMT_COLUMN_WIDTHS = [22, 34, 34, 40, 45, 55, 36, 18, 25, 27, 24, 18, 20, 36, 30, 28, 23, 24]

HEADER_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FONT = Font(bold=True, color="FFFFFF")
STATUS_FILLS = {
    "MEETS_STOCK": PatternFill("solid", fgColor="C6EFCE"),
    "STOCK_SHORTAGE": PatternFill("solid", fgColor="FFC7CE"),
    "REVIEW_REQUIRED_QUANTITY": PatternFill("solid", fgColor="FFEB9C"),
    "REVIEW_UNIT_MISMATCH": PatternFill("solid", fgColor="FFEB9C"),
    "NOT_IN_VERIFIED_STOCK": PatternFill("solid", fgColor="FFC7CE"),
}

THIN_GRAY = Side(style="thin", color="D9E2F3")
TBMT_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


def _fold(text: str) -> str:
    text = text.replace("Đ", "D").replace("đ", "d")
    normalized = unicodedata.normalize("NFD", text)
    return "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn").lower()


def _raw_lines(notice: Notice) -> list[str]:
    raw_text = html.unescape(notice.raw_text or "")
    return [re.sub(r"\s+", " ", line).strip() for line in raw_text.splitlines() if line.strip()]


def _raw_value(notice: Notice, *labels: str) -> str | None:
    lines = _raw_lines(notice)
    folded_labels = tuple(_fold(label).rstrip(":") for label in labels)
    for index, line in enumerate(lines):
        folded_line = _fold(line)
        for label in folded_labels:
            if folded_line.rstrip(":") == label and index + 1 < len(lines):
                return lines[index + 1]
            if folded_line.startswith(label):
                value = line.split(":", 1)[1].strip() if ":" in line else ""
                if value:
                    return value
    return None


def _notice_kind(notice: Notice) -> str:
    labels = {
        "tbmt": "Thông báo mời thầu",
        "khlcnt": "Kế hoạch lựa chọn nhà thầu",
        "kqlcnt": "Kết quả lựa chọn nhà thầu",
        "kqmt": "Kết quả mở thầu",
    }
    return labels.get((notice.notice_type or "tbmt").lower(), "Thông báo mời thầu")


def _main_content(notice: Notice) -> str | None:
    parts: list[str] = []
    if notice.title:
        parts.append(notice.title)
    if notice.notice_code:
        parts.append(f"Số thông báo: {notice.notice_code}")
    if notice.investor:
        parts.append(f"Chủ đầu tư: {notice.investor}")
    if notice.published_at:
        parts.append(f"Thời điểm đăng tải: {notice.published_at}")
    return "\n".join(parts) or None


def _split_selection_method(notice: Notice) -> tuple[str | None, str | None]:
    procedure = _raw_value(notice, "Phương thức lựa chọn nhà thầu")
    form = _raw_value(notice, "Hình thức lựa chọn nhà thầu")
    fallback = notice.selection_method
    if fallback:
        if not procedure and "giai doan" in _fold(fallback):
            procedure = fallback
        elif not form:
            form = fallback
    return procedure, form


def _related_opening_date(notice: Notice) -> str | None:
    openings = getattr(notice, "bid_openings", [])
    return next((item.opening_date for item in openings if item.opening_date), None)


def _related_contract_duration(notice: Notice) -> str | None:
    results = getattr(notice, "bid_results", [])
    return next((item.contract_duration for item in results if item.contract_duration), None)


def _tbmt_row(notice: Notice, sequence: int) -> list[object]:
    procedure, selection_form = _split_selection_method(notice)
    project_name = _raw_value(notice, "Tên dự án", "Dự án", "Tên kế hoạch")
    buyer_address = _raw_value(
        notice,
        "Địa chỉ bên mời thầu",
        "Địa chỉ của bên mời thầu",
        "Địa chỉ chủ đầu tư",
    )
    release_time = _raw_value(
        notice,
        "Thời gian phát hành E-HSMT",
        "Thời gian phát hành HSMT",
        "Thời điểm phát hành E-HSMT",
        "Thời điểm phát hành HSMT",
    )
    document_price = _raw_value(
        notice,
        "Giá bán 1 bộ E-HSMT",
        "Giá bán 1 bộ HSMT",
        "Giá E-HSMT",
        "Giá HSMT",
    )
    bid_security = _raw_value(
        notice,
        "Giá trị bảo đảm dự thầu",
        "Bảo đảm dự thầu",
    )
    security_form = _raw_value(notice, "Hình thức bảo đảm dự thầu")
    issue_location = _raw_value(
        notice,
        "Địa điểm phát hành E-HSMT",
        "Địa điểm phát hành HSMT",
        "Địa điểm phát hành",
    )
    opening_time = _raw_value(
        notice,
        "Thời gian mở thầu",
        "Thời điểm mở thầu",
        "Ngày mở thầu",
    ) or _related_opening_date(notice)
    contract_duration = _raw_value(
        notice,
        "Thời gian thực hiện hợp đồng",
    ) or _related_contract_duration(notice)

    return [
        f"{sequence}. {_notice_kind(notice)}",
        notice.buyer,
        buyer_address,
        project_name,
        notice.title,
        _main_content(notice),
        notice.funding_source,
        notice.package_price,
        procedure,
        selection_form,
        release_time,
        document_price,
        bid_security,
        security_form,
        issue_location or notice.source_url,
        notice.closing_at,
        opening_time,
        contract_duration,
    ]


def _add_tbmt_sheet(workbook: Workbook, notices: list[Notice]) -> None:
    sheet = workbook.active
    sheet.title = "Bản tin điện tử"
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:R2")
    sheet["A1"] = "QI-CRAWLER - BẢN TIN GÓI THẦU"
    sheet["A1"].font = Font(name="Calibri", size=20, bold=True, color="17365D")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.merge_cells("A3:R3")
    sheet["A3"] = "Biểu mẫu tổng hợp dữ liệu do QI-Crawler thu thập"
    sheet["A3"].font = Font(name="Calibri", size=11, italic=True, color="5B6573")
    sheet.merge_cells("A4:R4")
    sheet["A4"] = f"Thời điểm xuất: {datetime.now(UTC).astimezone().strftime('%d/%m/%Y %H:%M:%S')}"
    sheet["A4"].font = Font(name="Calibri", size=10, color="5B6573")
    sheet.merge_cells("A6:R6")
    sheet["A6"] = (
        "Lưu ý: ô trống nghĩa là nguồn chưa cung cấp hoặc QI-Crawler chưa xác minh được; "
        "cần đối chiếu thông báo và E-HSMT gốc trước khi sử dụng."
    )
    sheet["A6"].font = Font(name="Calibri", size=10, italic=True, color="9C5700")

    sheet.cell(row=9, column=1)
    sheet.append(TBMT_HEADERS)
    for sequence, notice in enumerate(notices, start=1):
        sheet.append(safe_excel_row(_tbmt_row(notice, sequence)))

    sheet.freeze_panes = "A11"
    sheet.auto_filter.ref = f"A10:R{max(sheet.max_row, 10)}"
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[10].height = 54
    for index, width in enumerate(TBMT_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[sheet.cell(10, index).column_letter].width = width

    for cell in sheet[10]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = TBMT_BORDER
    for row_number in range(11, sheet.max_row + 1):
        sheet.row_dimensions[row_number].height = 72
        for cell in sheet[row_number]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = TBMT_BORDER
        sheet.cell(row_number, 8).number_format = "#,##0"
        source_cell = sheet.cell(row_number, 15)
        if isinstance(source_cell.value, str) and source_cell.value.startswith(("http://", "https://")):
            source_cell.hyperlink = source_cell.value
            source_cell.style = "Hyperlink"


def _quantity_details(notice: Notice) -> str:
    if not notice.tender_items:
        return "NOT_AVAILABLE"
    return "; ".join(
        f"{item.product_name}: "
        f"{item.quantity if item.quantity is not None else 'REVIEW'}"
        f" {item.unit or ''}".rstrip()
        for item in notice.tender_items
    )


def _checks(notice: Notice, inventory: list[InventoryItem]) -> list[StockCheck]:
    return [check_stock(item, inventory) for item in notice.tender_items]


def _response_table(checks: list[StockCheck]) -> str:
    if not checks:
        return "NO_QUANTITY_DATA"
    counts: dict[str, int] = {}
    for result in checks:
        counts[result.status] = counts.get(result.status, 0) + 1
    return "; ".join(f"{status}={counts[status]}" for status in sorted(counts))


def _style_sheet(sheet, *, status_column: int | None = None) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), 60)
    if status_column:
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, status_column)
            cell.fill = STATUS_FILLS.get(str(cell.value), PatternFill())


def _data(db: Database) -> tuple[list[Notice], list[InventoryItem]]:
    with db.session() as session:
        notices = session.scalars(
            select(Notice)
            .options(
                selectinload(Notice.attachments),
                selectinload(Notice.tender_items),
                selectinload(Notice.bid_results),
                selectinload(Notice.bid_openings),
            )
            .order_by(Notice.id.desc())
        ).all()
        inventory = session.scalars(select(InventoryItem).order_by(InventoryItem.sku)).all()
        session.expunge_all()
        return list(notices), list(inventory)


def _rows(db: Database) -> list[list[object]]:
    notices, inventory = _data(db)
    return [
            [
                notice.id,
                notice.notice_code,
                notice.title,
                notice.buyer,
                notice.investor,
                notice.package_price,
                notice.currency,
                notice.published_at,
                notice.closing_at,
                notice.location,
                notice.sector,
                notice.selection_method,
                notice.notice_version,
                notice.source_url,
                notice.source_kind,
                notice.data_quality_status,
                len(notice.tender_items),
                _quantity_details(notice),
                _response_table(_checks(notice, inventory)),
                len(notice.attachments),
                sum(item.download_status == "downloaded" for item in notice.attachments),
                sum(item.download_status in {"failed", "manual_review"} for item in notice.attachments),
                notice.first_seen_at.isoformat() if notice.first_seen_at else None,
                notice.last_seen_at.isoformat() if notice.last_seen_at else None,
            ]
            for notice in notices
        ]


def export_csv(db: Database, output: Path) -> Path:
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(HEADERS)
        writer.writerows(safe_excel_row(row) for row in _rows(db))
    return output


def export_xlsx(db: Database, output: Path) -> Path:
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    notices, inventory = _data(db)
    _add_tbmt_sheet(workbook, notices)

    sheet = workbook.create_sheet("Notices")
    sheet.append(HEADERS)
    for notice in notices:
        checks = _checks(notice, inventory)
        row = [
            notice.id,
            notice.notice_code,
            notice.title,
            notice.buyer,
            notice.investor,
            notice.package_price,
            notice.currency,
            notice.published_at,
            notice.closing_at,
            notice.location,
            notice.sector,
            notice.selection_method,
            notice.notice_version,
            notice.source_url,
            notice.source_kind,
            notice.data_quality_status,
            len(notice.tender_items),
            _quantity_details(notice),
            _response_table(checks),
            len(notice.attachments),
            sum(item.download_status == "downloaded" for item in notice.attachments),
            sum(
                item.download_status in {"failed", "manual_review"}
                for item in notice.attachments
            ),
            notice.first_seen_at.isoformat() if notice.first_seen_at else None,
            notice.last_seen_at.isoformat() if notice.last_seen_at else None,
        ]
        sheet.append(safe_excel_row(row))

    item_sheet = workbook.create_sheet("Response Table")
    item_headers = [
        "notice_id",
        "notice_code",
        "requested_product",
        "specification",
        "required_quantity",
        "unit",
        "matched_sku",
        "available_quantity",
        "shortage_quantity",
        "response_status",
        "match_confidence",
        "quantity_source",
        "source_location",
        "extraction_confidence",
        "human_review_required",
        "note",
        "source_url",
    ]
    item_sheet.append(item_headers)
    for notice in notices:
        checks = {result.tender_item_id: result for result in _checks(notice, inventory)}
        for item in notice.tender_items:
            result = checks[item.id]
            item_sheet.append(
                safe_excel_row([
                    notice.id,
                    notice.notice_code,
                    item.product_name,
                    item.specification,
                    result.required_quantity,
                    item.unit,
                    result.inventory_sku,
                    result.available_quantity,
                    result.shortage_quantity,
                    result.status,
                    result.match_confidence,
                    item.source_document,
                    item.source_location,
                    item.extraction_confidence,
                    item.needs_human_review,
                    result.note,
                    notice.source_url,
                ])
            )

    inventory_sheet = workbook.create_sheet("QI Inventory")
    inventory_sheet.append(
        [
            "sku",
            "product_name",
            "aliases",
            "quantity_available",
            "unit",
            "warehouse",
            "verified",
            "updated_at",
            "source_file",
        ]
    )
    for item in inventory:
        inventory_sheet.append(
            safe_excel_row([
                item.sku,
                item.product_name,
                item.aliases,
                item.quantity_available,
                item.unit,
                item.warehouse,
                item.verified,
                item.updated_at.isoformat() if item.updated_at else None,
                item.source_file,
            ])
        )

    _style_sheet(sheet)
    _style_sheet(item_sheet, status_column=10)
    _style_sheet(inventory_sheet)
    workbook.save(output)
    return output
