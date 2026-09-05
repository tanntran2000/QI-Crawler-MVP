"""Bounded, fail-closed importer for synthetic or supplied TBMT workbooks."""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from datetime import UTC, datetime
from enum import StrEnum
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .opportunity_contract import (
    OpportunityCandidate,
    OpportunityImportBatch,
    OpportunitySourceType,
)
from .tbmt_normalization import (
    compact_source_text,
    parse_tbmt_notice_identity,
    parse_tbmt_package_price,
)
from .tbmt_schema import REQUIRED_TBMT_HEADERS, canonical_tbmt_header

MAX_HEADER_SCAN_ROWS = 50
SCHEMA_VERSION = "mi-tbmt-1"

_EXECUTION_LOCATION_FIELD_NAMES = (
    "địa điểm thực hiện gói thầu",
    "địa điểm thực hiện",
    "địa điểm thi công",
    "execution location",
)


def _execution_location_from_raw_fields(raw_fields: dict[str, Any]) -> str | None:
    """Return only an explicitly-labelled execution location from source data.

    The TBMT list export often omits this field. When a detail-shaped source is
    supplied, preserve explicit workbook fields before provinces/structured
    detail, then location -> workAddress, without treating the procuring-entity
    address or issue location as the place where the contract is performed.
    """

    normalized_fields = {
        " ".join(str(key).split()).casefold(): value for key, value in raw_fields.items()
    }
    for field_name in _EXECUTION_LOCATION_FIELD_NAMES:
        text = compact_source_text(normalized_fields.get(field_name))
        if text is not None:
            return text

    provinces = normalized_fields.get("provinces")
    if isinstance(provinces, (list, tuple)):
        names: list[str] = []
        for province in provinces:
            if isinstance(province, dict):
                name = compact_source_text(province.get("name"))
            else:
                name = compact_source_text(province)
            if name is not None:
                names.append(name)
        if names:
            return ", ".join(names)

    structured_names = tuple(
        text
        for field_name in ("provname", "districtname", "wardname")
        if (text := compact_source_text(normalized_fields.get(field_name))) is not None
    )
    if structured_names:
        return ", ".join(structured_names)

    for field_name in ("location", "workaddress"):
        text = compact_source_text(normalized_fields.get(field_name))
        if text is not None:
            return text
    return None


class TBMTIssueCode(StrEnum):
    MISSING_REQUIRED_HEADER = "MISSING_REQUIRED_HEADER"
    DUPLICATE_HEADER = "DUPLICATE_HEADER"
    INVALID_IB_IDENTITY = "INVALID_IB_IDENTITY"
    INVALID_PRICE = "INVALID_PRICE"
    EMPTY_PACKAGE_NAME = "EMPTY_PACKAGE_NAME"
    UNSUPPORTED_WORKBOOK = "UNSUPPORTED_WORKBOOK"
    NO_USABLE_SHEET = "NO_USABLE_SHEET"


class TBMTImportError(ValueError):
    def __init__(self, code: TBMTIssueCode, message: str) -> None:
        self.code = code
        super().__init__(f"{code.value}: {message}")


@dataclass(frozen=True, slots=True)
class TBMTImportIssue:
    code: TBMTIssueCode
    message: str
    source_row: int | None = None
    source_field: str | None = None


@dataclass(frozen=True, slots=True)
class TBMTImportResult:
    batch: OpportunityImportBatch
    candidates: tuple[OpportunityCandidate, ...]
    issues: tuple[TBMTImportIssue, ...]
    headers: tuple[str, ...]
    source_row_count: int


@dataclass(frozen=True, slots=True)
class _HeaderSelection:
    sheet: Worksheet
    row_number: int
    headers: tuple[str | None, ...]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _canonical_headers(values: tuple[Any, ...]) -> tuple[str | None, ...]:
    result: list[str | None] = []
    seen: set[str] = set()
    for value in values:
        header = canonical_tbmt_header(value)
        if header is None:
            result.append(None)
            continue
        key = header.casefold()
        if key in seen:
            raise TBMTImportError(
                TBMTIssueCode.DUPLICATE_HEADER,
                f"Duplicate source header: {header}",
            )
        seen.add(key)
        result.append(header)
    return tuple(result)


def _select_header(workbook: Any, requested_sheet: str | None) -> _HeaderSelection:
    if requested_sheet is not None:
        if requested_sheet not in workbook.sheetnames:
            raise TBMTImportError(
                TBMTIssueCode.NO_USABLE_SHEET,
                f"Requested sheet does not exist: {requested_sheet}",
            )
        sheets = [workbook[requested_sheet]]
    else:
        sheets = list(workbook.worksheets)

    best_partial: tuple[str, set[str]] | None = None
    for sheet in sheets:
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=MAX_HEADER_SCAN_ROWS, values_only=True),
            start=1,
        ):
            canonical_values = tuple(canonical_tbmt_header(value) for value in row)
            keys = {value for value in canonical_values if value is not None}
            overlap = keys.intersection(REQUIRED_TBMT_HEADERS)
            if best_partial is None or len(overlap) > len(best_partial[1]):
                best_partial = (sheet.title, overlap)
            if REQUIRED_TBMT_HEADERS.issubset(keys):
                return _HeaderSelection(
                    sheet=sheet,
                    row_number=row_number,
                    headers=_canonical_headers(tuple(row)),
                )

    if best_partial and best_partial[1]:
        missing = sorted(REQUIRED_TBMT_HEADERS - best_partial[1])
        raise TBMTImportError(
            TBMTIssueCode.MISSING_REQUIRED_HEADER,
            f"Missing required header(s): {', '.join(missing)}",
        )
    raise TBMTImportError(
        TBMTIssueCode.NO_USABLE_SHEET,
        "No sheet contains the required TBMT header contract",
    )


def _raw_row(headers: tuple[str | None, ...], values: tuple[Any, ...]) -> dict[str, Any]:
    return {
        header: values[index] if index < len(values) else None
        for index, header in enumerate(headers)
        if header is not None
    }


def import_tbmt_workbook(
    path: Path,
    *,
    sheet_name: str | None = None,
    imported_at: datetime | None = None,
) -> TBMTImportResult:
    """Read one TBMT workbook into source-backed candidates without persistence."""

    path = Path(path).resolve()
    if path.suffix.lower() != ".xlsx" or not path.is_file():
        raise TBMTImportError(
            TBMTIssueCode.UNSUPPORTED_WORKBOOK,
            "TBMT importer requires an existing .xlsx workbook",
        )
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise TBMTImportError(
            TBMTIssueCode.UNSUPPORTED_WORKBOOK,
            f"Workbook could not be opened: {type(exc).__name__}",
        ) from exc

    try:
        selected = _select_header(workbook, sheet_name)
        batch = OpportunityImportBatch(
            source_filename=path.name,
            source_sha256=_sha256(path),
            sheet=selected.sheet.title,
            imported_at=imported_at or datetime.now(UTC),
            schema_version=SCHEMA_VERSION,
            source_type=OpportunitySourceType.TBMT,
        )
        candidates: list[OpportunityCandidate] = []
        issues: list[TBMTImportIssue] = []
        source_row_count = 0
        for row_number, values in enumerate(
            selected.sheet.iter_rows(
                min_row=selected.row_number + 1,
                values_only=True,
            ),
            start=selected.row_number + 1,
        ):
            raw_fields = _raw_row(selected.headers, tuple(values))
            if not any(value not in (None, "") for value in raw_fields.values()):
                continue
            source_row_count += 1

            package_source = raw_fields.get("GÓI THẦU")
            package_name = compact_source_text(package_source)
            if package_name is None:
                issues.append(
                    TBMTImportIssue(
                        TBMTIssueCode.EMPTY_PACKAGE_NAME,
                        "Package name is missing",
                        source_row=row_number,
                        source_field="GÓI THẦU",
                    )
                )
                continue
            identity = parse_tbmt_notice_identity(package_source)
            if identity is None:
                issues.append(
                    TBMTImportIssue(
                        TBMTIssueCode.INVALID_IB_IDENTITY,
                        "Package source does not contain exactly one revisioned IB identity",
                        source_row=row_number,
                        source_field="GÓI THẦU",
                    )
                )
                continue

            price_source = raw_fields.get("GIÁ GÓI THẦU")
            package_price_raw = compact_source_text(price_source)
            package_price = parse_tbmt_package_price(price_source)
            if package_price_raw is not None and package_price is None:
                issues.append(
                    TBMTImportIssue(
                        TBMTIssueCode.INVALID_PRICE,
                        "Package price is not an unambiguous non-negative amount",
                        source_row=row_number,
                        source_field="GIÁ GÓI THẦU",
                    )
                )

            candidates.append(
                OpportunityCandidate(
                    identity=identity,
                    import_batch=batch,
                    source_row=row_number,
                    package_name=package_name,
                    project=compact_source_text(raw_fields.get("DỰ ÁN")),
                    package_price_raw=package_price_raw,
                    package_price=package_price,
                    funding_source=compact_source_text(raw_fields.get("NGUỒN VỐN")),
                    location_detail_raw=_execution_location_from_raw_fields(raw_fields),
                    raw_fields=raw_fields,
                    provenance={
                        "source_filename": batch.source_filename,
                        "source_sha256": batch.source_sha256,
                        "sheet": batch.sheet,
                        "source_row": row_number,
                    },
                )
            )
        return TBMTImportResult(
            batch=batch,
            candidates=tuple(candidates),
            issues=tuple(issues),
            headers=tuple(header for header in selected.headers if header is not None),
            source_row_count=source_row_count,
        )
    finally:
        workbook.close()
