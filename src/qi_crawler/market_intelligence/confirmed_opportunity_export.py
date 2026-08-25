"""Source-neutral XLSX export for the latest Human-confirmed observations."""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from datetime import UTC
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..excel_safety import safe_excel_row
from .opportunity_contract import OpportunitySourceType
from .opportunity_radar import OpportunityRadarItem
from .opportunity_review import OpportunityReviewRecord, OpportunityReviewService
from .source_integrity import verify_source_integrity

BUSINESS_HEADERS = (
    "LOẠI NGUỒN",
    "NAMESPACE",
    "MÃ CƠ HỘI",
    "MÃ GỐC",
    "REVISION",
    "TÊN GÓI THẦU",
    "TÊN DỰ ÁN",
    "GIÁ GÓI THẦU GỐC",
    "GIÁ GÓI THẦU",
    "NGUỒN VỐN",
    "CHỦ ĐẦU TƯ",
    "BÊN MỜI THẦU",
    "NỘI DUNG PHÊ DUYỆT",
    "NỘI DUNG CHÍNH GÓI THẦU",
    "HÌNH THỨC LỰA CHỌN",
    "PHƯƠNG THỨC LỰA CHỌN",
    "ĐỊA ĐIỂM",
    "MÃ TỈNH/THÀNH PHỐ",
    "TỈNH/THÀNH PHỐ",
)

AUDIT_HEADERS = (
    "QUYẾT ĐỊNH REVIEW",
    "NGƯỜI REVIEW",
    "GHI CHÚ REVIEW",
    "ID SỰ KIỆN REVIEW",
    "THỜI ĐIỂM REVIEW",
    "OBSERVATION KEY",
    "FILE NGUỒN",
    "SHA-256 NGUỒN",
    "SHEET NGUỒN",
    "DÒNG NGUỒN",
    "SCHEMA NGUỒN",
)

DEFAULT_CONFIRMED_OPPORTUNITY_EXPORT_FILENAME = "CÁC GÓI ĐÃ XÁC NHẬN.xlsx"


class ConfirmedOpportunityExportError(ValueError):
    """Raised when confirmed observations cannot be exported safely."""


@dataclass(frozen=True, slots=True)
class ConfirmedOpportunityExportResult:
    """Summary of one source-neutral confirmed-opportunity export."""

    output: Path
    exported_rows: int
    source_type: OpportunitySourceType
    source_sha256: str


def export_confirmed_opportunity(
    review_service: OpportunityReviewService,
    items: Iterable[OpportunityRadarItem],
    *,
    source_type: OpportunitySourceType | str,
    source_path: str | Path,
    expected_source_sha256: str,
    output: str | Path,
) -> ConfirmedOpportunityExportResult:
    """Write only the latest CONFIRMED observations after source verification."""

    proof = verify_source_integrity(source_path, expected_source_sha256)
    normalized_source_type = OpportunitySourceType(source_type)
    universe = tuple(items)
    item_by_key = {item.observation_key: item for item in universe}
    records = review_service.current_confirmed(universe)
    confirmed: list[tuple[OpportunityRadarItem, OpportunityReviewRecord]] = []
    for record in records:
        item = item_by_key.get(record.identity.observation_key)
        if item is None:
            raise ConfirmedOpportunityExportError(
                "confirmed review refers to an observation outside the loaded source"
            )
        if item.source_type is not normalized_source_type:
            raise ConfirmedOpportunityExportError("confirmed source type does not match loaded source")
        if record.identity.source_type is not normalized_source_type:
            raise ConfirmedOpportunityExportError("review source type does not match loaded source")
        if (
            record.identity.source_sheet != item.sheet
            or record.identity.source_row != item.source_row
        ):
            raise ConfirmedOpportunityExportError("review provenance does not match loaded source")
        if item.source_sha256.casefold() != proof.actual_sha256.casefold():
            raise ConfirmedOpportunityExportError("confirmed observation source SHA does not match source")
        if record.identity.source_sha256.casefold() != proof.actual_sha256.casefold():
            raise ConfirmedOpportunityExportError("review source SHA does not match source")
        confirmed.append((item, record))
    confirmed.sort(key=lambda pair: _sort_key(*pair))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CONFIRMED"
    headers = BUSINESS_HEADERS + AUDIT_HEADERS
    sheet.append(headers)
    for item, record in confirmed:
        sheet.append(safe_excel_row(_row(item, record)))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}{max(sheet.max_row, 1)}"
    )
    destination = Path(output).resolve()
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return ConfirmedOpportunityExportResult(
        output=destination,
        exported_rows=len(confirmed),
        source_type=normalized_source_type,
        source_sha256=proof.actual_sha256,
    )


def _sort_key(item: OpportunityRadarItem, record: OpportunityReviewRecord) -> tuple[object, ...]:
    return (
        item.source_sha256.casefold(),
        item.sheet.casefold(),
        item.source_row,
        item.identity.raw_id.casefold(),
        record.event_id,
    )


def _row(item: OpportunityRadarItem, record: OpportunityReviewRecord) -> tuple[object, ...]:
    return (
        item.source_type.value,
        item.identity.namespace.value,
        item.identity.raw_id,
        item.identity.base_id,
        item.identity.revision,
        item.package_name,
        item.project,
        item.package_price_raw,
        item.package_price,
        item.funding_source,
        item.investor,
        item.procuring_entity,
        item.approval_content,
        item.package_main_content,
        item.selection_method,
        item.procurement_method,
        item.location_detail_raw,
        item.province_city_code,
        item.province_city_name,
        record.decision.value,
        record.reviewer,
        record.note,
        record.event_id,
        record.created_at.astimezone(UTC).isoformat(),
        item.observation_key,
        item.source_filename,
        item.source_sha256,
        item.sheet,
        item.source_row,
        item.schema_version,
    )


__all__ = [
    "AUDIT_HEADERS",
    "BUSINESS_HEADERS",
    "DEFAULT_CONFIRMED_OPPORTUNITY_EXPORT_FILENAME",
    "ConfirmedOpportunityExportError",
    "ConfirmedOpportunityExportResult",
    "export_confirmed_opportunity",
]
