"""Fail-closed source detection for Excel intake.

The detector is deliberately evidence based: filename is only a hint, while
headers and embedded PL/IB identifiers provide content evidence.  It never
converts one source schema into another.
"""

from __future__ import annotations

import hashlib
import re
import unicodedata
from dataclasses import dataclass
from enum import StrEnum
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class SourceType(StrEnum):
    KHMT = "KHMT"
    TBMT = "TBMT"
    UNKNOWN = "UNKNOWN"


@dataclass(frozen=True, slots=True)
class SourceTypeDetection:
    original_filename: str
    source_sha256: str
    filename_type: SourceType
    content_type: SourceType
    identity_namespace: str | None
    identity_values: tuple[str, ...]
    identity_raw_values: tuple[str, ...]
    auto_type: SourceType
    requires_human: bool
    evidence: tuple[str, ...]
    reasons: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class ResolvedSourceType:
    final_type: SourceType
    authority: str
    identity_namespace: str | None
    source_sha256: str


_IDENTITY_RE = re.compile(
    r"\b(?P<namespace>PL|IB)\s*(?P<number>\d{8,14})\s*-\s*(?P<revision>\d{2})\b",
    re.IGNORECASE,
)
_KHMT_SIGNATURE = frozenset(
    {
        "SỐ KẾ HOẠCH",
        "TÊN GÓI THẦU",
        "TÊN CHỦ ĐẦU TƯ",
    }
)
_TBMT_SIGNATURE = frozenset(
    {
        "BÊN MỜI THẦU",
        "GÓI THẦU",
        "DỰ ÁN",
        "GIÁ GÓI THẦU",
    }
)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _text(value: Any) -> str:
    return " ".join(unicodedata.normalize("NFKC", str(value)).split())


def _header_key(value: Any) -> str:
    return _text(value).upper()


def _filename_type(path: Path) -> SourceType:
    stem = path.stem.upper()
    if stem.startswith("KHMT"):
        return SourceType.KHMT
    if stem.startswith("TBMT"):
        return SourceType.TBMT
    return SourceType.UNKNOWN


def _content_type(headers: set[str]) -> SourceType:
    has_khmt = _KHMT_SIGNATURE.issubset(headers)
    has_tbmt = _TBMT_SIGNATURE.issubset(headers)
    if has_khmt and has_tbmt:
        return SourceType.UNKNOWN
    if has_khmt:
        return SourceType.KHMT
    if has_tbmt:
        return SourceType.TBMT
    return SourceType.UNKNOWN


def _identities(values: list[str]) -> tuple[tuple[str, ...], tuple[str, ...], str | None]:
    canonical: list[str] = []
    raw: list[str] = []
    namespaces: set[str] = set()
    for value in values:
        for match in _IDENTITY_RE.finditer(value):
            namespace = match.group("namespace").upper()
            raw_value = match.group(0).strip()
            normalized = f"{namespace}{match.group('number')}-{match.group('revision')}"
            namespaces.add(namespace)
            if normalized not in canonical:
                canonical.append(normalized)
            if raw_value not in raw:
                raw.append(raw_value)
    namespace = next(iter(namespaces)) if len(namespaces) == 1 else None
    return tuple(canonical), tuple(raw), namespace


def detect_source_type(path: Path) -> SourceTypeDetection:
    """Inspect a workbook without importing it into any MI authority."""

    path = Path(path).resolve()
    filename_type = _filename_type(path)
    reasons: list[str] = []
    evidence: list[str] = []
    headers: set[str] = set()
    cell_text: list[str] = []
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [_text(value) for value in row if value not in (None, "")]
                if not values:
                    continue
                if row_number <= 50:
                    headers.update(_header_key(value) for value in values)
                cell_text.extend(values)
                if len(cell_text) >= 5000:
                    break
                if len(cell_text) >= 5000:
                    break
    except (BadZipFile, InvalidFileException, KeyError, OSError, TypeError, ValueError) as exc:
        reasons.append(f"content inspection failed: {type(exc).__name__}")
    finally:
        if workbook is not None:
            workbook.close()

    content_type = _content_type(headers)
    if _KHMT_SIGNATURE.issubset(headers) and _TBMT_SIGNATURE.issubset(headers):
        reasons.append("both KHMT and TBMT schema signatures present")
    if _KHMT_SIGNATURE.issubset(headers):
        evidence.append("KHMT headers: SỐ KẾ HOẠCH + TÊN GÓI THẦU + TÊN CHỦ ĐẦU TƯ")
    if _TBMT_SIGNATURE.issubset(headers):
        evidence.append("TBMT headers: BÊN MỜI THẦU + GÓI THẦU + DỰ ÁN + GIÁ GÓI THẦU")
    identity_values, identity_raw_values, identity_namespace = _identities(cell_text)
    if identity_values:
        evidence.append(f"{identity_namespace or 'mixed'} identity in workbook content")
        namespaces = {value[:2] for value in identity_values}
        if len(namespaces) > 1:
            reasons.append("conflicting identity namespaces in workbook content")

    if filename_type is not SourceType.UNKNOWN:
        evidence.append(f"filename hint: {filename_type.value}")
    if filename_type is SourceType.UNKNOWN:
        reasons.append("filename has no trusted KHMT/TBMT prefix")
    if filename_type is not SourceType.UNKNOWN and content_type is not SourceType.UNKNOWN:
        if filename_type is not content_type:
            reasons.append("conflict between filename hint and workbook schema")
        elif filename_type is SourceType.KHMT and identity_namespace not in (None, "PL"):
            reasons.append("KHMT schema contains a non-PL identity")
        elif filename_type is SourceType.TBMT and identity_namespace not in (None, "IB"):
            reasons.append("TBMT schema contains a non-IB identity")
    elif filename_type is not SourceType.UNKNOWN and content_type is SourceType.UNKNOWN:
        reasons.append("filename hint is not confirmed by workbook schema")
    elif filename_type is SourceType.UNKNOWN and content_type is not SourceType.UNKNOWN:
        reasons.append("content suggestion requires human source selection")

    compatible = (
        filename_type is not SourceType.UNKNOWN
        and filename_type is content_type
        and not reasons[-1:].count("filename hint is not confirmed by workbook schema")
        and not any(reason.startswith("conflict") or "non-" in reason for reason in reasons)
    )
    auto_type = filename_type if compatible else SourceType.UNKNOWN
    return SourceTypeDetection(
        original_filename=path.name,
        source_sha256=_sha256(path),
        filename_type=filename_type,
        content_type=content_type,
        identity_namespace=identity_namespace,
        identity_values=identity_values,
        identity_raw_values=identity_raw_values,
        auto_type=auto_type,
        requires_human=auto_type is SourceType.UNKNOWN,
        evidence=tuple(evidence),
        reasons=tuple(reasons),
    )


def resolve_source_type(
    detection: SourceTypeDetection,
    selected: SourceType | None = None,
) -> ResolvedSourceType:
    """Resolve the UI choice; ``None`` means automatic classification."""

    final_type = detection.auto_type if selected is None else SourceType(selected)
    if final_type is SourceType.UNKNOWN:
        raise ValueError("Cần chọn rõ nguồn KHMT hoặc TBMT trước khi nhập.")
    return ResolvedSourceType(
        final_type=final_type,
        authority="AUTO" if selected is None else "HUMAN",
        identity_namespace=detection.identity_namespace,
        source_sha256=detection.source_sha256,
    )
