"""Bounded, fail-closed Excel importer for KHMT procurement-plan rows."""

from __future__ import annotations

import hashlib
import unicodedata
from dataclasses import dataclass
from datetime import UTC, datetime
from enum import StrEnum
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .khmt_contract import (
    OBSERVED_KHMT_HEADERS,
    KHMTImportBatch,
    PlanPackage,
    ProcurementPlan,
)
from .khmt_normalization import (
    compact_text,
    normalize_selection_method,
    parse_package_price,
    parse_plan_identity,
)
from .location_resolver import resolve_province_city

REQUIRED_KHMT_HEADERS = frozenset({"SỐ KẾ HOẠCH", "TÊN GÓI THẦU"})
MAX_HEADER_SCAN_ROWS = 50
SCHEMA_VERSION = "mi-1"


class KHMTIssueCode(StrEnum):
    MISSING_REQUIRED_HEADER = "MISSING_REQUIRED_HEADER"
    DUPLICATE_HEADER = "DUPLICATE_HEADER"
    INVALID_PLAN_ID = "INVALID_PLAN_ID"
    INVALID_PRICE = "INVALID_PRICE"
    EMPTY_PACKAGE_NAME = "EMPTY_PACKAGE_NAME"
    UNSUPPORTED_WORKBOOK = "UNSUPPORTED_WORKBOOK"
    NO_USABLE_SHEET = "NO_USABLE_SHEET"
    LOCATION_AMBIGUOUS = "LOCATION_AMBIGUOUS"


class KHMTImportError(ValueError):
    def __init__(self, code: KHMTIssueCode, message: str) -> None:
        self.code = code
        super().__init__(f"{code.value}: {message}")


@dataclass(frozen=True, slots=True)
class KHMTImportIssue:
    code: KHMTIssueCode
    message: str
    source_row: int | None = None
    source_field: str | None = None


@dataclass(frozen=True, slots=True)
class KHMTImportResult:
    batch: KHMTImportBatch
    packages: tuple[PlanPackage, ...]
    issues: tuple[KHMTImportIssue, ...]
    headers: tuple[str, ...]
    source_row_count: int


@dataclass(frozen=True, slots=True)
class _HeaderSelection:
    sheet: Worksheet
    row_number: int
    headers: tuple[str | None, ...]


def _header_text(value: Any) -> str | None:
    if value is None:
        return None
    text = " ".join(unicodedata.normalize("NFKC", str(value)).strip().split())
    return text or None


def _header_key(value: Any) -> str | None:
    text = _header_text(value)
    return text.upper() if text else None


def _canonical_headers(values: tuple[Any, ...]) -> tuple[str | None, ...]:
    known = {_header_key(header): header for header in OBSERVED_KHMT_HEADERS}
    result: list[str | None] = []
    seen: set[str] = set()
    for value in values:
        text = _header_text(value)
        if text is None:
            result.append(None)
            continue
        key = _header_key(text)
        assert key is not None
        if key in seen:
            raise KHMTImportError(
                KHMTIssueCode.DUPLICATE_HEADER,
                f"Duplicate source header: {text}",
            )
        seen.add(key)
        result.append(known.get(key, text))
    return tuple(result)


def _select_header(workbook: Any, requested_sheet: str | None) -> _HeaderSelection:
    if requested_sheet is not None:
        if requested_sheet not in workbook.sheetnames:
            raise KHMTImportError(
                KHMTIssueCode.NO_USABLE_SHEET,
                f"Requested sheet does not exist: {requested_sheet}",
            )
        sheets = [workbook[requested_sheet]]
    else:
        sheets = list(workbook.worksheets)

    best_partial: tuple[int, set[str]] | None = None
    for sheet in sheets:
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=MAX_HEADER_SCAN_ROWS, values_only=True),
            start=1,
        ):
            keys = {_header_key(value) for value in row if _header_key(value) is not None}
            overlap = keys.intersection(OBSERVED_KHMT_HEADERS)
            if best_partial is None or len(overlap) > len(best_partial[1]):
                best_partial = (row_number, overlap)
            if REQUIRED_KHMT_HEADERS.issubset(keys):
                return _HeaderSelection(
                    sheet=sheet,
                    row_number=row_number,
                    headers=_canonical_headers(tuple(row)),
                )

    if best_partial and best_partial[1]:
        missing = sorted(REQUIRED_KHMT_HEADERS - best_partial[1])
        raise KHMTImportError(
            KHMTIssueCode.MISSING_REQUIRED_HEADER,
            f"Missing required header(s): {', '.join(missing)}",
        )
    raise KHMTImportError(
        KHMTIssueCode.NO_USABLE_SHEET,
        "No sheet contains the required KHMT header contract",
    )


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _raw_row(headers: tuple[str | None, ...], values: tuple[Any, ...]) -> dict[str, Any]:
    return {
        header: values[index] if index < len(values) else None
        for index, header in enumerate(headers)
        if header is not None
    }


def import_khmt_workbook(
    path: Path,
    *,
    sheet_name: str | None = None,
    imported_at: datetime | None = None,
) -> KHMTImportResult:
    path = path.resolve()
    if path.suffix.lower() != ".xlsx" or not path.is_file():
        raise KHMTImportError(
            KHMTIssueCode.UNSUPPORTED_WORKBOOK,
            "KHMT importer requires an existing .xlsx workbook",
        )
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise KHMTImportError(
            KHMTIssueCode.UNSUPPORTED_WORKBOOK,
            f"Workbook could not be opened: {type(exc).__name__}",
        ) from exc

    try:
        selected = _select_header(workbook, sheet_name)
        batch = KHMTImportBatch(
            source_filename=path.name,
            source_sha256=_sha256(path),
            sheet=selected.sheet.title,
            imported_at=imported_at or datetime.now(UTC),
            schema_version=SCHEMA_VERSION,
        )
        packages: list[PlanPackage] = []
        issues: list[KHMTImportIssue] = []
        plans: dict[tuple[str, str | None], ProcurementPlan] = {}
        source_row_count = 0
        for row_number, values in enumerate(
            selected.sheet.iter_rows(
                min_row=selected.row_number + 1,
                values_only=True,
            ),
            start=selected.row_number + 1,
        ):
            raw_fields = _raw_row(selected.headers, tuple(values))
            if not any(value not in (None, "") for value in raw_fields.values()):
                continue
            source_row_count += 1
            identity = parse_plan_identity(raw_fields.get("SỐ KẾ HOẠCH"))
            if identity is None:
                issues.append(
                    KHMTImportIssue(
                        KHMTIssueCode.INVALID_PLAN_ID,
                        "Plan ID is missing or malformed",
                        source_row=row_number,
                        source_field="SỐ KẾ HOẠCH",
                    )
                )
                continue
            package_name = compact_text(raw_fields.get("TÊN GÓI THẦU"))
            if package_name is None:
                issues.append(
                    KHMTImportIssue(
                        KHMTIssueCode.EMPTY_PACKAGE_NAME,
                        "Package name is missing",
                        source_row=row_number,
                        source_field="TÊN GÓI THẦU",
                    )
                )
                continue

            price_source = raw_fields.get("GIÁ GÓI THẦU")
            price_raw = compact_text(price_source)
            package_price = parse_package_price(price_source)
            if price_raw is not None and package_price is None:
                issues.append(
                    KHMTImportIssue(
                        KHMTIssueCode.INVALID_PRICE,
                        "Package price is not an unambiguous integer amount",
                        source_row=row_number,
                        source_field="GIÁ GÓI THẦU",
                    )
                )

            location = resolve_province_city(raw_fields)
            if location.evidence.startswith("Conflicting"):
                issues.append(
                    KHMTImportIssue(
                        KHMTIssueCode.LOCATION_AMBIGUOUS,
                        location.evidence,
                        source_row=row_number,
                    )
                )

            plan_key = (identity.base_id, identity.revision)
            plan = plans.setdefault(
                plan_key,
                ProcurementPlan(
                    plan_id_raw=identity.raw,
                    plan_base_id=identity.base_id,
                    plan_revision=identity.revision,
                    import_batch=batch,
                ),
            )
            packages.append(
                PlanPackage(
                    plan=plan,
                    source_row=row_number,
                    package_name=package_name,
                    investor=compact_text(raw_fields.get("TÊN CHỦ ĐẦU TƯ")),
                    project=compact_text(raw_fields.get("TÊN DỰ ÁN")),
                    package_price_raw=price_raw,
                    package_price=package_price,
                    total_investment_raw=compact_text(raw_fields.get("TỔNG MỨC ĐẦU TƯ")),
                    approval_content_raw=compact_text(raw_fields.get("NỘI DUNG PHÊ DUYỆT")),
                    funding_source=compact_text(raw_fields.get("NGUỒN VỐN")),
                    selection_method_raw=compact_text(raw_fields.get("HÌNH THỨC LỰA CHỌN")),
                    selection_method=normalize_selection_method(
                        raw_fields.get("HÌNH THỨC LỰA CHỌN")
                    ),
                    selection_schedule_raw=compact_text(raw_fields.get("THỜI GIAN LỰA CHỌN")),
                    contract_type_raw=compact_text(raw_fields.get("HÌNH THỨC HỢP ĐỒNG")),
                    execution_duration_raw=compact_text(raw_fields.get("THỜI GIAN THỰC HIỆN")),
                    location_detail_raw=location.location_detail_raw,
                    province_city_code=location.code,
                    province_city_name=location.name,
                    province_city_status=location.status,
                    province_city_evidence=location.evidence,
                    raw_fields=raw_fields,
                    provenance={
                        "source_filename": batch.source_filename,
                        "source_sha256": batch.source_sha256,
                        "sheet": batch.sheet,
                        "source_row": row_number,
                    },
                    source_notice_id=None,
                )
            )
        return KHMTImportResult(
            batch=batch,
            packages=tuple(packages),
            issues=tuple(issues),
            headers=tuple(header for header in selected.headers if header is not None),
            source_row_count=source_row_count,
        )
    finally:
        workbook.close()
