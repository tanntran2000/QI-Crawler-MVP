from __future__ import annotations

import csv
import re
import unicodedata
from collections.abc import Iterable
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .crawler import CrawlerService
from .models import CrawlRun, Notice
from .parser import ParsedAttachment, ParsedNotice, parse_datetime_value, parse_money

COLUMN_ALIASES: dict[str, set[str]] = {
    "notice_code": {"notice_code", "ma_tbmt", "ma_thong_bao", "so_tbmt", "ma_goi_thau"},
    "source_notice_id": {"source_notice_id", "ma_nguon", "source_id", "website_notice_id"},
    "source_name": {"source_name", "nguon", "website", "source"},
    "title": {"title", "ten_goi_thau", "ten_du_an", "tieu_de"},
    "buyer": {"buyer", "ben_moi_thau", "don_vi_moi_thau"},
    "procuring_entity_address": {
        "procuring_entity_address",
        "dia_chi_ben_moi_thau",
        "dia_chi_cua_ben_moi_thau",
    },
    "investor": {"investor", "chu_dau_tu"},
    "project_name": {"project_name", "du_an", "ten_du_an"},
    "package_description": {
        "package_description",
        "noi_dung_chinh_cua_goi_thau",
        "mo_ta_goi_thau",
    },
    "package_price": {"package_price", "gia_goi_thau", "gia_du_toan", "gia_tri_goi_thau"},
    "currency": {"currency", "tien_te", "loai_tien"},
    "published_at": {"published_at", "ngay_dang_tai", "thoi_gian_dang_tai"},
    "closing_at": {"closing_at", "thoi_diem_dong_thau", "thoi_gian_dong_thau"},
    "location": {"location", "dia_diem", "dia_diem_thuc_hien"},
    "sector": {"sector", "linh_vuc", "phan_loai_linh_vuc"},
    "selection_method": {
        "selection_method",
        "phuong_thuc_lua_chon_nha_thau",
    },
    "selection_form": {"selection_form", "hinh_thuc_lua_chon_nha_thau"},
    "funding_source": {"funding_source", "nguon_von"},
    "document_issue_at": {
        "document_issue_at",
        "thoi_gian_phat_hanh_hsmt",
        "thoi_gian_phat_hanh_e_hsmt",
    },
    "document_price": {"document_price", "gia_ban_1_bo_hsmt", "gia_hsmt"},
    "bid_security_amount": {
        "bid_security_amount",
        "bao_dam_du_thau",
        "gia_tri_bao_dam_du_thau",
    },
    "bid_security_method": {
        "bid_security_method",
        "hinh_thuc_bao_dam_du_thau",
    },
    "issue_location": {"issue_location", "dia_diem_phat_hanh"},
    "bid_open_at": {"bid_open_at", "thoi_gian_mo_thau", "thoi_diem_mo_thau"},
    "contract_duration": {"contract_duration", "thoi_gian_thuc_hien_hop_dong"},
    "review_status": {"review_status", "trang_thai_duyet"},
    "notice_version": {"notice_version", "phien_ban", "lan_dang", "lan_thay_doi"},
    "source_url": {"source_url", "url", "duong_dan", "link"},
    "attachments": {"attachments", "tep_dinh_kem", "file_dinh_kem", "attachment_urls"},
}


@dataclass(slots=True)
class ImportSummary:
    rows_found: int = 0
    inserted: int = 0
    updated: int = 0
    unchanged: int = 0
    rejected: int = 0
    reject_file: Path | None = None
    errors: list[str] = field(default_factory=list)


def _fold_key(value: Any) -> str:
    raw = str(value or "").replace("\u0110", "D").replace("\u0111", "d")
    text = unicodedata.normalize("NFD", raw)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")


def _canonical_key(header: Any) -> str | None:
    folded = _fold_key(header)
    for canonical, aliases in COLUMN_ALIASES.items():
        if folded in aliases:
            return canonical
    return None


def _format_value(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip() or None


def _iter_csv(path: Path) -> Iterable[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(handle, dialect=dialect)
        yield from reader


def _iter_xlsx(path: Path) -> Iterable[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return
        for values in rows:
            yield dict(zip(headers, values, strict=False))
    finally:
        workbook.close()


def iter_rows(path: Path) -> Iterable[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        yield from _iter_csv(path)
    elif suffix == ".xlsx":
        yield from _iter_xlsx(path)
    else:
        raise ValueError("Chi ho tro import .csv hoac .xlsx")


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for header, value in row.items():
        key = _canonical_key(header)
        if key and key not in normalized:
            normalized[key] = value
    return normalized


def _attachments(value: Any, base_source: str) -> list[ParsedAttachment]:
    if value is None:
        return []
    items = re.split(r"[;\n|]+", str(value))
    result: list[ParsedAttachment] = []
    for item in items:
        url = item.strip()
        if not url:
            continue
        result.append(ParsedAttachment(source_url=url, file_name=None))
    return result


def import_file(service: CrawlerService, path: Path) -> ImportSummary:
    path = path.resolve()
    summary = ImportSummary()
    rejects: list[dict[str, Any]] = []

    with service.db.session() as session:
        run = CrawlRun(status="running", source_name=f"import:{path.name}")
        session.add(run)
        session.flush()
        run_id = run.id

    for row_number, raw_row in enumerate(iter_rows(path), start=2):
        summary.rows_found += 1
        row = _normalize_row(raw_row)
        source_url = _format_value(row.get("source_url")) or f"import://{path.as_posix()}#row={row_number}"
        price, detected_currency = parse_money(_format_value(row.get("package_price")))
        currency = _format_value(row.get("currency")) or detected_currency
        parsed = ParsedNotice(
            source_url=source_url,
            notice_code=_format_value(row.get("notice_code")),
            source_notice_id=_format_value(row.get("source_notice_id")),
            source_name=_format_value(row.get("source_name")),
            title=_format_value(row.get("title")),
            buyer=_format_value(row.get("buyer")),
            procuring_entity_address=_format_value(row.get("procuring_entity_address")),
            investor=_format_value(row.get("investor")),
            project_name=_format_value(row.get("project_name")),
            package_description=_format_value(row.get("package_description")),
            package_price=price,
            currency=currency,
            published_at=_format_value(row.get("published_at")),
            closing_at=_format_value(row.get("closing_at")),
            location=_format_value(row.get("location")),
            sector=_format_value(row.get("sector")),
            selection_method=_format_value(row.get("selection_method")),
            selection_form=_format_value(row.get("selection_form")),
            notice_version=_format_value(row.get("notice_version")),
            funding_source=_format_value(row.get("funding_source")),
            document_issue_at=parse_datetime_value(_format_value(row.get("document_issue_at"))),
            document_price=parse_money(_format_value(row.get("document_price")))[0],
            bid_security_amount=parse_money(
                _format_value(row.get("bid_security_amount"))
            )[0],
            bid_security_method=_format_value(row.get("bid_security_method")),
            issue_location=_format_value(row.get("issue_location")),
            published_at_dt=parse_datetime_value(_format_value(row.get("published_at"))),
            closing_at_dt=parse_datetime_value(_format_value(row.get("closing_at"))),
            bid_open_at=parse_datetime_value(_format_value(row.get("bid_open_at"))),
            contract_duration=_format_value(row.get("contract_duration")),
            raw_text="\n".join(
                f"{key}: {_format_value(value)}" for key, value in row.items() if value not in (None, "")
            ),
            attachments=_attachments(row.get("attachments"), source_url),
        )
        try:
            imported_notice, created, changed = service.upsert_parsed_notice(
                parsed,
                source_kind="import",
                strict_validation=True,
                crawl_run_id=run_id,
            )
            if _format_value(row.get("review_status")):
                with service.db.session() as session:
                    stored_notice = session.get(Notice, imported_notice.id)
                    if stored_notice is not None:
                        stored_notice.review_status = _format_value(row.get("review_status"))
            if created:
                summary.inserted += 1
            elif changed:
                summary.updated += 1
            else:
                summary.unchanged += 1
        except Exception as exc:  # noqa: BLE001 - reject one bad source row and continue the batch
            summary.rejected += 1
            message = f"Dong {row_number}: {exc}"
            summary.errors.append(message)
            rejects.append(
                {
                    "row_number": row_number,
                    "error": str(exc),
                    **{str(key): value for key, value in raw_row.items()},
                }
            )

    if rejects:
        service.config.storage.rejects_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
        reject_path = service.config.storage.rejects_dir / f"{path.stem}_{stamp}_rejects.csv"
        headers: list[str] = []
        for item in rejects:
            for key in item:
                if key not in headers:
                    headers.append(key)
        with reject_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rejects)
        summary.reject_file = reject_path

    with service.db.session() as session:
        run = session.get(CrawlRun, run_id)
        if run:
            run.finished_at = datetime.now(UTC)
            run.status = "completed" if summary.rejected == 0 else "partial"
            run.records_found = summary.rows_found
            run.records_inserted = summary.inserted
            run.records_updated = summary.updated
            run.records_failed = summary.rejected
            run.error_message = "\n".join(summary.errors[:20]) or None
    return summary
