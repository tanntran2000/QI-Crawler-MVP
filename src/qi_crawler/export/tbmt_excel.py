from __future__ import annotations

import logging
from copy import copy
from dataclasses import dataclass
from datetime import UTC, date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from .. import __version__
from ..db import Database
from ..excel_safety import safe_excel_value
from ..models import Notice
from .tbmt_formatter import fold_text
from .tbmt_mapper import TBMTExcelMapper
from .tbmt_schema import (
    DATA_START_ROW,
    DATETIME_COLUMNS,
    HEADER_ROW,
    META_SHEET_NAME,
    MONEY_COLUMNS,
    SCHEMA_VERSION,
    SHEET_NAME,
    TBMT_COLUMNS,
    NormalizedTenderRecord,
    TBMTExcelRow,
)
from .tbmt_validator import DataQuality, TBMTValidation, validate_tbmt_record

logger = logging.getLogger(__name__)

TEMPLATE_PATH = Path(__file__).resolve().parents[3] / "templates" / "TBMT_template_v1.xlsx"
DATETIME_NUMBER_FORMAT = 'hh" giờ "mm" ngày "dd/mm/yyyy'


@dataclass(slots=True)
class TBMTExportResult:
    output: Path
    reject_output: Path | None
    total_records: int
    exported_records: int
    warning_records: int
    rejected_records: int
    crawl_run_id: int | None


def _unique_output(path: Path) -> Path:
    if not path.exists():
        return path
    version = 2
    while True:
        candidate = path.with_name(f"{path.stem}_v{version}{path.suffix}")
        if not candidate.exists():
            return candidate
        version += 1


def _default_name(report_date: date) -> str:
    return f"TBMT_{report_date.day}_{report_date.month}_{report_date.year}.xlsx"


def _excel_datetime(value: object) -> object:
    if not isinstance(value, datetime):
        return value
    if value.tzinfo is not None:
        return value.astimezone().replace(tzinfo=None)
    return value


def _load_notices(db: Database) -> list[Notice]:
    with db.session() as session:
        notices = session.scalars(
            select(Notice)
            .options(selectinload(Notice.bid_results), selectinload(Notice.bid_openings))
            .order_by(Notice.id.asc())
        ).all()
        session.expunge_all()
        return list(notices)


def _record_date(record: NormalizedTenderRecord) -> date | None:
    return record.published_at.date() if record.published_at else None


def _matches_filters(
    record: NormalizedTenderRecord,
    *,
    on_date: date | None,
    from_date: date | None,
    to_date: date | None,
    status: str | None,
    keyword: str | None,
) -> bool:
    record_date = _record_date(record)
    if on_date is not None and record_date != on_date:
        return False
    if from_date is not None and (record_date is None or record_date < from_date):
        return False
    if to_date is not None and (record_date is None or record_date > to_date):
        return False
    if status and fold_text(record.review_status) != fold_text(status):
        return False
    if keyword:
        haystack = " ".join(
            item or ""
            for item in (
                record.notice_id,
                record.package_name,
                record.package_description,
                record.procuring_entity,
                record.project_name,
            )
        )
        terms = [term for term in fold_text(keyword).split() if term]
        if not all(term in fold_text(haystack) for term in terms):
            return False
    return True


def _copy_template_row_style(sheet, target_row: int, styles: list[object]) -> None:
    for column, style in enumerate(styles, start=1):
        cell = sheet.cell(target_row, column)
        cell._style = copy(style)
        cell.alignment = copy(cell.alignment)


def _write_main_sheet(
    workbook,
    rows: list[tuple[TBMTExcelRow, TBMTValidation]],
    *,
    highlight: bool,
) -> None:
    sheet = workbook[SHEET_NAME]
    if tuple(sheet.cell(HEADER_ROW, column).value for column in range(1, 19)) != TBMT_COLUMNS:
        raise ValueError(f"Template khong dung schema {SCHEMA_VERSION}")

    style_row = [copy(sheet.cell(DATA_START_ROW, column)._style) for column in range(1, 19)]
    if sheet.max_row >= DATA_START_ROW:
        sheet.delete_rows(DATA_START_ROW, sheet.max_row - HEADER_ROW)

    now = datetime.now().astimezone().replace(tzinfo=None)
    for row_number, (mapped, validation) in enumerate(rows, start=DATA_START_ROW):
        _copy_template_row_style(sheet, row_number, style_row)
        for column_number, column_name in enumerate(TBMT_COLUMNS, start=1):
            value = _excel_datetime(mapped.values[column_name])
            cell = sheet.cell(row_number, column_number, safe_excel_value(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_number in MONEY_COLUMNS:
            sheet.cell(row_number, column_number).number_format = "#,##0"
        for column_number in DATETIME_COLUMNS:
            sheet.cell(row_number, column_number).number_format = DATETIME_NUMBER_FORMAT

        source_cell = sheet.cell(row_number, 15)
        if mapped.record.source_url:
            source_cell.hyperlink = mapped.record.source_url
            source_cell.style = "Hyperlink"

        if highlight:
            if mapped.record.bid_close_at:
                closing = _excel_datetime(mapped.record.bid_close_at)
                if isinstance(closing, datetime):
                    if closing < now:
                        sheet.cell(row_number, 16).fill = PatternFill("solid", fgColor="F4CCCC")
                    elif closing <= now + timedelta(hours=24):
                        sheet.cell(row_number, 16).fill = PatternFill("solid", fgColor="FCE5CD")
            if mapped.record.package_price is None:
                sheet.cell(row_number, 8).fill = PatternFill("solid", fgColor="FFF2CC")
            if mapped.record.document_issue_at is None:
                sheet.cell(row_number, 11).fill = PatternFill("solid", fgColor="FFF2CC")
            if validation.status == DataQuality.WARNING:
                sheet.cell(row_number, 1).fill = PatternFill("solid", fgColor="FFF2CC")

    last_row = max(sheet.max_row, HEADER_ROW)
    sheet.freeze_panes = "A11"
    sheet.auto_filter.ref = f"A10:R{last_row}"


def _write_meta_sheet(
    workbook,
    rows: list[tuple[TBMTExcelRow, TBMTValidation]],
    *,
    generated_at: datetime,
    crawl_run_id: int | None,
) -> None:
    if META_SHEET_NAME in workbook.sheetnames:
        del workbook[META_SHEET_NAME]
    sheet = workbook.create_sheet(META_SHEET_NAME)
    summary = [
        ("Schema Version", SCHEMA_VERSION),
        ("Generated At", generated_at.isoformat()),
        ("Crawler Version", __version__),
        ("Crawl Run", crawl_run_id),
        ("Total Records", len(rows)),
        ("Reviewed Records", sum(row.record.review_status == "approved" for row, _ in rows)),
        ("Source", ", ".join(sorted({row.record.source_kind or "unknown" for row, _ in rows}))),
    ]
    for key, value in summary:
        sheet.append([key, safe_excel_value(value)])

    sheet.append([])
    headers = [
        "Database ID",
        "Notice ID",
        "Notice Version",
        "Source URL",
        "Content Hash",
        "Crawl Time",
        "Crawl Run",
        "Crawl Status",
        "Review Status",
        "Data Quality",
        "Quality Notes",
    ]
    sheet.append(headers)
    for mapped, validation in rows:
        record = mapped.record
        notes = "; ".join(validation.errors + validation.warnings)
        sheet.append(
            [
                record.database_id,
                record.notice_id,
                record.notice_version,
                safe_excel_value(record.source_url),
                record.content_hash,
                _excel_datetime(record.crawled_at),
                record.crawl_run_id,
                record.crawl_status,
                record.review_status,
                validation.status.value,
                safe_excel_value(notes),
            ]
        )
    for cell in sheet[9]:
        cell.font = Font(bold=True)
    sheet.sheet_state = "hidden"


def _write_rejects(
    path: Path,
    rejected: list[tuple[NormalizedTenderRecord, TBMTValidation]],
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path = _unique_output(path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rejected"
    headers = ["Database ID", "Notice ID", "Tên gói thầu", "Lý do", "Nguồn"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for record, validation in rejected:
        sheet.append(
            [
                record.database_id,
                record.notice_id,
                safe_excel_value(record.package_name),
                safe_excel_value("; ".join(validation.errors)),
                safe_excel_value(record.source_url),
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)
    return path


def export_tbmt(
    db: Database,
    *,
    report_dir: Path = Path("data/reports"),
    rejects_dir: Path = Path("data/rejects"),
    output: Path | None = None,
    on_date: date | None = None,
    from_date: date | None = None,
    to_date: date | None = None,
    status: str | None = None,
    keyword: str | None = None,
    highlight: bool = False,
    latest_run_only: bool = True,
    template_path: Path = TEMPLATE_PATH,
) -> TBMTExportResult:
    """Export the stable TBMT-1.0 workbook without mutating the source template."""
    if not template_path.exists():
        raise FileNotFoundError(f"Khong tim thay template TBMT: {template_path}")

    notices = _load_notices(db)
    mapper = TBMTExcelMapper()
    records = [mapper.normalize(notice) for notice in notices]
    run_ids = [record.crawl_run_id for record in records if record.crawl_run_id is not None]
    crawl_run_id = max(run_ids) if run_ids and latest_run_only else None
    if crawl_run_id is not None:
        records = [record for record in records if record.crawl_run_id == crawl_run_id]
    records = [
        record
        for record in records
        if _matches_filters(
            record,
            on_date=on_date,
            from_date=from_date,
            to_date=to_date,
            status=status,
            keyword=keyword,
        )
    ]

    accepted: list[tuple[TBMTExcelRow, TBMTValidation]] = []
    rejected: list[tuple[NormalizedTenderRecord, TBMTValidation]] = []
    for record in records:
        validation = validate_tbmt_record(record)
        if validation.status == DataQuality.INVALID:
            rejected.append((record, validation))
            continue
        if validation.status == DataQuality.WARNING:
            logger.warning(
                "TBMT %s co canh bao: %s",
                record.notice_id,
                "; ".join(validation.warnings),
            )
        accepted.append((mapper.map(record, len(accepted) + 1), validation))

    generated_at = datetime.now(UTC)
    report_date = on_date or generated_at.astimezone().date()
    destination = output or report_dir / _default_name(report_date)
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination = _unique_output(destination)

    workbook = load_workbook(template_path)
    _write_main_sheet(workbook, accepted, highlight=highlight)
    _write_meta_sheet(
        workbook,
        accepted,
        generated_at=generated_at,
        crawl_run_id=crawl_run_id,
    )
    workbook.save(destination)

    reject_output = None
    if rejected:
        reject_output = _write_rejects(
            rejects_dir / f"{destination.stem}_rejects.xlsx",
            rejected,
        )

    return TBMTExportResult(
        output=destination,
        reject_output=reject_output,
        total_records=len(records),
        exported_records=len(accepted),
        warning_records=sum(
            validation.status == DataQuality.WARNING for _, validation in accepted
        ),
        rejected_records=len(rejected),
        crawl_run_id=crawl_run_id,
    )
