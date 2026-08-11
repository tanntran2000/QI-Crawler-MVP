from __future__ import annotations

import smtplib
from collections.abc import Iterable
from datetime import UTC, date, datetime, timedelta
from email.message import EmailMessage
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from .config import AppConfig
from .datetime_utils import parse_datetime_utc
from .db import Database
from .excel_safety import safe_excel_row
from .models import Attachment, Notice

NOTICE_HEADERS = [
    "id",
    "notice_code",
    "title",
    "buyer",
    "investor",
    "package_price",
    "currency",
    "published_at",
    "closing_at",
    "location",
    "sector",
    "selection_method",
    "notice_version",
    "source_url",
    "attachment_count",
]


def _notice_row(notice: Notice) -> list[object]:
    return [
        notice.id,
        notice.notice_code,
        notice.title,
        notice.buyer,
        notice.investor,
        notice.package_price,
        notice.currency,
        notice.published_at,
        notice.closing_at,
        notice.location,
        notice.sector,
        notice.selection_method,
        notice.notice_version,
        notice.source_url,
        len(notice.attachments),
    ]


def _format_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for column in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), 60)


def _append_notices(sheet, notices: Iterable[Notice]) -> None:
    sheet.append(NOTICE_HEADERS)
    for notice in notices:
        sheet.append(safe_excel_row(_notice_row(notice)))
    _format_sheet(sheet)


def build_daily_report(
    db: Database,
    output: Path,
    report_date: date | None = None,
    days_ahead: int = 7,
) -> Path:
    report_date = report_date or datetime.now(UTC).date()
    closing_end = report_date + timedelta(days=days_ahead)
    with db.session() as session:
        notices = list(
            session.scalars(
                select(Notice)
                .options(selectinload(Notice.attachments))
                .order_by(Notice.id.desc())
            ).all()
        )
        failed_attachments = list(
            session.scalars(
                select(Attachment)
                .where(Attachment.download_status.in_(["failed", "manual_review"]))
                .order_by(Attachment.id.desc())
            ).all()
        )

    new_notices = [
        item
        for item in notices
        if item.first_seen_at and item.first_seen_at.date() == report_date
    ]
    closing_soon: list[Notice] = []
    for item in notices:
        parsed = parse_datetime_utc(item.closing_at)
        if parsed and report_date <= parsed.date() <= closing_end:
            closing_soon.append(item)
    quality_issues = [item for item in notices if item.data_quality_status != "valid"]

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Goi thau moi"
    _append_notices(sheet, new_notices)

    sheet = workbook.create_sheet("Sap dong thau")
    _append_notices(sheet, closing_soon)

    sheet = workbook.create_sheet("Tat ca goi thau")
    _append_notices(sheet, notices)

    sheet = workbook.create_sheet("Tep tai loi")
    sheet.append(
        [
            "attachment_id",
            "notice_id",
            "file_name",
            "source_url",
            "status",
            "attempts",
            "error",
            "last_attempt_at",
        ]
    )
    for item in failed_attachments:
        sheet.append(
            safe_excel_row([
                item.id,
                item.notice_id,
                item.file_name,
                item.source_url,
                item.download_status,
                item.download_attempts,
                item.download_error,
                item.last_attempt_at.isoformat() if item.last_attempt_at else None,
            ])
        )
    _format_sheet(sheet)

    sheet = workbook.create_sheet("Chat luong du lieu")
    _append_notices(sheet, quality_issues)

    workbook.save(output)
    return output


def send_report_email(
    config: AppConfig,
    report_path: Path,
    subject: str | None = None,
    body: str | None = None,
) -> None:
    settings = config.reporting
    missing = [
        name
        for name, value in {
            "smtp_host": settings.smtp_host,
            "email_from": settings.email_from,
            "email_to": settings.email_to,
        }.items()
        if not value
    ]
    if missing:
        raise ValueError(f"Thieu cau hinh email: {', '.join(missing)}")

    message = EmailMessage()
    message["Subject"] = subject or f"Bao cao goi thau ngay {datetime.now(UTC).date().isoformat()}"
    message["From"] = settings.email_from
    message["To"] = ", ".join(settings.email_to)
    message.set_content(
        body
        or "Bao cao tu dong tu he thong QI Crawler. Vui long xem tep Excel dinh kem."
    )
    content = report_path.read_bytes()
    message.add_attachment(
        content,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=report_path.name,
    )

    with smtplib.SMTP(settings.smtp_host, settings.smtp_port, timeout=30) as smtp:
        if settings.smtp_use_tls:
            smtp.starttls()
        if settings.smtp_username:
            if not settings.smtp_password:
                raise ValueError("Thieu mat khau SMTP; hay dat QI_CRAWLER_SMTP_PASSWORD trong .env")
            smtp.login(settings.smtp_username, settings.smtp_password)
        smtp.send_message(message)
