"""Auditable, immutable intake for manual and future web tender documents."""

from __future__ import annotations

import hashlib
import json
import logging
import mimetypes
import os
import re
import stat
import tempfile
import unicodedata
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path, PurePosixPath
from urllib.parse import urlsplit
from xml.etree import ElementTree

from openpyxl import load_workbook
from pypdf import PdfReader
from pypdf.errors import PdfReadError
from sqlalchemy import func, or_, select

from .db import Database
from .document_taxonomy import classify_document
from .models import Document, Notice

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = frozenset({".pdf", ".docx", ".xlsx", ".zip"})
DOCUMENT_TYPES = {
    ".pdf": "PDF",
    ".docx": "DOCX",
    ".xlsx": "XLSX",
    ".zip": "ZIP",
}
MIME_TYPES = {
    ".pdf": "application/pdf",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".zip": "application/zip",
}
_NOTICE_ID_PATTERN = re.compile(r"\bIB\d{6,}(?:-\d{2,3})?\b", re.IGNORECASE)
_CONTENT_SCAN_MAX_PAGES = 3
_CONTENT_SCAN_MAX_ROWS = 100
WINDOWS_RESERVED_NAMES = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{index}" for index in range(1, 10)),
    *(f"LPT{index}" for index in range(1, 10)),
}


class DocumentIntakeError(RuntimeError):
    """Base class for a fail-closed document intake error."""


class DocumentValidationError(DocumentIntakeError):
    """The selected input is unsupported, empty, unreadable or unsafe."""


class DocumentStorageError(DocumentIntakeError):
    """The original could not be stored without risking overwrite or loss."""


class DocumentIdentityMismatch(DocumentValidationError):
    """Trusted identity evidence points at a different tender."""

    def __init__(self, expected: str, detected: str):
        self.expected = expected
        self.detected = detected
        super().__init__(
            f"CRITICAL_MISMATCH: expected tender {expected}; detected {detected}."
        )


@dataclass(frozen=True)
class BundleMembershipClaim:
    """Explicit non-filename evidence for a document without a content identifier."""

    kind: str
    evidence_locator: str
    base_notice_id: str | None = None
    revision: str | None = None
    reference_document_id: int | None = None
    confirmed_by: str | None = None


@dataclass(frozen=True)
class DocumentIntakeResult:
    outcome: str
    document_id: int
    original_filename: str
    stored_path: Path
    document_type: str
    mime_type: str
    file_size: int
    sha256: str
    version: int
    tender_id: int | None
    tender_identifier: str | None
    discovered_files: tuple[str, ...] = ()
    identity_status: str = "UNLINKED"
    expected_identity: str | None = None
    detected_identity: str | None = None
    file_format: str | None = None
    template_code: str | None = None
    package_type: str | None = None
    selection_method: str | None = None
    classification_status: str = "UNKNOWN"
    raw_notice_id: str | None = None
    base_notice_id: str | None = None
    notice_revision: str | None = None
    identity_source: str | None = None
    identity_evidence_locator: str | None = None
    identity_match_status: str | None = None
    identity_candidates: tuple[str, ...] = ()
    bundle_base_notice_id: str | None = None
    bundle_revision: str | None = None
    bundle_membership_status: str | None = None
    bundle_membership_source: str | None = None
    bundle_membership_evidence: str | None = None


@dataclass(frozen=True)
class DocumentBatchResult:
    results: tuple[DocumentIntakeResult, ...]
    extraction_warnings: tuple[str, ...] = ()

    @property
    def imported(self) -> int:
        return sum(item.outcome == "IMPORTED" for item in self.results)

    @property
    def duplicates(self) -> int:
        return sum(item.outcome == "DUPLICATE" for item in self.results)


@dataclass(frozen=True)
class DocumentManifestEntry:
    document_id: int
    document_type: str
    file_format: str | None
    template_code: str | None
    package_type: str | None
    selection_method: str | None
    classification_status: str
    filename: str
    sha256: str
    version: int
    source: str
    status: str
    stored_path: Path
    uploaded_at: datetime
    raw_notice_id: str | None = None
    base_notice_id: str | None = None
    notice_revision: str | None = None
    identity_match_status: str | None = None
    bundle_base_notice_id: str | None = None
    bundle_revision: str | None = None
    bundle_membership_status: str | None = None


@dataclass(frozen=True)
class TenderDocumentManifest:
    tender_id: int
    tender_identifier: str
    tender_title: str
    source: str
    identity_status: str
    documents: tuple[DocumentManifestEntry, ...]


@dataclass(frozen=True)
class TenderDocumentTarget:
    """Minimal verified tender context exposed to document acquisition services."""

    tender_id: int
    identifier: str
    source_name: str
    source_url: str


@dataclass(frozen=True)
class DocumentContentIdentity:
    """Raw tender identifiers found in native document content, never filenames."""

    raw_notice_id: str | None = None
    base_notice_id: str | None = None
    revision: str | None = None
    identity_source: str | None = None
    evidence_locator: str | None = None
    candidates: tuple[str, ...] = ()
    status: str = "NO_CONTENT_ID"


@dataclass(frozen=True)
class _IdentityResolution:
    status: str
    tender: Notice | None
    expected: str | None
    detected: str | None
    content_identity: DocumentContentIdentity | None = None
    match_status: str | None = None


@dataclass(frozen=True)
class _BundleMembership:
    """Logical bundle placement, distinct from the document's own identity evidence."""

    base_notice_id: str | None
    revision: str | None
    status: str
    source: str | None = None
    evidence_locator: str | None = None


def sanitize_filename(filename: str) -> str:
    """Return a Windows-safe leaf name while retaining the supported extension."""
    leaf = Path(filename.replace("\\", "/")).name
    normalized = unicodedata.normalize("NFKC", leaf)
    suffix = Path(normalized).suffix.lower()
    stem = normalized[: -len(suffix)] if suffix else normalized
    stem = re.sub(r"[^\w .()\-]+", "_", stem, flags=re.UNICODE)
    stem = re.sub(r"\s+", " ", stem).strip(" .")
    if not stem:
        stem = "document"
    if stem.upper() in WINDOWS_RESERVED_NAMES:
        stem = f"_{stem}"
    return f"{stem[:120]}{suffix}"


def _safe_scope(value: str) -> str:
    safe = sanitize_filename(value).rsplit(".", 1)[0]
    return safe or "unlinked"


def extract_document_identity(path: Path) -> DocumentContentIdentity:
    """Read early native content only; no filename-derived identity is accepted."""
    backend = _identity_backend(path)
    try:
        regions = _identity_content_regions(path)
    except (OSError, PdfReadError, ValueError, zipfile.BadZipFile, ElementTree.ParseError) as exc:
        logger.info(
            "IDENTITY_PRECHECK backend=%s file=%s pages_scanned=0 text_chars=0 "
            "candidates=() read_error=%s: %s",
            backend,
            path.name,
            type(exc).__name__,
            str(exc)[:200],
        )
        return DocumentContentIdentity()
    except Exception as exc:  # noqa: BLE001 - fail closed at the native-reader boundary.
        logger.warning(
            "IDENTITY_EXTRACTION_FAILED backend=%s file=%s exception=%s: %s",
            backend,
            path.name,
            type(exc).__name__,
            str(exc)[:200],
        )
        return DocumentContentIdentity(
            identity_source="DOCUMENT_CONTENT",
            status="EXTRACTION_FAILED",
        )
    candidates: list[tuple[str, str]] = []
    for locator, text in regions:
        for raw_value in _NOTICE_ID_PATTERN.findall(text):
            raw = raw_value.upper()
            if raw not in {value for value, _locator in candidates}:
                candidates.append((raw, locator))
    logger.info(
        "IDENTITY_PRECHECK backend=%s file=%s pages_scanned=%s text_chars=%s candidates=%s",
        backend,
        path.name,
        len(regions),
        sum(len(text) for _locator, text in regions),
        tuple(raw for raw, _locator in candidates),
    )
    if not candidates:
        return DocumentContentIdentity()
    bases = {_notice_id_parts(raw)[0] for raw, _locator in candidates}
    raw_values = tuple(raw for raw, _locator in candidates)
    if len(bases) != 1:
        return DocumentContentIdentity(
            identity_source="DOCUMENT_CONTENT",
            evidence_locator="; ".join(locator for _raw, locator in candidates),
            candidates=raw_values,
            status="AMBIGUOUS",
        )
    raw, locator = candidates[0]
    base, revision = _notice_id_parts(raw)
    return DocumentContentIdentity(
        raw_notice_id=raw,
        base_notice_id=base,
        revision=revision,
        identity_source="DOCUMENT_CONTENT",
        evidence_locator=locator,
        candidates=raw_values,
        status="FOUND",
    )


def _identity_backend(path: Path) -> str:
    return {
        ".pdf": "pypdf",
        ".docx": "zipfile_xml",
        ".xlsx": "openpyxl",
    }.get(path.suffix.lower(), "unsupported")


def _notice_id_parts(value: str) -> tuple[str, str | None]:
    normalized = value.strip().upper()
    base, separator, revision = normalized.rpartition("-")
    if separator and revision.isdigit():
        return base, revision
    return normalized, None


def _identity_content_regions(path: Path) -> list[tuple[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        reader = PdfReader(str(path))
        return [
            (f"file:{path.name}:page:{number}", page.extract_text() or "")
            for number, page in enumerate(reader.pages[:_CONTENT_SCAN_MAX_PAGES], start=1)
        ]
    if suffix == ".docx":
        with zipfile.ZipFile(path) as archive:
            root = ElementTree.fromstring(archive.read("word/document.xml"))
        text = " ".join(node.text or "" for node in root.iter() if node.text)
        return [(f"file:{path.name}:word/document.xml", text)]
    if suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            regions: list[tuple[str, str]] = []
            for worksheet in workbook.worksheets[:_CONTENT_SCAN_MAX_PAGES]:
                for row_number, row in enumerate(
                    worksheet.iter_rows(values_only=True), start=1
                ):
                    text = " ".join(str(value) for value in row if value is not None)
                    if text:
                        regions.append(
                            (f"file:{path.name}:sheet:{worksheet.title}!{row_number}", text)
                        )
                    if row_number >= _CONTENT_SCAN_MAX_ROWS:
                        break
            return regions
        finally:
            workbook.close()
    return []


class DocumentIntakeService:
    """Store one immutable original and create one auditable database record."""

    def __init__(self, database: Database, document_root: Path):
        self.database = database
        self.database.require_current_schema()
        self.document_root = document_root.expanduser().resolve()

    def resolve_tender_target(self, tender_reference: str) -> TenderDocumentTarget:
        """Resolve one unique tender before any web document acquisition starts."""
        tender = self._lookup_tender(tender_reference)
        if tender is None:
            raise DocumentValidationError(
                "Không tìm thấy tender đã xác định; không tải hoặc tự đoán tài liệu từ web."
            )
        return TenderDocumentTarget(
            tender_id=tender.id,
            identifier=self._notice_identifier(tender),
            source_name=tender.source_name or tender.source_kind or "web",
            source_url=tender.source_url,
        )

    def verify_tender_identity(
        self,
        tender_reference: str,
        *,
        detected_tender_reference: str | None,
        source_url: str | None,
    ) -> str:
        """Run the existing Identity Guard before a staged web file is hashed."""
        return self._resolve_identity(
            tender_reference,
            detected_tender_reference=detected_tender_reference,
            source_url=source_url,
        ).status

    def intake_path(
        self,
        input_path: Path,
        *,
        tender_reference: str | None = None,
        document_name: str | None = None,
        document_source: str = "manual_upload",
        source_url: str | None = None,
        uploaded_by: str | None = None,
        detected_tender_reference: str | None = None,
        bundle_claim: BundleMembershipClaim | None = None,
    ) -> DocumentBatchResult:
        candidate = input_path.expanduser()
        if candidate.is_symlink():
            raise DocumentValidationError("Không chấp nhận đường dẫn liên kết (symlink).")
        if candidate.is_dir():
            supported = sorted(
                path
                for path in candidate.rglob("*")
                if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS
            )
            if not supported:
                raise DocumentValidationError("Thư mục không có PDF, DOCX, XLSX hoặc ZIP.")
            results = tuple(
                self.intake_file(
                    path,
                    tender_reference=tender_reference,
                    document_name=document_name if len(supported) == 1 else None,
                    document_source=document_source,
                    source_url=source_url,
                    uploaded_by=uploaded_by,
                    detected_tender_reference=detected_tender_reference,
                    bundle_claim=bundle_claim,
                )
                for path in supported
            )
            return DocumentBatchResult(results)
        return DocumentBatchResult(
            (
                self.intake_file(
                    candidate,
                    tender_reference=tender_reference,
                    document_name=document_name,
                    document_source=document_source,
                    source_url=source_url,
                    uploaded_by=uploaded_by,
                    detected_tender_reference=detected_tender_reference,
                    bundle_claim=bundle_claim,
                ),
            )
        )

    def intake_file(
        self,
        input_path: Path,
        *,
        tender_reference: str | None = None,
        document_name: str | None = None,
        document_source: str = "manual_upload",
        source_url: str | None = None,
        uploaded_by: str | None = None,
        detected_tender_reference: str | None = None,
        bundle_claim: BundleMembershipClaim | None = None,
    ) -> DocumentIntakeResult:
        logger.info(
            "DOCUMENT_INTAKE_START source=%s filename=%s",
            document_source,
            input_path.name,
        )
        path = self._validate_input(input_path)
        if document_source not in {"manual_upload", "web"}:
            raise DocumentValidationError("Nguồn tài liệu không được hỗ trợ.")
        if document_source == "web" and not source_url:
            raise DocumentValidationError("Tài liệu từ web phải có source_url.")

        sha256, file_size = self._hash_file(path)
        logger.info("HASH_DONE sha256=%s file_size=%s", sha256, file_size)
        discovered_files = self._inspect_zip(path) if path.suffix.lower() == ".zip" else ()
        content_identity = extract_document_identity(path)
        identity = self._resolve_identity(
            tender_reference,
            detected_tender_reference=detected_tender_reference,
            source_url=source_url,
            content_identity=content_identity,
        )
        membership = self._resolve_bundle_membership(
            identity,
            document_source=document_source,
            source_url=source_url,
            claim=bundle_claim,
        )
        duplicate = self._find_duplicate(sha256)
        logger.info("DUPLICATE_CHECK_DONE duplicate=%s", duplicate is not None)
        if duplicate is not None:
            duplicate = self._guard_duplicate_identity(duplicate, identity, membership)
            stored_path = Path(duplicate.stored_path)
            if not stored_path.is_file():
                raise DocumentStorageError(
                    "Bản ghi trùng tồn tại nhưng file gốc không còn trong Document Store."
                )
            result = self._result(
                duplicate,
                "DUPLICATE",
                self._tender_identifier(duplicate.tender_id),
                identity_status=self._document_identity_status(duplicate),
                expected_identity=identity.expected,
                detected_identity=identity.detected,
            )
            logger.info(
                "DOCUMENT_INTAKE_DONE outcome=DUPLICATE document_id=%s",
                duplicate.id,
            )
            return result

        tender = identity.tender
        classification = classify_document(
            metadata_title=(document_name or "").strip() or None,
            filename=path.name,
            identity_status=identity.status,
            package_type=tender.contract_type if tender else None,
            selection_method=tender.selection_method if tender else None,
        )
        version = self._next_version(tender.id if tender else None)
        safe_filename = sanitize_filename(path.name)
        scope = self._tender_scope(tender)
        source_scope = self._source_scope(tender, document_source)
        destination = self.document_root / source_scope / scope / sha256 / safe_filename
        stored_path, created = self._atomic_store(path, destination)
        logger.info("FILE_STORED sha256=%s stored_path=%s", sha256, stored_path)
        mime_type = MIME_TYPES.get(path.suffix.lower()) or mimetypes.guess_type(path.name)[0]
        try:
            document = self._create_record(
                tender_id=tender.id if tender else None,
                document_source=document_source,
                document_type=classification.document_type.value,
                file_format=DOCUMENT_TYPES[path.suffix.lower()],
                template_code=classification.template_code,
                package_type=classification.package_type,
                selection_method=classification.selection_method,
                classification_status=classification.status.value,
                display_name=(document_name or "").strip() or None,
                original_filename=path.name,
                stored_path=stored_path,
                mime_type=mime_type or "application/octet-stream",
                file_size=file_size,
                sha256=sha256,
                version=version,
                source_url=source_url,
                uploaded_by=(uploaded_by or "").strip() or None,
                zip_supported_entries=discovered_files,
                status=identity.status,
                content_identity=identity.content_identity or DocumentContentIdentity(),
                identity_match_status=identity.match_status,
                bundle_membership=membership,
            )
        except Exception:
            if created:
                self._cleanup_orphan(stored_path)
            raise
        logger.info(
            "DOCUMENT_RECORD_CREATED document_id=%s tender_id=%s version=%s",
            document.id,
            document.tender_id,
            document.version,
        )
        result = self._result(
            document,
            "IMPORTED",
            self._notice_identifier(tender) if tender else None,
            identity_status=identity.status,
            expected_identity=identity.expected,
            detected_identity=identity.detected,
        )
        logger.info(
            "DOCUMENT_INTAKE_DONE outcome=IMPORTED document_id=%s",
            document.id,
        )
        return result

    @staticmethod
    def _validate_input(input_path: Path) -> Path:
        try:
            if input_path.is_symlink():
                raise DocumentValidationError(
                    "Không chấp nhận đường dẫn liên kết (symlink)."
                )
            path = input_path.expanduser().resolve(strict=True)
            if not path.is_file():
                raise DocumentValidationError("Đường dẫn không phải là một file.")
            extension = path.suffix.lower()
            if extension not in SUPPORTED_EXTENSIONS:
                raise DocumentValidationError("Chỉ hỗ trợ PDF, DOCX, XLSX và ZIP.")
            if path.stat().st_size <= 0:
                raise DocumentValidationError("Không thể nhập file rỗng.")
            with path.open("rb") as stream:
                stream.read(1)
            return path
        except DocumentIntakeError:
            raise
        except (OSError, PermissionError) as exc:
            raise DocumentValidationError("Không thể đọc file đã chọn.") from exc

    @staticmethod
    def _hash_file(path: Path) -> tuple[str, int]:
        digest = hashlib.sha256()
        size = 0
        try:
            with path.open("rb") as stream:
                for chunk in iter(lambda: stream.read(1024 * 1024), b""):
                    digest.update(chunk)
                    size += len(chunk)
        except OSError as exc:
            raise DocumentValidationError("Không thể đọc file để tính SHA-256.") from exc
        return digest.hexdigest(), size

    @staticmethod
    def _inspect_zip(path: Path) -> tuple[str, ...]:
        supported: list[str] = []
        try:
            with zipfile.ZipFile(path) as archive:
                for entry in archive.infolist():
                    normalized = entry.filename.replace("\\", "/")
                    pure_path = PurePosixPath(normalized)
                    mode = entry.external_attr >> 16
                    if (
                        pure_path.is_absolute()
                        or ".." in pure_path.parts
                        or (pure_path.parts and ":" in pure_path.parts[0])
                        or stat.S_ISLNK(mode)
                    ):
                        raise DocumentValidationError(
                            f"ZIP chứa đường dẫn không an toàn: {entry.filename}"
                        )
                    if not entry.is_dir() and pure_path.suffix.lower() in SUPPORTED_EXTENSIONS:
                        supported.append(normalized)
        except DocumentIntakeError:
            raise
        except (OSError, zipfile.BadZipFile, zipfile.LargeZipFile) as exc:
            raise DocumentValidationError("File ZIP không hợp lệ hoặc không thể đọc.") from exc
        return tuple(dict.fromkeys(supported))

    def _find_duplicate(self, sha256: str) -> Document | None:
        with self.database.session() as session:
            return session.scalar(select(Document).where(Document.sha256 == sha256))

    def _lookup_tender(self, reference: str | None) -> Notice | None:
        value = (reference or "").strip()
        if not value:
            return None
        predicates = [Notice.notice_code == value, Notice.source_notice_id == value]
        if value.isdigit():
            predicates.append(Notice.id == int(value))
        with self.database.session() as session:
            matches = tuple(
                session.scalars(select(Notice).where(or_(*predicates)).limit(2))
            )
        if len(matches) > 1:
            raise DocumentValidationError(
                "Mã tender không duy nhất trong database; không tự động lựa chọn."
            )
        return matches[0] if matches else None

    def _lookup_tender_by_source_url(self, source_url: str | None) -> Notice | None:
        value = (source_url or "").strip()
        if not value:
            return None
        with self.database.session() as session:
            matches = tuple(
                session.scalars(
                    select(Notice).where(Notice.source_url == value).limit(2)
                )
            )
        if len(matches) > 1:
            raise DocumentValidationError(
                "URL nguồn khớp nhiều tender; không tự động lựa chọn."
            )
        return matches[0] if matches else None

    def _resolve_identity(
        self,
        tender_reference: str | None,
        *,
        detected_tender_reference: str | None,
        source_url: str | None,
        content_identity: DocumentContentIdentity | None = None,
    ) -> _IdentityResolution:
        content_identity = content_identity or DocumentContentIdentity()
        expected_value = (tender_reference or "").strip() or None
        detected_value = (detected_tender_reference or "").strip() or None
        logger.info(
            "DOCUMENT_IDENTITY_CHECK expected=%s detected=%s source_url_present=%s",
            expected_value,
            detected_value,
            bool(source_url),
        )
        expected = self._lookup_tender(expected_value)
        detected = self._lookup_tender(detected_value)
        if detected_value is None:
            detected = self._lookup_tender_by_source_url(source_url)
            if detected is not None:
                detected_value = self._notice_identifier(detected)

        unknown_expected = expected_value is not None and expected is None
        unknown_detected = detected_value is not None and detected is None
        if unknown_expected or unknown_detected:
            logger.info(
                "DOCUMENT_IDENTITY_UNLINKED status=NEEDS_REVIEW expected=%s detected=%s",
                expected_value,
                detected_value,
            )
            return _IdentityResolution(
                status="NEEDS_REVIEW",
                tender=None,
                expected=expected_value,
                detected=detected_value,
            )

        if expected is not None and detected is not None and expected.id != detected.id:
            expected_identifier = self._notice_identifier(expected)
            detected_identifier = self._notice_identifier(detected)
            logger.error(
                "DOCUMENT_IDENTITY_MISMATCH expected=%s detected=%s",
                expected_identifier,
                detected_identifier,
            )
            raise DocumentIdentityMismatch(expected_identifier, detected_identifier)

        tender = expected or detected
        if content_identity.status == "EXTRACTION_FAILED":
            logger.warning(
                "DOCUMENT_IDENTITY_UNLINKED status=NEEDS_REVIEW reason=EXTRACTION_FAILED"
            )
            return _IdentityResolution(
                status="NEEDS_REVIEW",
                tender=tender,
                expected=expected_value,
                detected=detected_value,
                content_identity=content_identity,
                match_status="EXTRACTION_FAILED",
            )
        if tender is None:
            logger.info("DOCUMENT_IDENTITY_UNLINKED status=UNLINKED")
            return _IdentityResolution(
                status="UNLINKED",
                tender=None,
                expected=expected_value,
                detected=detected_value,
            )

        identifier = self._notice_identifier(tender)
        if content_identity.status == "AMBIGUOUS":
            logger.warning(
                "DOCUMENT_IDENTITY_UNLINKED status=AMBIGUOUS expected=%s candidates=%s",
                identifier,
                content_identity.candidates,
            )
            return _IdentityResolution(
                status="NEEDS_REVIEW",
                tender=tender,
                expected=expected_value or identifier,
                detected=detected_value,
                content_identity=content_identity,
                match_status="AMBIGUOUS",
            )
        if content_identity.status == "FOUND":
            expected_base, expected_revision = _notice_id_parts(identifier)
            if expected_base != content_identity.base_notice_id:
                logger.error(
                    "DOCUMENT_IDENTITY_MISMATCH expected=%s detected=%s source=DOCUMENT_CONTENT",
                    identifier,
                    content_identity.raw_notice_id,
                )
                raise DocumentIdentityMismatch(identifier, content_identity.raw_notice_id or "UNKNOWN")
            match_status = (
                "SAME_TENDER"
                if expected_revision == content_identity.revision
                else "SAME_TENDER_DIFFERENT_REVISION"
            )
            if expected_revision is None and not self._has_other_document_revision(
                tender.id,
                content_identity.revision,
            ):
                match_status = "SAME_TENDER"
            logger.info(
                "DOCUMENT_IDENTITY_VERIFIED tender=%s detected=%s match=%s",
                identifier,
                content_identity.raw_notice_id,
                match_status,
            )
            return _IdentityResolution(
                status="DOCUMENT_VERIFIED",
                tender=tender,
                expected=expected_value or identifier,
                detected=content_identity.raw_notice_id,
                content_identity=content_identity,
                match_status=match_status,
            )
        identity_status = (
            "HUMAN_DECLARED"
            if tender.source_origin == "MANUAL_TEAM_BID"
            else "VERIFIED_LINKED"
        )
        logger.info("DOCUMENT_IDENTITY_VERIFIED tender=%s status=%s", identifier, identity_status)
        return _IdentityResolution(
            status=identity_status,
            tender=tender,
            expected=expected_value or identifier,
            detected=detected_value,
            content_identity=content_identity,
            match_status="NO_CONTENT_ID",
        )

    def _resolve_bundle_membership(
        self,
        identity: _IdentityResolution,
        *,
        document_source: str,
        source_url: str | None,
        claim: BundleMembershipClaim | None,
    ) -> _BundleMembership:
        """Place a document in a revision bundle without weakening Identity Guard."""
        tender = identity.tender
        content = identity.content_identity or DocumentContentIdentity()
        if tender is None:
            return _BundleMembership(None, None, "NEEDS_REVIEW")

        expected_base, expected_revision = _notice_id_parts(
            self._notice_identifier(tender)
        )
        if content.status == "FOUND":
            # _resolve_identity already rejects a different content-derived base.
            assert content.base_notice_id == expected_base
            status = (
                "EXACT_BUNDLE"
                if expected_revision in {None, content.revision}
                else "SAME_TENDER_DIFFERENT_REVISION"
            )
            return _BundleMembership(
                content.base_notice_id,
                content.revision,
                status,
                "DOCUMENT_CONTENT",
                content.evidence_locator,
            )

        if content.status in {"AMBIGUOUS", "EXTRACTION_FAILED"}:
            return _BundleMembership(expected_base, expected_revision, "NEEDS_REVIEW")
        if claim is None:
            return _BundleMembership(expected_base, expected_revision, "NEEDS_REVIEW")

        kind = claim.kind.strip().upper()
        evidence = claim.evidence_locator.strip()
        if kind not in {"REFERENCE_LINKED", "PROVENANCE_LINKED", "HUMAN_LINKED"}:
            raise DocumentValidationError("Loại bằng chứng bundle không hợp lệ.")
        if not evidence:
            raise DocumentValidationError("Liên kết bundle phải có bằng chứng cụ thể.")

        if kind == "REFERENCE_LINKED":
            if claim.reference_document_id is None:
                raise DocumentValidationError("REFERENCE_LINKED phải chỉ rõ tài liệu HSMT tham chiếu.")
            with self.database.session() as session:
                reference = session.get(Document, claim.reference_document_id)
            if (
                reference is None
                or reference.tender_id != tender.id
                or reference.document_type != "E_HSMT"
                or not reference.bundle_base_notice_id
                or not reference.bundle_revision
            ):
                raise DocumentValidationError(
                    "Tài liệu HSMT tham chiếu không thuộc đúng bundle đang làm việc."
                )
            return _BundleMembership(
                reference.bundle_base_notice_id,
                reference.bundle_revision,
                kind,
                "PRIMARY_HSMT_REFERENCE",
                evidence,
            )

        claim_base = (claim.base_notice_id or "").strip().upper()
        claim_revision = (claim.revision or "").strip()
        if claim_base != expected_base or not claim_revision:
            raise DocumentValidationError(
                "Liên kết bundle không chứng minh được mã gói và revision đang làm việc."
            )
        if expected_revision is not None and claim_revision != expected_revision:
            raise DocumentValidationError(
                "Liên kết bundle chỉ hợp lệ cho đúng revision của gói đang làm việc."
            )
        if kind == "PROVENANCE_LINKED":
            if (
                document_source != "web"
                or not source_url
                or not tender.source_url
                or not self._same_origin(tender.source_url, source_url)
            ):
                raise DocumentValidationError(
                    "PROVENANCE_LINKED phải thuộc cùng nguồn web chính thức của gói."
                )
            source = "OFFICIAL_DOWNLOAD_BATCH"
        else:
            if not (claim.confirmed_by or "").strip():
                raise DocumentValidationError("HUMAN_LINKED phải ghi nhận người xác nhận Team Bid.")
            source = "TEAM_BID_CONFIRMATION"
        return _BundleMembership(claim_base, claim_revision, kind, source, evidence)

    @staticmethod
    def _same_origin(first_url: str, second_url: str) -> bool:
        first = urlsplit(first_url)
        second = urlsplit(second_url)
        return (
            first.scheme in {"http", "https"}
            and first.scheme == second.scheme
            and first.netloc.casefold() == second.netloc.casefold()
        )

    def _guard_duplicate_identity(
        self,
        duplicate: Document,
        identity: _IdentityResolution,
        membership: _BundleMembership,
    ) -> Document:
        if (
            identity.status == "NEEDS_REVIEW"
            and identity.expected is not None
            and duplicate.tender_id is not None
        ):
            detected = self._tender_identifier(duplicate.tender_id) or "UNLINKED"
            logger.error(
                "DOCUMENT_IDENTITY_MISMATCH expected=%s detected=%s duplicate_document_id=%s",
                identity.expected,
                detected,
                duplicate.id,
            )
            raise DocumentIdentityMismatch(identity.expected, detected)
        if identity.tender is None:
            return duplicate
        if duplicate.tender_id == identity.tender.id:
            self._guard_duplicate_bundle(duplicate, membership)
            return duplicate
        if duplicate.tender_id is None and identity.status == "DOCUMENT_VERIFIED":
            return self._link_verified_unlinked_duplicate(duplicate, identity, membership)
        expected = self._notice_identifier(identity.tender)
        detected = self._tender_identifier(duplicate.tender_id) or "UNLINKED"
        logger.error(
            "DOCUMENT_IDENTITY_MISMATCH expected=%s detected=%s duplicate_document_id=%s",
            expected,
            detected,
            duplicate.id,
        )
        raise DocumentIdentityMismatch(expected, detected)

    @staticmethod
    def _guard_duplicate_bundle(
        duplicate: Document,
        membership: _BundleMembership,
    ) -> None:
        """A SHA duplicate may only reuse a known equivalent logical bundle."""
        if not duplicate.bundle_base_notice_id or not duplicate.bundle_revision:
            return
        if (
            duplicate.bundle_base_notice_id == membership.base_notice_id
            and duplicate.bundle_revision == membership.revision
        ):
            return
        raise DocumentValidationError(
            "SHA-256 trùng nhưng thuộc bundle/revision khác; cần kiểm tra thủ công."
        )

    def _link_verified_unlinked_duplicate(
        self,
        duplicate: Document,
        identity: _IdentityResolution,
        membership: _BundleMembership,
    ) -> Document:
        """Link an old unlinked SHA duplicate only after content verifies its tender."""
        tender = identity.tender
        assert tender is not None
        content_identity = identity.content_identity
        assert content_identity is not None
        with self.database.session() as session:
            document = session.get(Document, duplicate.id)
            assert document is not None
            document.tender_id = tender.id
            document.version = self._next_version(tender.id)
            document.status = identity.status
            document.raw_notice_id = content_identity.raw_notice_id
            document.base_notice_id = content_identity.base_notice_id
            document.notice_revision = content_identity.revision
            document.identity_source = content_identity.identity_source
            document.identity_evidence_locator = content_identity.evidence_locator
            document.identity_match_status = identity.match_status
            document.identity_candidates_json = json.dumps(
                content_identity.candidates, ensure_ascii=False
            )
            document.bundle_base_notice_id = membership.base_notice_id
            document.bundle_revision = membership.revision
            document.bundle_membership_status = membership.status
            document.bundle_membership_source = membership.source
            document.bundle_membership_evidence = membership.evidence_locator
            session.flush()
            logger.info(
                "DOCUMENT_UNLINKED_DUPLICATE_LINKED document_id=%s tender=%s",
                document.id,
                self._notice_identifier(tender),
            )
            return document

    def _next_version(self, tender_id: int | None) -> int:
        with self.database.session() as session:
            predicate = (
                Document.tender_id == tender_id
                if tender_id is not None
                else Document.tender_id.is_(None)
            )
            current = session.scalar(select(func.max(Document.version)).where(predicate))
        return int(current or 0) + 1

    def _has_other_document_revision(
        self,
        tender_id: int,
        revision: str | None,
    ) -> bool:
        if revision is None:
            return False
        with self.database.session() as session:
            return session.scalar(
                select(Document.id)
                .where(
                    Document.tender_id == tender_id,
                    Document.notice_revision.is_not(None),
                    Document.notice_revision != revision,
                )
                .limit(1)
            ) is not None

    @staticmethod
    def _tender_scope(tender: Notice | None) -> str:
        if tender is None:
            return "unlinked"
        identifier = tender.notice_code or tender.source_notice_id or f"tender-{tender.id}"
        return _safe_scope(identifier)

    @staticmethod
    def _source_scope(tender: Notice | None, document_source: str) -> str:
        return _safe_scope(tender.source_name or document_source) if tender else _safe_scope(
            document_source
        )

    @staticmethod
    def _notice_identifier(tender: Notice) -> str:
        return tender.notice_code or tender.source_notice_id or str(tender.id)

    def _tender_identifier(self, tender_id: int | None) -> str | None:
        if tender_id is None:
            return None
        with self.database.session() as session:
            tender = session.get(Notice, tender_id)
            return self._notice_identifier(tender) if tender else None

    @staticmethod
    def _document_identity_status(document: Document) -> str:
        if document.status in {
            "VERIFIED_LINKED",
            "DOCUMENT_VERIFIED",
            "HUMAN_DECLARED",
            "UNLINKED",
            "MISMATCH",
            "NEEDS_REVIEW",
        }:
            return document.status
        return "VERIFIED_LINKED" if document.tender_id is not None else "UNLINKED"

    def manifest_for_tender(self, tender_reference: str) -> TenderDocumentManifest:
        """Return the database-backed immutable document manifest for one tender."""
        tender = self._lookup_tender(tender_reference)
        if tender is None:
            raise DocumentValidationError("Không tìm thấy tender để lập document manifest.")
        with self.database.session() as session:
            documents = tuple(
                session.scalars(
                    select(Document)
                    .where(Document.tender_id == tender.id)
                    .order_by(Document.version, Document.id)
                )
            )
        source = tender.source_name or "unknown"
        entries = tuple(
            DocumentManifestEntry(
                document_id=document.id,
                document_type=document.document_type,
                file_format=document.file_format or self._legacy_file_format(document),
                template_code=document.template_code,
                package_type=document.package_type,
                selection_method=document.selection_method,
                classification_status=document.classification_status or "UNKNOWN",
                filename=document.original_filename,
                sha256=document.sha256,
                version=document.version,
                source=document.document_source,
                status=self._document_identity_status(document),
                stored_path=Path(document.stored_path),
                uploaded_at=document.uploaded_at,
                raw_notice_id=document.raw_notice_id,
                base_notice_id=document.base_notice_id,
                notice_revision=document.notice_revision,
                identity_match_status=document.identity_match_status,
                bundle_base_notice_id=document.bundle_base_notice_id,
                bundle_revision=document.bundle_revision,
                bundle_membership_status=document.bundle_membership_status,
            )
            for document in documents
        )
        return TenderDocumentManifest(
            tender_id=tender.id,
            tender_identifier=self._notice_identifier(tender),
            tender_title=tender.title or "Chưa có tên gói",
            source=source,
            identity_status=(
                tender.identity_status
                or ("HUMAN_DECLARED" if tender.source_origin == "MANUAL_TEAM_BID" else "VERIFIED_LINKED")
            ),
            documents=entries,
        )

    def _atomic_store(self, source: Path, destination: Path) -> tuple[Path, bool]:
        temporary_path: Path | None = None
        try:
            destination.parent.mkdir(parents=True, exist_ok=True)
            if destination.exists():
                raise DocumentStorageError(
                    "Document Store đã có file cùng đường dẫn; không ghi đè."
                )
            with tempfile.NamedTemporaryFile(
                mode="wb",
                dir=destination.parent,
                prefix=".intake-",
                suffix=".tmp",
                delete=False,
            ) as target, source.open("rb") as origin:
                temporary_path = Path(target.name)
                for chunk in iter(lambda: origin.read(1024 * 1024), b""):
                    target.write(chunk)
                target.flush()
                os.fsync(target.fileno())
            os.link(temporary_path, destination)
            temporary_path.unlink()
            return destination.resolve(), True
        except FileExistsError as exc:
            raise DocumentStorageError("Document Store từ chối ghi đè file gốc.") from exc
        except DocumentIntakeError:
            raise
        except OSError as exc:
            raise DocumentStorageError("Không thể lưu file vào Document Store.") from exc
        finally:
            if temporary_path is not None and temporary_path.exists():
                temporary_path.unlink(missing_ok=True)

    def _create_record(
        self,
        *,
        tender_id: int | None,
        document_source: str,
        document_type: str,
        file_format: str,
        template_code: str | None,
        package_type: str | None,
        selection_method: str | None,
        classification_status: str,
        display_name: str | None,
        original_filename: str,
        stored_path: Path,
        mime_type: str,
        file_size: int,
        sha256: str,
        version: int,
        source_url: str | None,
        uploaded_by: str | None,
        zip_supported_entries: tuple[str, ...],
        status: str,
        content_identity: DocumentContentIdentity,
        identity_match_status: str | None,
        bundle_membership: _BundleMembership,
    ) -> Document:
        document = Document(
            tender_id=tender_id,
            document_source=document_source,
            document_type=document_type,
            file_format=file_format,
            template_code=template_code,
            package_type=package_type,
            selection_method=selection_method,
            classification_status=classification_status,
            display_name=display_name,
            original_filename=original_filename,
            stored_path=str(stored_path),
            mime_type=mime_type,
            file_size=file_size,
            sha256=sha256,
            version=version,
            source_url=source_url,
            uploaded_by=uploaded_by,
            status=status,
            raw_notice_id=content_identity.raw_notice_id,
            base_notice_id=content_identity.base_notice_id,
            notice_revision=content_identity.revision,
            identity_source=content_identity.identity_source,
            identity_evidence_locator=content_identity.evidence_locator,
            identity_match_status=identity_match_status,
            identity_candidates_json=(
                json.dumps(content_identity.candidates, ensure_ascii=False)
                if content_identity.candidates
                else None
            ),
            bundle_base_notice_id=bundle_membership.base_notice_id,
            bundle_revision=bundle_membership.revision,
            bundle_membership_status=bundle_membership.status,
            bundle_membership_source=bundle_membership.source,
            bundle_membership_evidence=bundle_membership.evidence_locator,
            zip_supported_entries=(
                json.dumps(zip_supported_entries, ensure_ascii=False)
                if zip_supported_entries
                else None
            ),
        )
        with self.database.session() as session:
            session.add(document)
            session.flush()
        return document

    @staticmethod
    def _cleanup_orphan(path: Path) -> None:
        path.unlink(missing_ok=True)
        try:
            path.parent.rmdir()
        except OSError:
            return

    @staticmethod
    def _result(
        document: Document,
        outcome: str,
        tender_identifier: str | None,
        *,
        identity_status: str,
        expected_identity: str | None,
        detected_identity: str | None,
    ) -> DocumentIntakeResult:
        entries = tuple(json.loads(document.zip_supported_entries or "[]"))
        return DocumentIntakeResult(
            outcome=outcome,
            document_id=document.id,
            original_filename=document.original_filename,
            stored_path=Path(document.stored_path),
            document_type=document.document_type,
            mime_type=document.mime_type,
            file_size=document.file_size,
            sha256=document.sha256,
            version=document.version,
            tender_id=document.tender_id,
            tender_identifier=tender_identifier,
            discovered_files=entries,
            identity_status=identity_status,
            expected_identity=expected_identity,
            detected_identity=detected_identity,
            file_format=document.file_format or DocumentIntakeService._legacy_file_format(
                document
            ),
            template_code=document.template_code,
            package_type=document.package_type,
            selection_method=document.selection_method,
            classification_status=document.classification_status or "UNKNOWN",
            raw_notice_id=document.raw_notice_id,
            base_notice_id=document.base_notice_id,
            notice_revision=document.notice_revision,
            identity_source=document.identity_source,
            identity_evidence_locator=document.identity_evidence_locator,
            identity_match_status=document.identity_match_status,
            identity_candidates=tuple(json.loads(document.identity_candidates_json or "[]")),
            bundle_base_notice_id=document.bundle_base_notice_id,
            bundle_revision=document.bundle_revision,
            bundle_membership_status=document.bundle_membership_status,
            bundle_membership_source=document.bundle_membership_source,
            bundle_membership_evidence=document.bundle_membership_evidence,
        )

    @staticmethod
    def _legacy_file_format(document: Document) -> str | None:
        if document.document_type in DOCUMENT_TYPES.values():
            return document.document_type
        return Path(document.original_filename).suffix.lstrip(".").upper() or None
