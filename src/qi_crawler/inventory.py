from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select

from .db import Database
from .keywords import normalize_keyword
from .models import InventoryItem, Notice, TenderItem


@dataclass(frozen=True)
class ImportQuantitySummary:
    rows: int
    inserted: int
    updated: int
    rejected: int


def _read_rows(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Supported files: .csv, .xlsx, .xlsm")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not values:
        return []
    headers = [str(value or "").strip() for value in values[0]]
    return [dict(zip(headers, row)) for row in values[1:] if any(value is not None for value in row)]


def _value(row: dict[str, Any], aliases: tuple[str, ...]) -> Any:
    def header(value: str) -> str:
        return " ".join(normalize_keyword(value).replace("_", " ").replace("-", " ").split())

    normalized = {header(str(key)): value for key, value in row.items()}
    return next(
        (normalized.get(header(alias)) for alias in aliases if header(alias) in normalized),
        None,
    )


def _number(value: Any) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None


def _boolean(value: Any) -> bool:
    return normalize_keyword(str(value or "")) in {"1", "true", "yes", "y", "verified", "da xac minh"}


def import_inventory(db: Database, path: Path) -> ImportQuantitySummary:
    rows = _read_rows(path)
    inserted = updated = rejected = 0
    db.create_all()
    with db.session() as session:
        for row in rows:
            sku = str(_value(row, ("sku", "item_code", "ma hang")) or "").strip()
            name = str(_value(row, ("product_name", "product", "name", "ten hang")) or "").strip()
            quantity = _number(_value(row, ("quantity_available", "quantity", "stock", "so luong ton")))
            if not sku or not name or quantity is None or quantity < 0:
                rejected += 1
                continue
            item = session.scalar(select(InventoryItem).where(InventoryItem.sku == sku))
            if item is None:
                item = InventoryItem(sku=sku, product_name=name)
                session.add(item)
                inserted += 1
            else:
                updated += 1
            item.product_name = name
            item.aliases = str(_value(row, ("aliases", "keywords", "ten khac")) or "").strip() or None
            item.quantity_available = quantity
            item.unit = str(_value(row, ("unit", "don vi")) or "").strip() or None
            item.warehouse = str(_value(row, ("warehouse", "kho")) or "").strip() or None
            item.source_file = str(path)
            item.verified = _boolean(_value(row, ("verified", "xac minh")))
            item.updated_at = datetime.now(UTC)
    return ImportQuantitySummary(len(rows), inserted, updated, rejected)


def import_tender_items(db: Database, notice_id: int, path: Path) -> ImportQuantitySummary:
    rows = _read_rows(path)
    inserted = updated = rejected = 0
    db.create_all()
    with db.session() as session:
        notice = session.get(Notice, notice_id)
        if notice is None:
            raise ValueError(f"Notice id {notice_id} does not exist")
        for index, row in enumerate(rows, start=1):
            code = str(_value(row, ("item_code", "sku", "stt", "ma hang")) or f"row-{index}").strip()
            name = str(_value(row, ("product_name", "product", "name", "ten hang")) or "").strip()
            quantity = _number(_value(row, ("quantity", "required_quantity", "so luong")))
            if not name:
                rejected += 1
                continue
            item = session.scalar(
                select(TenderItem).where(
                    TenderItem.notice_id == notice_id,
                    TenderItem.item_code == code,
                )
            )
            if item is None:
                item = TenderItem(notice_id=notice_id, item_code=code, product_name=name)
                session.add(item)
                inserted += 1
            else:
                updated += 1
            item.product_name = name
            item.specification = str(_value(row, ("specification", "spec", "technical_spec")) or "").strip() or None
            item.quantity = quantity
            item.minimum_quantity = _number(_value(row, ("minimum_quantity", "min_quantity")))
            item.maximum_quantity = _number(_value(row, ("maximum_quantity", "max_quantity")))
            item.unit = str(_value(row, ("unit", "don vi")) or "").strip() or None
            item.source_document = path.name
            item.source_location = f"row {index + 1}"
            item.extraction_confidence = 1.0
            item.needs_human_review = quantity is None
    return ImportQuantitySummary(len(rows), inserted, updated, rejected)
