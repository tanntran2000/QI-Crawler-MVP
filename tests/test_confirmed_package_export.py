from __future__ import annotations

from dataclasses import replace
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import func, select

from qi_crawler.db import Database
from qi_crawler.market_intelligence.candidate_review import (
    CandidateReviewService,
    HumanReviewDecision,
)
from qi_crawler.market_intelligence.confirmed_package_export import (
    AUDIT_HEADERS,
    BUSINESS_HEADERS,
    DEFAULT_CONFIRMED_EXPORT_FILENAME,
    export_confirmed_packages,
)
from qi_crawler.market_intelligence.khmt_contract import (
    OBSERVED_KHMT_HEADERS,
    load_sanitized_khmt_fixture,
)
from qi_crawler.models import CandidateReviewEvent

FIXTURE = Path(__file__).parent / "fixtures" / "khmt" / "khmt_sanitized_golden.json"


def _packages():
    packages = load_sanitized_khmt_fixture(FIXTURE)[1]
    return tuple(
        replace(
            package,
            provenance={
                **package.provenance,
                "source_filename": package.plan.import_batch.source_filename,
                "source_sha256": package.plan.import_batch.source_sha256,
                "sheet": package.plan.import_batch.sheet,
                "source_row": package.source_row,
            },
        )
        for package in packages
    )


def _service(tmp_path: Path) -> tuple[Database, CandidateReviewService]:
    database = Database(f"sqlite:///{tmp_path / 'confirmed-export.db'}")
    return database, CandidateReviewService(database)


def _with_batch(package, **changes):
    batch = replace(package.plan.import_batch, **changes)
    plan = replace(package.plan, import_batch=batch)
    provenance = dict(package.provenance)
    if "source_filename" in changes:
        provenance["source_filename"] = changes["source_filename"]
    if "source_sha256" in changes:
        provenance["source_sha256"] = changes["source_sha256"]
    return replace(package, plan=plan, provenance=provenance)


def test_synthetic_golden_exports_latest_confirmed_rows_with_exact_contract(
    tmp_path: Path,
) -> None:
    _, service = _service(tmp_path)
    first, second, third = _packages()[:3]
    first = replace(
        first,
        package_name="Thiết bị mạng — Thành phố Hồ Chí Minh",
        raw_fields={
            **first.raw_fields,
            "TÊN GÓI THẦU": "Thiết bị mạng — Thành phố Hồ Chí Minh",
            "GIÁ GÓI THẦU": Decimal("125000000.50"),
        },
    )
    events = [
        service.record_decision(
            package,
            decision=HumanReviewDecision.CONFIRMED,
            reviewer=f"Reviewer {index}",
            note="Đã kiểm tra",
        )
        for index, package in enumerate((first, second, third), start=1)
    ]
    output = tmp_path / DEFAULT_CONFIRMED_EXPORT_FILENAME

    result = export_confirmed_packages(
        service,
        (third, first, second, first),
        output=output,
    )

    assert result.output == output
    assert result.exported_rows == 3
    workbook = load_workbook(output, data_only=False)
    sheet = workbook.active
    assert BUSINESS_HEADERS == OBSERVED_KHMT_HEADERS
    assert tuple(cell.value for cell in sheet[1]) == OBSERVED_KHMT_HEADERS + AUDIT_HEADERS
    assert sheet.max_row == 4
    rows = tuple(sheet.iter_rows(min_row=2, values_only=True))
    assert rows[0][6] == "Thiết bị mạng — Thành phố Hồ Chí Minh"
    assert Decimal(str(rows[0][8])) == Decimal("125000000.5")
    assert [row[15] for row in rows] == [event.id for event in events]
    assert rows[0][13:] == (
        events[0].decision,
        events[0].reviewer,
        events[0].id,
        events[0].created_at.isoformat(),
        first.plan.import_batch.source_filename,
        events[0].source_sha256,
        events[0].source_sheet,
        events[0].source_row,
        events[0].plan_base_id,
        events[0].plan_revision,
    )
    assert all(cell.data_type != "f" for row in sheet.iter_rows() for cell in row)


def test_formula_like_raw_and_audit_values_are_exported_as_text(tmp_path: Path) -> None:
    _, service = _service(tmp_path)
    package = _with_batch(_packages()[0], source_filename="=SOURCE.xlsx")
    package = replace(
        package,
        raw_fields={
            **package.raw_fields,
            "NỘI DUNG PHÊ DUYỆT": " \t=WEBSERVICE(\"https://example.test\")",
        },
    )
    service.record_decision(
        package,
        decision="CONFIRMED",
        reviewer='=HYPERLINK("https://example.test","reviewer")',
    )

    result = export_confirmed_packages(
        service,
        (package,),
        output=tmp_path / "safe.xlsx",
    )

    sheet = load_workbook(result.output, data_only=False).active
    row = next(iter(sheet.iter_rows(min_row=2, max_row=2)))
    assert all(cell.data_type != "f" for cell in row)
    assert row[5].value.startswith("'")
    assert row[14].value.startswith("'")
    assert row[17].value.startswith("'")


def test_empty_export_uses_exact_default_filename(
    tmp_path: Path,
    monkeypatch,
) -> None:
    _, service = _service(tmp_path)
    monkeypatch.chdir(tmp_path)

    result = export_confirmed_packages(service, ())

    assert result.output == Path(DEFAULT_CONFIRMED_EXPORT_FILENAME)
    assert result.output.name == "CÁC GÓI ĐÃ XÁC NHẬN.xlsx"
    assert result.exported_rows == 0
    assert load_workbook(result.output).active.max_row == 1


def test_rejected_needs_review_and_unreviewed_never_export(tmp_path: Path) -> None:
    _, service = _service(tmp_path)
    confirmed, rejected, needs_review, unreviewed = _packages()[:4]
    service.record_decision(confirmed, decision="CONFIRMED", reviewer="Team Bid")
    service.record_decision(rejected, decision="REJECTED", reviewer="Team Bid")
    service.record_decision(
        needs_review,
        decision="NEEDS_REVIEW",
        reviewer="Team Bid",
    )

    result = export_confirmed_packages(
        service,
        (unreviewed, needs_review, rejected, confirmed),
        output=tmp_path / "confirmed.xlsx",
    )

    rows = tuple(
        load_workbook(result.output).active.iter_rows(min_row=2, values_only=True)
    )
    assert result.exported_rows == 1
    assert [row[6] for row in rows] == [confirmed.raw_fields["TÊN GÓI THẦU"]]


def test_historical_confirmation_disappears_after_rejection(tmp_path: Path) -> None:
    _, service = _service(tmp_path)
    packages = _packages()[:3]
    for package in packages:
        service.record_decision(package, decision="CONFIRMED", reviewer="Team Bid")
    first = export_confirmed_packages(
        service,
        packages,
        output=tmp_path / "first.xlsx",
    )
    service.record_decision(packages[1], decision="REJECTED", reviewer="Team Bid")

    second = export_confirmed_packages(
        service,
        packages,
        output=tmp_path / "second.xlsx",
    )

    assert first.exported_rows == 3
    assert second.exported_rows == 2
    assert load_workbook(second.output).active.max_row == 3


def test_source_identity_rename_sha_revision_and_row_rules_are_preserved(
    tmp_path: Path,
) -> None:
    _, service = _service(tmp_path)
    original, same_revision_other_row, other_revision = _packages()[:3]
    service.record_decision(original, decision="CONFIRMED", reviewer="Team Bid")
    service.record_decision(other_revision, decision="CONFIRMED", reviewer="Team Bid")
    renamed = _with_batch(original, source_filename="renamed-source.xlsx")
    changed_sha = _with_batch(original, source_sha256="b" * 64)

    result = export_confirmed_packages(
        service,
        (
            same_revision_other_row,
            changed_sha,
            other_revision,
            renamed,
            renamed,
        ),
        output=tmp_path / "identity.xlsx",
    )

    rows = tuple(
        load_workbook(result.output).active.iter_rows(min_row=2, values_only=True)
    )
    assert result.exported_rows == 2
    assert {(row[21], row[22]) for row in rows} == {
        (original.plan.plan_base_id, original.plan.plan_revision),
        (other_revision.plan.plan_base_id, other_revision.plan.plan_revision),
    }
    assert any(row[17] == "renamed-source.xlsx" for row in rows)


def test_export_and_reopen_do_not_mutate_review_history(tmp_path: Path) -> None:
    database, service = _service(tmp_path)
    packages = _packages()[:2]
    for package in packages:
        service.record_decision(package, decision="CONFIRMED", reviewer="Team Bid")
    with database.session() as session:
        before = session.scalar(select(func.count(CandidateReviewEvent.id)))

    result = export_confirmed_packages(
        service,
        packages,
        output=tmp_path / "read-only.xlsx",
    )
    load_workbook(result.output)

    with database.session() as session:
        after = session.scalar(select(func.count(CandidateReviewEvent.id)))
    assert before == after == 2
    assert [len(service.list_history(package)) for package in packages] == [1, 1]
