import csv
from pathlib import Path

from openpyxl import load_workbook

from qi_crawler.db import Database
from qi_crawler.exporter import export_csv, export_xlsx
from qi_crawler.inventory import import_inventory, import_tender_items
from qi_crawler.models import InventoryItem, Notice, TenderItem
from qi_crawler.stock import check_stock


def _db(tmp_path: Path) -> Database:
    db = Database(f"sqlite:///{tmp_path / 'stock.db'}")
    db.create_all()
    return db


def _notice(db: Database) -> int:
    with db.session() as session:
        notice = Notice(
            source_url="https://example.test/tender/1",
            url_hash="1" * 64,
            notice_code="QI-001",
            title="Router supply",
        )
        session.add(notice)
        session.flush()
        return notice.id


def test_import_inventory_and_tender_items(tmp_path: Path):
    db = _db(tmp_path)
    notice_id = _notice(db)
    inventory_file = tmp_path / "inventory.csv"
    inventory_file.write_text(
        "SKU,Product Name,Aliases,Quantity Available,Unit,Warehouse,Verified\n"
        "RTR-001,5G Router,router 5g,10,pieces,Main,yes\n",
        encoding="ascii",
    )
    boq_file = tmp_path / "boq.csv"
    boq_file.write_text(
        "Item Code,Product Name,Quantity,Unit,Specification\n"
        "1,5G Router,8,pieces,Three antenna ports\n",
        encoding="ascii",
    )

    inventory_summary = import_inventory(db, inventory_file)
    tender_summary = import_tender_items(db, notice_id, boq_file)

    assert inventory_summary.inserted == 1
    assert tender_summary.inserted == 1
    with db.session() as session:
        stock = session.query(InventoryItem).one()
        tender_item = session.query(TenderItem).one()
        result = check_stock(tender_item, [stock])
    assert result.status == "MEETS_STOCK"
    assert result.required_quantity == 8
    assert result.available_quantity == 10
    assert result.shortage_quantity == 0


def test_stock_check_requires_review_for_unit_mismatch(tmp_path: Path):
    db = _db(tmp_path)
    notice_id = _notice(db)
    with db.session() as session:
        item = TenderItem(
            notice_id=notice_id,
            item_code="1",
            product_name="White sand",
            quantity=10,
            unit="tonnes",
        )
        stock = InventoryItem(
            sku="SAND-01",
            product_name="White sand",
            quantity_available=100,
            unit="kg",
            verified=True,
        )
        session.add_all([item, stock])
        session.flush()
        result = check_stock(item, [stock])
    assert result.status == "REVIEW_UNIT_MISMATCH"
    assert result.shortage_quantity is None


def test_export_contains_quantity_and_response_table(tmp_path: Path):
    db = _db(tmp_path)
    notice_id = _notice(db)
    with db.session() as session:
        session.add_all(
            [
                TenderItem(
                    notice_id=notice_id,
                    item_code="1",
                    product_name="5G Router",
                    quantity=12,
                    unit="pieces",
                    source_document="boq.xlsx",
                    source_location="row 2",
                    extraction_confidence=1.0,
                    needs_human_review=False,
                ),
                InventoryItem(
                    sku="RTR-001",
                    product_name="5G Router",
                    quantity_available=10,
                    unit="pieces",
                    verified=True,
                ),
            ]
        )

    output = export_xlsx(db, tmp_path / "response.xlsx")
    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == [
            "Bản tin điện tử",
            "Notices",
            "Response Table",
            "QI Inventory",
        ]
        notice_headers = [cell.value for cell in workbook["Notices"][1]]
        response_index = notice_headers.index("response_table") + 1
        assert workbook["Notices"].cell(2, response_index).value == "STOCK_SHORTAGE=1"

        response_headers = [cell.value for cell in workbook["Response Table"][1]]
        status_index = response_headers.index("response_status") + 1
        shortage_index = response_headers.index("shortage_quantity") + 1
        assert workbook["Response Table"].cell(2, status_index).value == "STOCK_SHORTAGE"
        assert workbook["Response Table"].cell(2, shortage_index).value == 2
    finally:
        workbook.close()


def test_excel_export_starts_with_tbmt_template(tmp_path: Path) -> None:
    db = _db(tmp_path)
    with db.session() as session:
        session.add(
            Notice(
                source_url="https://example.test/tender/tbmt",
                url_hash="t" * 64,
                notice_code="IB2600000001-00",
                title="Cung cấp thiết bị mạng",
                buyer="Đơn vị mời thầu QI",
                investor="Chủ đầu tư QI",
                package_price=1250000000,
                published_at="07/08/2026 09:00",
                closing_at="17/08/2026 09:00",
                funding_source="Ngân sách nhà nước",
                selection_method="Một giai đoạn một túi hồ sơ",
                raw_text=(
                    "Địa chỉ bên mời thầu: 12 Lê Lợi, Hà Nội\n"
                    "Tên dự án: Nâng cấp hạ tầng mạng\n"
                    "Hình thức lựa chọn nhà thầu: Đấu thầu rộng rãi trong nước\n"
                    "Thời gian phát hành E-HSMT: 07/08/2026 09:00\n"
                    "Giá E-HSMT: 330.000 VND\n"
                    "Bảo đảm dự thầu: 20.000.000 VND\n"
                    "Hình thức bảo đảm dự thầu: Thư bảo lãnh\n"
                    "Thời điểm mở thầu: 17/08/2026 09:00\n"
                    "Thời gian thực hiện hợp đồng: 90 ngày"
                ),
            )
        )

    output = export_xlsx(db, tmp_path / "tbmt.xlsx")
    workbook = load_workbook(output, read_only=False, data_only=True)
    try:
        sheet = workbook["Bản tin điện tử"]
        assert [cell.value for cell in sheet[10]] == [
            "GÓI TIN",
            "BÊN MỜI THẦU",
            "ĐỊA CHỈ BÊN MỜI THẦU",
            "DỰ ÁN",
            "GÓI THẦU",
            "NỘI DUNG CHÍNH CỦA GÓI THẦU",
            "NGUỒN VỐN",
            "GIÁ GÓI THẦU",
            "PHƯƠNG THỨC LỰA CHỌN NHÀ THẦU",
            "HÌNH THỨC LỰA CHỌN NHÀ THẦU",
            "THỜI GIAN PHÁT HÀNH HSMT",
            "GIÁ BÁN 1 BỘ HSMT",
            "BẢO ĐẢM DỰ THẦU",
            "HÌNH THỨC BẢO ĐẢM DỰ THẦU",
            "ĐỊA ĐIỂM PHÁT HÀNH",
            "THỜI GIAN ĐÓNG THẦU(HẠN CUỐI TIẾP NHẬN BG)",
            "THỜI GIAN MỞ THẦU",
            "THỜI GIAN THỰC HIỆN HỢP ĐỒNG",
        ]
        assert sheet["A11"].value == "1. Thông báo mời thầu"
        assert sheet["B11"].value == "Đơn vị mời thầu QI"
        assert sheet["C11"].value == "12 Lê Lợi, Hà Nội"
        assert sheet["D11"].value == "Nâng cấp hạ tầng mạng"
        assert sheet["H11"].value == 1250000000
        assert sheet["I11"].value == "Một giai đoạn một túi hồ sơ"
        assert sheet["J11"].value == "Đấu thầu rộng rãi trong nước"
        assert sheet["O11"].value == "https://example.test/tender/tbmt"
        assert sheet.freeze_panes == "A11"
        assert sheet.auto_filter.ref == "A10:R11"
    finally:
        workbook.close()


def test_excel_export_preserves_vietnamese_unicode_on_windows(tmp_path: Path):
    db = _db(tmp_path)
    title = "Cung cap c\u00e1p quang tai L\u00e3nh Binh Th\u0103ng"
    with db.session() as session:
        session.add(
            Notice(
                source_url="https://example.test/tender/unicode",
                url_hash="u" * 64,
                notice_code="QI-UNICODE",
                title=title,
                buyer="Don vi mua s\u1eafm",
            )
        )

    output = export_xlsx(db, tmp_path / "unicode-report.xlsx")
    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        assert workbook["Notices"].cell(2, 3).value == title
        assert workbook["Notices"].cell(2, 4).value == "Don vi mua s\u1eafm"
    finally:
        workbook.close()


def test_export_neutralizes_spreadsheet_formula_injection(tmp_path: Path) -> None:
    db = _db(tmp_path)
    malicious = '=HYPERLINK("https://evil.test","open")'
    with db.session() as session:
        session.add(
            Notice(
                source_url="https://example.test/tender/formula",
                url_hash="f" * 64,
                notice_code="QI-FORMULA",
                title=malicious,
            )
        )

    xlsx_output = export_xlsx(db, tmp_path / "formula.xlsx")
    workbook = load_workbook(xlsx_output, read_only=True, data_only=False)
    try:
        assert workbook["Notices"].cell(2, 3).value == f"'{malicious}"
    finally:
        workbook.close()

    csv_output = export_csv(db, tmp_path / "formula.csv")
    with csv_output.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    assert rows[1][2] == f"'{malicious}"
