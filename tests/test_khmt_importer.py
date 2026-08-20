from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from qi_crawler.market_intelligence.khmt_contract import OBSERVED_KHMT_HEADERS
from qi_crawler.market_intelligence.khmt_importer import (
    KHMTImportError,
    KHMTIssueCode,
    import_khmt_workbook,
)


def _source_row(**overrides: object) -> dict[str, object]:
    row: dict[str, object] = {
        "GÓI TIN": "Synthetic bulletin",
        "SỐ KẾ HOẠCH": "PL2600000001-00",
        "TÊN DỰ ÁN": "Synthetic network project",
        "TÊN CHỦ ĐẦU TƯ": "Synthetic Services TP.HCM",
        "TỔNG MỨC ĐẦU TƯ": "900.000.000",
        "NỘI DUNG PHÊ DUYỆT": "Approved synthetic scope",
        "TÊN GÓI THẦU": "Synthetic network package",
        "NGUỒN VỐN": "Synthetic operating fund",
        "GIÁ GÓI THẦU": "477.850.000",
        "HÌNH THỨC LỰA CHỌN": "Chỉ định thầu rút gọn",
        "HÌNH THỨC HỢP ĐỒNG": "Trọn gói",
        "THỜI GIAN LỰA CHỌN": "Q4/2026",
        "THỜI GIAN THỰC HIỆN": "45 ngày",
    }
    row.update(overrides)
    return row


def _write_workbook(
    path: Path,
    rows: list[dict[str, object]],
    *,
    headers: list[str] | None = None,
    sheet_name: str = "KHMT",
) -> Path:
    selected_headers = headers or list(rows[0])
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(["Synthetic KHMT contract fixture"])
    sheet.append([])
    sheet.append(selected_headers)
    for row in rows:
        sheet.append([row.get(header) for header in selected_headers])
    workbook.save(path)
    return path


def test_valid_workbook_imports_rows_and_preserves_source_provenance(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path / "synthetic-khmt.xlsx",
        [
            _source_row(),
            _source_row(
                **{
                    "TÊN GÓI THẦU": "Synthetic installation package",
                    "GIÁ GÓI THẦU": 0,
                }
            ),
        ],
    )
    imported_at = datetime(2026, 8, 20, tzinfo=UTC)

    result = import_khmt_workbook(path, imported_at=imported_at)

    assert result.source_row_count == 2
    assert len(result.packages) == 2
    assert result.batch.source_filename == path.name
    assert len(result.batch.source_sha256) == 64
    assert result.batch.sheet == "KHMT"
    assert result.batch.imported_at == imported_at
    assert result.packages[0].source_row == 4
    assert result.packages[0].provenance["source_sha256"] == result.batch.source_sha256
    assert result.packages[1].package_price == 0
    assert result.packages[0].plan is result.packages[1].plan


def test_all_known_headers_and_extra_source_column_are_preserved(tmp_path: Path) -> None:
    row = _source_row(EXTRA_SOURCE_FIELD="Synthetic extra value")
    path = _write_workbook(tmp_path / "extra.xlsx", [row])

    package = import_khmt_workbook(path).packages[0]

    assert tuple(package.raw_fields) == (*OBSERVED_KHMT_HEADERS, "EXTRA_SOURCE_FIELD")
    assert package.raw_fields["EXTRA_SOURCE_FIELD"] == "Synthetic extra value"


def test_blank_header_column_does_not_shift_following_source_values(tmp_path: Path) -> None:
    row = _source_row()
    headers: list[str | None] = [*OBSERVED_KHMT_HEADERS[:2], None, *OBSERVED_KHMT_HEADERS[2:]]
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    sheet.append([row.get(header) if header is not None else "ignored" for header in headers])
    path = tmp_path / "blank-column.xlsx"
    workbook.save(path)

    package = import_khmt_workbook(path).packages[0]

    assert package.project == row["TÊN DỰ ÁN"]
    assert package.package_name == row["TÊN GÓI THẦU"]


@pytest.mark.parametrize("missing_header", ["SỐ KẾ HOẠCH", "TÊN GÓI THẦU"])
def test_missing_required_header_is_explicit_schema_failure(
    tmp_path: Path, missing_header: str
) -> None:
    row = _source_row()
    headers = [header for header in row if header != missing_header]
    path = _write_workbook(tmp_path / "missing.xlsx", [row], headers=headers)

    with pytest.raises(KHMTImportError, match="MISSING_REQUIRED_HEADER") as raised:
        import_khmt_workbook(path)

    assert raised.value.code is KHMTIssueCode.MISSING_REQUIRED_HEADER


def test_duplicate_header_is_explicit_ambiguity_failure(tmp_path: Path) -> None:
    row = _source_row()
    headers = [*row, "TÊN GÓI THẦU"]
    path = _write_workbook(tmp_path / "duplicate.xlsx", [row], headers=headers)

    with pytest.raises(KHMTImportError) as raised:
        import_khmt_workbook(path)

    assert raised.value.code is KHMTIssueCode.DUPLICATE_HEADER


def test_blank_optional_values_import_without_fabrication(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path / "optional.xlsx",
        [
            _source_row(
                **{
                    "TỔNG MỨC ĐẦU TƯ": None,
                    "NGUỒN VỐN": None,
                    "GIÁ GÓI THẦU": None,
                    "HÌNH THỨC HỢP ĐỒNG": None,
                }
            )
        ],
    )

    package = import_khmt_workbook(path).packages[0]

    assert package.total_investment_raw is None
    assert package.funding_source is None
    assert package.package_price_raw is None
    assert package.package_price is None
    assert package.contract_type_raw is None


def test_malformed_row_is_visible_without_hiding_valid_rows(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path / "row-issue.xlsx",
        [_source_row(), _source_row(**{"SỐ KẾ HOẠCH": "not-a-plan"})],
    )

    result = import_khmt_workbook(path)

    assert len(result.packages) == 1
    assert result.source_row_count == 2
    assert [issue.code for issue in result.issues] == [KHMTIssueCode.INVALID_PLAN_ID]
    assert result.issues[0].source_row == 5


def test_schema_error_never_returns_false_safe_empty_import(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active["A1"] = "Unrelated workbook"
    path = tmp_path / "unrelated.xlsx"
    workbook.save(path)

    with pytest.raises(KHMTImportError) as raised:
        import_khmt_workbook(path)

    assert raised.value.code is KHMTIssueCode.NO_USABLE_SHEET


def test_invalid_price_is_none_with_issue_while_zero_remains_numeric(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path / "prices.xlsx",
        [
            _source_row(**{"GIÁ GÓI THẦU": "ambiguous 12,34"}),
            _source_row(
                **{
                    "SỐ KẾ HOẠCH": "PL2600000002-00",
                    "TÊN GÓI THẦU": "Synthetic zero-price package",
                    "GIÁ GÓI THẦU": 0,
                }
            ),
        ],
    )

    result = import_khmt_workbook(path)

    assert result.packages[0].package_price is None
    assert result.packages[0].package_price_raw == "ambiguous 12,34"
    assert result.packages[1].package_price == 0
    assert KHMTIssueCode.INVALID_PRICE in {issue.code for issue in result.issues}
