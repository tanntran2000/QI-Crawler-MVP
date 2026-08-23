from __future__ import annotations

import hashlib
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from qi_crawler.market_intelligence.opportunity_contract import OpportunitySourceType
from qi_crawler.market_intelligence.tbmt_importer import (
    TBMTImportError,
    TBMTIssueCode,
    import_tbmt_workbook,
)
from qi_crawler.market_intelligence.tbmt_schema import OBSERVED_TBMT_HEADERS


def _write_workbook(
    tmp_path: Path,
    *,
    rows: list[dict[str, object]],
    headers: tuple[str, ...] = OBSERVED_TBMT_HEADERS,
    filename: str = "TBMT_synthetic.xlsx",
    sheet_name: str = "TBMT",
    header_row: int = 10,
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=header_row, column=column, value=header)
    for row_offset, values in enumerate(rows, start=header_row + 1):
        for column, header in enumerate(headers, start=1):
            sheet.cell(row=row_offset, column=column, value=values.get(header))
    path = tmp_path / filename
    workbook.save(path)
    return path


def _row(
    *,
    package: object = "Thiết bị mạng IB2600463290-00",
    price: object = "1.234.567",
    project: object = "Dự án mẫu",
    funding: object = "Ngân sách mẫu",
) -> dict[str, object]:
    return {
        "GÓI THẦU": package,
        "BÊN MỜI THẦU": "Bên mời thầu mẫu",
        "ĐỊA CHỈ BÊN MỜI THẦU": "Địa chỉ bên mời thầu",
        "DỰ ÁN": project,
        "NỘI DUNG CHÍNH CỦA GÓI THẦU": "Nội dung nguồn",
        "NGUỒN VỐN": funding,
        "GIÁ GÓI THẦU": price,
        "PHƯƠNG THỨC LỰA CHỌN NHÀ THẦU": "Một giai đoạn",
        "HÌNH THỨC LỰA CHỌN NHÀ THẦU": "Đấu thầu rộng rãi",
        "THỜI GIAN PHÁT HÀNH HSMT": "01/08/2026",
        "GIÁ BÁN 1 BỘ HSMT": 23,
        "BẢO ĐẢM DỰ THẦU": "23",
        "HÌNH THỨC BẢO ĐẢM DỰ THẦU": "Tiền mặt",
        "ĐỊA ĐIỂM PHÁT HÀNH": "Địa điểm phát hành",
        "THỜI GIAN ĐÓNG THẦU(HẠN CUỐI TIẾP NHẬN BG)": "10/08/2026",
        "THỜI GIAN MỞ THẦU": "10/08/2026",
        "THỜI GIAN THỰC HIỆN HỢP ĐỒNG": "120 ngày",
    }


def test_import_discovers_header_below_first_row_and_builds_tbmt_candidate(
    tmp_path: Path,
) -> None:
    path = _write_workbook(tmp_path, rows=[_row()])

    result = import_tbmt_workbook(
        path,
        imported_at=datetime(2026, 8, 23, tzinfo=UTC),
    )

    assert result.batch.source_type is OpportunitySourceType.TBMT
    assert result.batch.sheet == "TBMT"
    assert result.source_row_count == 1
    assert len(result.candidates) == 1
    assert result.candidates[0].identity.namespace.value == "IB"
    assert result.candidates[0].source_row == 11


def test_import_preserves_sha_provenance_and_tbmt_specific_raw_fields(tmp_path: Path) -> None:
    path = _write_workbook(tmp_path, rows=[_row()])
    expected_sha = hashlib.sha256(path.read_bytes()).hexdigest()

    result = import_tbmt_workbook(path)
    candidate = result.candidates[0]

    assert result.batch.source_sha256 == expected_sha
    assert candidate.provenance == {
        "source_filename": path.name,
        "source_sha256": expected_sha,
        "sheet": "TBMT",
        "source_row": 11,
    }
    assert candidate.raw_fields["BÊN MỜI THẦU"] == "Bên mời thầu mẫu"
    assert candidate.raw_fields["ĐỊA CHỈ BÊN MỜI THẦU"] == "Địa chỉ bên mời thầu"
    assert candidate.raw_fields["BẢO ĐẢM DỰ THẦU"] == "23"
    assert candidate.raw_fields["GIÁ BÁN 1 BỘ HSMT"] == 23


def test_import_maps_common_fields_without_inventing_execution_location(tmp_path: Path) -> None:
    path = _write_workbook(tmp_path, rows=[_row()])

    candidate = import_tbmt_workbook(path).candidates[0]

    assert candidate.project == "Dự án mẫu"
    assert candidate.funding_source == "Ngân sách mẫu"
    assert candidate.package_price_raw == "1.234.567"
    assert candidate.package_price == Decimal(1234567)
    assert candidate.location_detail_raw is None
    assert candidate.raw_fields["ĐỊA CHỈ BÊN MỜI THẦU"] == "Địa chỉ bên mời thầu"
    assert candidate.raw_fields["ĐỊA ĐIỂM PHÁT HÀNH"] == "Địa điểm phát hành"


def test_import_preserves_full_package_source_text_without_stripping_identity(
    tmp_path: Path,
) -> None:
    package = "Gói mua sắm IB2600463290- 00"
    path = _write_workbook(tmp_path, rows=[_row(package=package)])

    candidate = import_tbmt_workbook(path).candidates[0]

    assert candidate.package_name == package
    assert candidate.identity.raw_id == "IB2600463290- 00"


@pytest.mark.parametrize(
    "package",
    [
        "Không có mã IB",
        "PL2600245672-02",
        "IB2600463290",
        "IB2600463290-0A",
        "IB2600463290-00 và IB2600463291-00",
        "PL2600245672-02 / IB2600463290-00",
        "IB2600463290-00-01",
    ],
)
def test_invalid_ib_identity_creates_issue_without_candidate(
    tmp_path: Path,
    package: str,
) -> None:
    path = _write_workbook(tmp_path, rows=[_row(package=package)])

    result = import_tbmt_workbook(path)

    assert result.candidates == ()
    assert [(issue.code, issue.source_row, issue.source_field) for issue in result.issues] == [
        (TBMTIssueCode.INVALID_IB_IDENTITY, 11, "GÓI THẦU")
    ]


def test_invalid_non_empty_price_keeps_candidate_and_records_issue(tmp_path: Path) -> None:
    path = _write_workbook(tmp_path, rows=[_row(price="không rõ")])

    result = import_tbmt_workbook(path)

    assert len(result.candidates) == 1
    assert result.candidates[0].package_price_raw == "không rõ"
    assert result.candidates[0].package_price is None
    assert result.issues[0].code is TBMTIssueCode.INVALID_PRICE
    assert result.issues[0].source_row == 11


def test_blank_package_creates_issue_and_skips_candidate(tmp_path: Path) -> None:
    path = _write_workbook(tmp_path, rows=[_row(package="  ")])

    result = import_tbmt_workbook(path)

    assert result.candidates == ()
    assert result.issues[0].code is TBMTIssueCode.EMPTY_PACKAGE_NAME
    assert result.issues[0].source_field == "GÓI THẦU"


def test_duplicate_valid_identity_rows_remain_two_candidates(tmp_path: Path) -> None:
    path = _write_workbook(tmp_path, rows=[_row(), _row()])

    result = import_tbmt_workbook(path)

    assert len(result.candidates) == 2
    assert [candidate.source_row for candidate in result.candidates] == [11, 12]
    assert result.candidates[0].identity.base_id == result.candidates[1].identity.base_id


def test_same_ib_lineage_different_revisions_remain_two_candidates(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path,
        rows=[
            _row(package="Mã TBMT : IB2600462391-00-Gói mẫu lần đầu"),
            _row(package="Mã TBMT : IB2600462391-01-Gói mẫu cập nhật"),
        ],
    )

    result = import_tbmt_workbook(path)

    assert len(result.candidates) == 2
    first, second = result.candidates
    assert first.identity.base_id == second.identity.base_id == "IB2600462391"
    assert first.identity.revision == "00"
    assert second.identity.revision == "01"
    assert first.identity.raw_id != second.identity.raw_id
    assert [first.source_row, second.source_row] == [11, 12]
    assert first.provenance["source_row"] != second.provenance["source_row"]


def test_missing_required_header_fails_closed(tmp_path: Path) -> None:
    headers = tuple(header for header in OBSERVED_TBMT_HEADERS if header != "DỰ ÁN")
    path = _write_workbook(tmp_path, headers=headers, rows=[])

    with pytest.raises(TBMTImportError) as error:
        import_tbmt_workbook(path)

    assert error.value.code is TBMTIssueCode.MISSING_REQUIRED_HEADER


def test_duplicate_canonical_header_fails_closed(tmp_path: Path) -> None:
    headers = OBSERVED_TBMT_HEADERS + (" gói   thầu ",)
    path = _write_workbook(tmp_path, headers=headers, rows=[])

    with pytest.raises(TBMTImportError) as error:
        import_tbmt_workbook(path)

    assert error.value.code is TBMTIssueCode.DUPLICATE_HEADER


def test_missing_requested_sheet_fails_closed(tmp_path: Path) -> None:
    path = _write_workbook(tmp_path, rows=[])

    with pytest.raises(TBMTImportError) as error:
        import_tbmt_workbook(path, sheet_name="Missing")

    assert error.value.code is TBMTIssueCode.NO_USABLE_SHEET


def test_khmt_like_workbook_does_not_import_as_tbmt(tmp_path: Path) -> None:
    khmt_headers = ("SỐ KẾ HOẠCH", "TÊN GÓI THẦU", "TÊN CHỦ ĐẦU TƯ")
    path = _write_workbook(tmp_path, headers=khmt_headers, rows=[])

    with pytest.raises(TBMTImportError) as error:
        import_tbmt_workbook(path)

    assert error.value.code is TBMTIssueCode.NO_USABLE_SHEET


def test_import_reads_without_mutating_source_workbook(tmp_path: Path) -> None:
    path = _write_workbook(tmp_path, rows=[_row()])
    before = path.read_bytes()

    import_tbmt_workbook(path)

    assert path.read_bytes() == before


def test_nonexistent_or_unsupported_path_fails_closed(tmp_path: Path) -> None:
    missing = tmp_path / "missing.xlsx"
    with pytest.raises(TBMTImportError) as missing_error:
        import_tbmt_workbook(missing)
    assert missing_error.value.code is TBMTIssueCode.UNSUPPORTED_WORKBOOK

    text_file = tmp_path / "source.csv"
    text_file.write_text("not a workbook", encoding="utf-8")
    with pytest.raises(TBMTImportError) as suffix_error:
        import_tbmt_workbook(text_file)
    assert suffix_error.value.code is TBMTIssueCode.UNSUPPORTED_WORKBOOK
