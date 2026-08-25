from __future__ import annotations

import hashlib
from datetime import UTC, datetime
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook

from qi_crawler.gui_services import run_bid_radar_export, run_bid_radar_legal_docx
from qi_crawler.market_intelligence.khmt_contract import OBSERVED_KHMT_HEADERS
from qi_crawler.market_intelligence.opportunity_contract import OpportunitySourceType
from qi_crawler.market_intelligence.opportunity_intelligence import (
    OpportunityIntelligenceService,
)
from qi_crawler.market_intelligence.opportunity_review import (
    OpportunityReviewDecision,
    OpportunityReviewRecord,
    OpportunityReviewService,
)
from qi_crawler.market_intelligence.tbmt_schema import OBSERVED_TBMT_HEADERS


class MemoryReviewRepository:
    def __init__(self) -> None:
        self.records: list[OpportunityReviewRecord] = []

    def latest(self, observation_key: str) -> OpportunityReviewRecord | None:
        matches = [record for record in self.records if record.identity.observation_key == observation_key]
        return max(matches, key=lambda record: record.event_id, default=None)

    def history(self, observation_key: str) -> tuple[OpportunityReviewRecord, ...]:
        return tuple(record for record in self.records if record.identity.observation_key == observation_key)

    def append(self, write) -> OpportunityReviewRecord:
        record = OpportunityReviewRecord(
            event_id=len(self.records) + 1,
            identity=write.identity,
            decision=write.decision,
            reviewer=write.reviewer,
            note=write.note,
            opportunity_snapshot_json=write.opportunity_snapshot_json,
            snapshot_schema_version=write.snapshot_schema_version,
            created_at=datetime(2026, 8, 25, tzinfo=UTC),
        )
        self.records.append(record)
        return record

    def latest_for_keys(self, observation_keys: tuple[str, ...]):
        return {
            key: record
            for key in observation_keys
            if (record := self.latest(key)) is not None
        }


def _service() -> tuple[OpportunityIntelligenceService, MemoryReviewRepository]:
    repository = MemoryReviewRepository()
    return OpportunityIntelligenceService(OpportunityReviewService(repository)), repository


def _write_khmt(path: Path, *, plan_id: str = "PL2600000001-00") -> Path:
    values = {
        "GÓI TIN": "Synthetic bulletin",
        "SỐ KẾ HOẠCH": plan_id,
        "TÊN DỰ ÁN": "Synthetic project",
        "TÊN CHỦ ĐẦU TƯ": "Synthetic investor",
        "TỔNG MỨC ĐẦU TƯ": "900.000.000",
        "NỘI DUNG PHÊ DUYỆT": "Approved synthetic scope",
        "TÊN GÓI THẦU": "Synthetic package",
        "NGUỒN VỐN": "Synthetic fund",
        "GIÁ GÓI THẦU": "477.850.000",
        "HÌNH THỨC LỰA CHỌN": "Đấu thầu rộng rãi",
        "HÌNH THỨC HỢP ĐỒNG": "Trọn gói",
        "THỜI GIAN LỰA CHỌN": "Q4/2026",
        "THỜI GIAN THỰC HIỆN": "45 ngày",
    }
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "KHMT"
    sheet.append(list(OBSERVED_KHMT_HEADERS))
    sheet.append([values.get(header) for header in OBSERVED_KHMT_HEADERS])
    workbook.save(path)
    return path


def _write_tbmt(path: Path, *, identity: str = "IB2600463290-00") -> Path:
    values = {
        "GÓI THẦU": f"Gói TBMT {identity}",
        "BÊN MỜI THẦU": "Synthetic procuring entity",
        "ĐỊA CHỈ BÊN MỜI THẦU": "Synthetic address",
        "DỰ ÁN": "Synthetic project",
        "NỘI DUNG CHÍNH CỦA GÓI THẦU": "Synthetic scope",
        "NGUỒN VỐN": "Synthetic fund",
        "GIÁ GÓI THẦU": "1.234.567",
        "PHƯƠNG THỨC LỰA CHỌN NHÀ THẦU": "Một giai đoạn",
        "HÌNH THỨC LỰA CHỌN NHÀ THẦU": "Đấu thầu rộng rãi",
        "THỜI GIAN PHÁT HÀNH HSMT": "01/08/2026",
        "GIÁ BÁN 1 BỘ HSMT": 23,
        "BẢO ĐẢM DỰ THẦU": "23",
        "HÌNH THỨC BẢO ĐẢM DỰ THẦU": "Tiền mặt",
        "ĐỊA ĐIỂM PHÁT HÀNH": "Synthetic location",
        "THỜI GIAN ĐÓNG THẦU(HẠN CUỐI TIẾP NHẬN BG)": "10/08/2026",
        "THỜI GIAN MỞ THẦU": "10/08/2026",
        "THỜI GIAN THỰC HIỆN HỢP ĐỒNG": "120 ngày",
    }
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TBMT"
    sheet.append(list(OBSERVED_TBMT_HEADERS))
    sheet.append([values.get(header) for header in OBSERVED_TBMT_HEADERS])
    workbook.save(path)
    return path


def test_confirmed_khmt_export_preserves_identity_and_review(tmp_path: Path) -> None:
    service, repository = _service()
    source = _write_khmt(tmp_path / "KHMT-synthetic.xlsx")
    loaded = service.load_workbook(source, OpportunitySourceType.KHMT)
    service.record_review(loaded.items[0], decision="CONFIRMED", reviewer="Team Bid")

    result = service.export_confirmed(loaded, output=tmp_path / "confirmed.xlsx")
    workbook = load_workbook(result.output, read_only=True, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))

    assert result.source_type is OpportunitySourceType.KHMT
    assert result.source_sha256 == loaded.source_sha256
    assert result.exported_rows == 1
    assert rows[0][:3] == ("LOẠI NGUỒN", "NAMESPACE", "MÃ CƠ HỘI")
    assert rows[1][0:5] == ("KHMT", "PL", "PL2600000001-00", "PL2600000001", "00")
    assert rows[1][25:30] == (
        "KHMT-synthetic.xlsx",
        loaded.source_sha256,
        "KHMT",
        2,
        "mi-1",
    )
    assert len(repository.records) == 1


def test_confirmed_tbmt_export_is_source_neutral_and_not_plan_package(tmp_path: Path) -> None:
    service, _ = _service()
    source = _write_tbmt(tmp_path / "TBMT-synthetic.xlsx")
    loaded = service.load_workbook(source, OpportunitySourceType.TBMT)
    service.record_review(loaded.items[0], decision="CONFIRMED", reviewer="Team Bid")

    result = service.export_confirmed(loaded, output=tmp_path / "tbmt-confirmed.xlsx")
    rows = list(load_workbook(result.output, read_only=True).active.iter_rows(values_only=True))

    assert result.source_type is OpportunitySourceType.TBMT
    assert rows[1][0:5] == ("TBMT", "IB", "IB2600463290-00", "IB2600463290", "00")


def test_only_latest_confirmed_observations_export(tmp_path: Path) -> None:
    service, _ = _service()
    source = _write_tbmt(tmp_path / "states.xlsx")
    loaded = service.load_workbook(source, OpportunitySourceType.TBMT)
    item = loaded.items[0]
    service.record_review(item, decision="CONFIRMED", reviewer="Team Bid")
    service.record_review(item, decision="REJECTED", reviewer="Team Bid")

    result = service.export_confirmed(loaded, output=tmp_path / "states-out.xlsx")

    assert result.exported_rows == 0
    rows = list(load_workbook(result.output, read_only=True).active.iter_rows(values_only=True))
    assert len(rows) == 1
    assert rows[0][0] == "LOẠI NGUỒN"


def test_same_path_mutation_fails_closed_and_preserves_existing_output(tmp_path: Path) -> None:
    service, _ = _service()
    source = _write_khmt(tmp_path / "mutable.xlsx")
    loaded = service.load_workbook(source, OpportunitySourceType.KHMT)
    service.record_review(loaded.items[0], decision=OpportunityReviewDecision.CONFIRMED, reviewer="Team Bid")
    output = tmp_path / "existing.xlsx"
    output.write_bytes(b"known output")
    source.write_bytes(b"changed source")

    from qi_crawler.market_intelligence.source_integrity import OpportunitySourceIntegrityError

    with pytest.raises(OpportunitySourceIntegrityError):
        service.export_confirmed(loaded, output=output)
    assert output.read_bytes() == b"known output"


def test_missing_source_fails_closed(tmp_path: Path) -> None:
    service, _ = _service()
    source = _write_khmt(tmp_path / "missing.xlsx")
    loaded = service.load_workbook(source, OpportunitySourceType.KHMT)
    service.record_review(loaded.items[0], decision="CONFIRMED", reviewer="Team Bid")
    source.unlink()

    from qi_crawler.market_intelligence.source_integrity import OpportunitySourceIntegrityError

    with pytest.raises(OpportunitySourceIntegrityError):
        service.export_confirmed(loaded, output=tmp_path / "missing-out.xlsx")


@pytest.mark.parametrize("adapter", [run_bid_radar_export, run_bid_radar_legal_docx])
def test_gui_delivery_adapters_verify_source_before_legacy_export(
    adapter,
    tmp_path: Path,
) -> None:
    source = tmp_path / "delivery.xlsx"
    source.write_bytes(b"source-a")
    expected_sha256 = hashlib.sha256(source.read_bytes()).hexdigest()
    source.write_bytes(b"source-b")
    config = SimpleNamespace(
        storage=SimpleNamespace(
            database_url=f"sqlite:///{tmp_path / 'unused.db'}",
            report_dir=tmp_path / "reports",
        )
    )

    from qi_crawler.market_intelligence.source_integrity import OpportunitySourceIntegrityError

    with pytest.raises(OpportunitySourceIntegrityError):
        adapter(
            config,
            (),
            source_path=source,
            expected_source_sha256=expected_sha256,
        )
