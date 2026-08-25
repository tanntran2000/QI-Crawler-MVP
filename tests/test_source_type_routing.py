from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from qi_crawler import gui_services
from qi_crawler.config import AppConfig
from qi_crawler.market_intelligence.search import TargetedSearchRequest
from qi_crawler.market_intelligence.source_detection import SourceType
from qi_crawler.market_intelligence.tbmt_schema import OBSERVED_TBMT_HEADERS


def test_tbmt_service_route_does_not_call_khmt_importer(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    called = False

    def forbidden(*args, **kwargs):
        nonlocal called
        called = True
        raise AssertionError("KHMT importer must not run for TBMT")

    monkeypatch.setattr(gui_services, "import_khmt_workbook", forbidden)
    source = tmp_path / "TBMT.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TBMT"
    values = {
        "GÓI THẦU": "Gói mẫu IB2600463290-00",
        "GIÁ GÓI THẦU": "1000",
    }
    for column, header in enumerate(OBSERVED_TBMT_HEADERS, start=1):
        sheet.cell(row=1, column=column, value=header)
        sheet.cell(row=2, column=column, value=values.get(header))
    workbook.save(source)
    config = AppConfig()
    config.storage.database_url = f"sqlite:///{tmp_path / 'routing.db'}"
    request = TargetedSearchRequest()

    result = gui_services.run_bid_radar_import_search(
        config,
        source,
        request,
        source_type=SourceType.TBMT,
    )

    assert called is False
    assert result.items[0].identity.namespace.value == "IB"
