from datetime import UTC, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from qi_crawler.models import CompanyEvidence, InventoryItem, Notice
from qi_crawler.monitoring import _write_ranked_report, assess_notice_feasibility
from qi_crawler.opportunity import KeywordGroup, assess_opportunity

NOW = datetime(2026, 7, 31, tzinfo=UTC)


def _notice(deadline: datetime, **overrides) -> Notice:
    values = {
        "id": 1,
        "source_url": "https://example.test/tender/1",
        "url_hash": "a" * 64,
        "notice_code": "IB260001",
        "notice_version": "1",
        "title": "Supply of outdoor network cable and network accessories",
        "raw_text": "The buyer requires outdoor LAN cable, switches and installation support.",
        "buyer": "Example Buyer",
        "package_price": 2_000_000_000,
        "currency": "VND",
        "published_at": "2026-07-30T08:00:00+00:00",
        "closing_at": deadline.isoformat(),
        "location": "Ho Chi Minh City",
        "sector": "Information Technology",
        "selection_method": "Open bidding",
        "first_seen_at": NOW - timedelta(hours=2),
    }
    values.update(overrides)
    return Notice(**values)


def _evidence(code: str, evidence_type: str, keywords: str = "network cable") -> CompanyEvidence:
    return CompanyEvidence(
        evidence_code=code,
        title=code,
        evidence_type=evidence_type,
        keywords=keywords,
        verified=True,
    )


def _inventory() -> InventoryItem:
    return InventoryItem(
        sku="NET-001",
        product_name="Outdoor network cable",
        aliases="LAN cable; network cable",
        quantity_available=100,
        unit="pieces",
        verified=True,
    )


def test_complete_opportunity_with_evidence_is_priority() -> None:
    assessment = assess_notice_feasibility(
        _notice(NOW + timedelta(days=14)),
        ("cap mang", "network cable", "LAN cable"),
        [
            _evidence("PROD-01", "product"),
            _evidence("EXP-01", "contract"),
            _evidence("FIN-01", "financial", "finance"),
            _evidence("SLA-01", "sla", "Ho Chi Minh City"),
        ],
        [_inventory()],
        now=NOW,
    )

    assert assessment.status == "PRIORITY"
    assert assessment.priority == "A"
    assert assessment.score is not None and assessment.score >= 75
    assert "EXP-01" in assessment.matched_evidence


def test_financial_and_payment_evidence_can_reach_full_component_score() -> None:
    assessment = assess_notice_feasibility(
        _notice(NOW + timedelta(days=14)),
        ("network cable",),
        [
            _evidence("FIN-01", "financial", "finance"),
            _evidence("PAY-01", "payment", "payment terms"),
        ],
        [_inventory()],
        now=NOW,
    )

    component = next(
        item for item in assessment.components if item.name == "financial_payment"
    )
    assert component.score == component.maximum == 10


def test_product_match_without_similar_contract_needs_review() -> None:
    assessment = assess_notice_feasibility(
        _notice(NOW + timedelta(days=14)),
        ("network cable",),
        [],
        [_inventory()],
        now=NOW,
    )

    assert assessment.status == "REVIEW"
    assert assessment.score is not None and 55 <= assessment.score < 75


def test_missing_price_is_not_ranked() -> None:
    assessment = assess_notice_feasibility(
        _notice(NOW + timedelta(days=14), package_price=None),
        ("network cable",),
        [_evidence("PROD-01", "product"), _evidence("EXP-01", "contract")],
        [_inventory()],
        now=NOW,
    )

    assert assessment.status == "INSUFFICIENT_DATA"
    assert assessment.score is None
    assert "package_price" in assessment.missing_fields


def test_expired_notice_is_skipped() -> None:
    assessment = assess_notice_feasibility(
        _notice(NOW - timedelta(days=1)),
        ("network cable",),
        [],
        now=NOW,
    )

    assert assessment.status == "SKIP"
    assert assessment.score == 0


def test_negative_and_boolean_keywords_are_explained() -> None:
    notice = _notice(
        NOW + timedelta(days=10),
        raw_text="Supply network switch for a civil construction project",
    )
    assessment = assess_opportunity(
        notice,
        (KeywordGroup("Network", ("switch", "router"), 30),),
        [],
        [],
        required_any=("switch", "firewall"),
        required_all=("network",),
        excluded_terms=("civil construction", "medicine"),
        now=NOW,
    )

    assert assessment.status == "SKIP"
    assert assessment.excluded_keywords == ("civil construction",)
    assert any("loai tru" in risk for risk in assessment.risks)


def test_matching_new_opportunity_near_deadline_has_alerts() -> None:
    assessment = assess_notice_feasibility(
        _notice(NOW + timedelta(days=2), first_seen_at=NOW - timedelta(hours=1)),
        ("network cable",),
        [],
        [_inventory()],
        now=NOW,
    )

    assert assessment.alerts == ("NEW_MATCH", "CLOSING_SOON")


def test_ranked_report_contains_explanations_and_safe_text(tmp_path: Path) -> None:
    notice = _notice(
        NOW + timedelta(days=10),
        title='=HYPERLINK("https://evil.test","open") network cable',
    )
    assessment = assess_notice_feasibility(
        notice,
        ("network cable",),
        [],
        [_inventory()],
        now=NOW,
    )

    output = _write_ranked_report(tmp_path / "ranked.xlsx", [(notice, assessment)])
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        sheet = workbook["Co hoi xep hang"]
        headers = [cell.value for cell in sheet[1]]
        assert "score_explanation" in headers
        assert "missing_data" in headers
        assert "next_action" in headers
        title_column = headers.index("title") + 1
        assert str(sheet.cell(2, title_column).value).startswith("'=")
    finally:
        workbook.close()
