from __future__ import annotations

import ast
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from qi_crawler.market_intelligence.khmt_contract import OBSERVED_KHMT_HEADERS
from qi_crawler.market_intelligence.opportunity_contract import OpportunitySourceType
from qi_crawler.market_intelligence.opportunity_review import (
    OpportunityReviewDecision,
    OpportunityReviewRecord,
    OpportunityReviewService,
)
from qi_crawler.market_intelligence.search import TargetedSearchRequest
from qi_crawler.market_intelligence.tbmt_schema import OBSERVED_TBMT_HEADERS


class MemoryReviewRepository:
    def __init__(self) -> None:
        self.records: list[OpportunityReviewRecord] = []

    def latest(self, observation_key: str) -> OpportunityReviewRecord | None:
        matches = [
            record
            for record in self.records
            if record.identity.observation_key == observation_key
        ]
        return max(matches, key=lambda record: record.event_id, default=None)

    def history(self, observation_key: str) -> tuple[OpportunityReviewRecord, ...]:
        return tuple(
            record
            for record in self.records
            if record.identity.observation_key == observation_key
        )

    def append(self, write) -> OpportunityReviewRecord:
        record = OpportunityReviewRecord(
            event_id=len(self.records) + 1,
            identity=write.identity,
            decision=write.decision,
            reviewer=write.reviewer,
            note=write.note,
            opportunity_snapshot_json=write.opportunity_snapshot_json,
            snapshot_schema_version=write.snapshot_schema_version,
            created_at=datetime(2026, 8, 25, tzinfo=UTC),
        )
        self.records.append(record)
        return record

    def latest_for_keys(self, observation_keys: tuple[str, ...]):
        return {
            key: record
            for key in observation_keys
            if (record := self.latest(key)) is not None
        }


def _write_khmt(path: Path, *, plan_id: str = "PL2600000001-00") -> Path:
    values = {
        "GÓI TIN": "Synthetic bulletin",
        "SỐ KẾ HOẠCH": plan_id,
        "TÊN DỰ ÁN": "Synthetic project",
        "TÊN CHỦ ĐẦU TƯ": "Synthetic investor",
        "TỔNG MỨC ĐẦU TƯ": "900.000.000",
        "NỘI DUNG PHÊ DUYỆT": "Approved synthetic scope",
        "TÊN GÓI THẦU": "Synthetic package",
        "NGUỒN VỐN": "Synthetic fund",
        "GIÁ GÓI THẦU": "477.850.000",
        "HÌNH THỨC LỰA CHỌN": "Đấu thầu rộng rãi",
        "HÌNH THỨC HỢP ĐỒNG": "Trọn gói",
        "THỜI GIAN LỰA CHỌN": "Q4/2026",
        "THỜI GIAN THỰC HIỆN": "45 ngày",
    }
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "KHMT"
    sheet.append(["Synthetic KHMT"])
    sheet.append([])
    sheet.append(list(OBSERVED_KHMT_HEADERS))
    sheet.append([values.get(header) for header in OBSERVED_KHMT_HEADERS])
    workbook.save(path)
    return path


def _write_tbmt(path: Path, *, packages: tuple[str, ...], price: object = "1.234.567") -> Path:
    values = {
        "GÓI THẦU": None,
        "BÊN MỜI THẦU": "Synthetic procuring entity",
        "ĐỊA CHỈ BÊN MỜI THẦU": "Synthetic address",
        "DỰ ÁN": "Synthetic project",
        "NỘI DUNG CHÍNH CỦA GÓI THẦU": "Synthetic scope",
        "NGUỒN VỐN": "Synthetic fund",
        "GIÁ GÓI THẦU": price,
        "PHƯƠNG THỨC LỰA CHỌN NHÀ THẦU": "Một giai đoạn",
        "HÌNH THỨC LỰA CHỌN NHÀ THẦU": "Đấu thầu rộng rãi",
        "THỜI GIAN PHÁT HÀNH HSMT": "01/08/2026",
        "GIÁ BÁN 1 BỘ HSMT": 23,
        "BẢO ĐẢM DỰ THẦU": "23",
        "HÌNH THỨC BẢO ĐẢM DỰ THẦU": "Tiền mặt",
        "ĐỊA ĐIỂM PHÁT HÀNH": "Synthetic location",
        "THỜI GIAN ĐÓNG THẦU(HẠN CUỐI TIẾP NHẬN BG)": "10/08/2026",
        "THỜI GIAN MỞ THẦU": "10/08/2026",
        "THỜI GIAN THỰC HIỆN HỢP ĐỒNG": "120 ngày",
    }
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TBMT"
    for column, header in enumerate(OBSERVED_TBMT_HEADERS, start=1):
        sheet.cell(row=10, column=column, value=header)
    for row_number, package in enumerate(packages, start=11):
        values["GÓI THẦU"] = package
        for column, header in enumerate(OBSERVED_TBMT_HEADERS, start=1):
            sheet.cell(row=row_number, column=column, value=values.get(header))
    workbook.save(path)
    return path


@pytest.fixture
def service():
    from qi_crawler.market_intelligence.opportunity_intelligence import (
        OpportunityIntelligenceService,
    )

    return OpportunityIntelligenceService(OpportunityReviewService(MemoryReviewRepository()))


def test_routes_khmt_and_tbmt_to_common_radar_items(tmp_path: Path, service) -> None:
    khmt_path = _write_khmt(tmp_path / "khmt.xlsx")
    tbmt_path = _write_tbmt(
        tmp_path / "tbmt.xlsx", packages=("Gói mẫu IB2600463290-00",)
    )
    khmt = service.load_workbook(khmt_path, OpportunitySourceType.KHMT)
    tbmt = service.load_workbook(
        tbmt_path,
        OpportunitySourceType.TBMT,
    )

    assert khmt.source_type is OpportunitySourceType.KHMT
    assert tbmt.source_type is OpportunitySourceType.TBMT
    assert khmt.items[0].identity.namespace.value == "PL"
    assert tbmt.items[0].identity.namespace.value == "IB"
    assert khmt.source_sha256 == khmt.items[0].source_sha256
    assert tbmt.source_row_count == 1
    assert khmt.sheet == "KHMT"
    assert tbmt.sheet == "TBMT"
    assert khmt.source_path == khmt_path.resolve()
    assert tbmt.source_filename == tbmt_path.name


def test_normalizes_import_issues_without_losing_coordinates(tmp_path: Path, service) -> None:
    path = _write_khmt(tmp_path / "issues.xlsx", plan_id="not-a-plan")

    result = service.load_workbook(path, "KHMT")

    assert len(result.issues) == 1
    issue = result.issues[0]
    assert issue.code == "INVALID_PLAN_ID"
    assert issue.source_row == 4
    assert issue.source_field == "SỐ KẾ HOẠCH"
    assert issue.message == "Plan ID is missing or malformed"


def test_search_delegates_and_keeps_indeterminate_bucket(tmp_path: Path, service) -> None:
    loaded = service.load_workbook(
        _write_tbmt(tmp_path / "unknown.xlsx", packages=("Gói mẫu IB2600463290-00",), price=None),
        OpportunitySourceType.TBMT,
    )

    result = service.search_opportunities(
        loaded.items,
        TargetedSearchRequest(min_budget=Decimal(1)),
    )

    assert result.total_examined == 1
    assert result.indeterminate_count == 1
    assert service.current_review(loaded.items[0]) is None
    assert service.current_confirmed(loaded.items) == ()


def test_review_is_explicit_and_revision_isolation_is_preserved(tmp_path: Path, service) -> None:
    loaded = service.load_workbook(
        _write_tbmt(
            tmp_path / "revisions.xlsx",
            packages=(
                "Gói lần đầu IB2600463290-00",
                "Gói cập nhật IB2600463290-01",
            ),
        ),
        OpportunitySourceType.TBMT,
    )
    first, second = loaded.items

    event = service.record_review(
        first,
        decision=OpportunityReviewDecision.CONFIRMED,
        reviewer="Team Bid",
    )

    assert service.current_review(first) == event
    assert service.current_review(second) is None
    assert service.current_confirmed(loaded.items) == (event,)


def test_changed_source_sha_does_not_inherit_review(tmp_path: Path, service) -> None:
    first = service.load_workbook(
        _write_tbmt(tmp_path / "same-name.xlsx", packages=("Gói A IB2600463290-00",)),
        OpportunitySourceType.TBMT,
    )
    service.record_review(
        first.items[0],
        decision="CONFIRMED",
        reviewer="Team Bid",
    )
    second = service.load_workbook(
        _write_tbmt(tmp_path / "same-name-replaced.xlsx", packages=("Gói B IB2600463290-00",)),
        OpportunitySourceType.TBMT,
    )

    assert first.items[0].source_row == second.items[0].source_row
    assert first.source_sha256 != second.source_sha256
    assert first.items[0].observation_key != second.items[0].observation_key
    assert service.current_review(second.items[0]) is None


def test_facade_has_no_delivery_or_persistence_imports() -> None:
    import qi_crawler.market_intelligence.opportunity_intelligence as facade

    tree = ast.parse(Path(facade.__file__).read_text(encoding="utf-8"))
    imported: list[str] = []
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            imported.extend(alias.name for alias in node.names)
        elif isinstance(node, ast.ImportFrom) and node.module:
            imported.append(node.module)
    forbidden = (
        "PySide",
        "PyQt",
        "qi_crawler.gui",
        "qi_crawler.gui_services",
        "qi_crawler.api",
        "qi_crawler.cli",
        "sqlalchemy",
        "qi_crawler.db",
        "qi_crawler.opportunity_review_persistence",
    )
    assert not any(module == item or module.startswith(item + ".") for module in imported for item in forbidden)
