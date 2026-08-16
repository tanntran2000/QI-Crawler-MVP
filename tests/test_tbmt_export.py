from __future__ import annotations

import asyncio
import hashlib
from datetime import UTC, datetime, timedelta

import pytest
from openpyxl import load_workbook

from qi_crawler import __version__
from qi_crawler.config import AppConfig
from qi_crawler.crawler import CrawlerService
from qi_crawler.db import Database
from qi_crawler.export import tbmt_excel
from qi_crawler.export.tbmt_excel import (
    LATEST_REPORT_NAME,
    TEMPLATE_PATH,
    ReportOutputError,
    TBMTExportResult,
    export_tbmt,
)
from qi_crawler.export.tbmt_mapper import TBMTExcelMapper
from qi_crawler.export.tbmt_schema import META_SHEET_NAME, SHEET_NAME, TBMT_COLUMNS
from qi_crawler.models import Notice
from qi_crawler.parser import ParsedNotice


def _database(tmp_path) -> Database:
    db = Database(f"sqlite:///{tmp_path / 'tbmt.db'}")
    db.create_all()
    return db


def _assert_reconciled(result: TBMTExportResult) -> None:
    assert result.mapped_records == (
        result.exported_records
        + result.rejected_records
        + result.deduplicated_records
        + result.skipped_records
    )


def test_tbmt_export_uses_stable_schema_and_typed_values(tmp_path) -> None:
    db = _database(tmp_path)
    with db.session() as session:
        session.add(
            Notice(
                source_url="https://example.test/tender/IB260001",
                url_hash="a" * 64,
                content_hash="b" * 64,
                notice_code="IB260001",
                title="Cung cấp cáp &amp; thiết bị mạng",
                package_description="Cáp quang\u00a0và router",
                buyer="QI Technologies",
                procuring_entity_address="123 Lãnh Binh Thăng",
                project_name="Nâng cấp hạ tầng",
                funding_source="Vốn doanh nghiệp",
                package_price=1_250_000_000,
                selection_method="Một giai đoạn một túi hồ sơ",
                selection_form="Đấu thầu rộng rãi",
                document_issue_at=datetime(2026, 8, 10, 8, 0, tzinfo=UTC),
                document_price=330_000,
                bid_security_amount=25_000_000,
                bid_security_method="Thư bảo lãnh",
                issue_location="Hệ thống e-GP",
                published_at="10/08/2026 07:30",
                published_at_dt=datetime(2026, 8, 10, 7, 30, tzinfo=UTC),
                closing_at_dt=datetime(2026, 8, 20, 9, 0, tzinfo=UTC),
                bid_open_at=datetime(2026, 8, 20, 9, 15, tzinfo=UTC),
                contract_duration="120 ngày",
                crawl_run_id=7,
                review_status="approved",
            )
        )

    before_hash = hashlib.sha256(TEMPLATE_PATH.read_bytes()).hexdigest()
    result = export_tbmt(
        db,
        output=tmp_path / "TBMT.xlsx",
        rejects_dir=tmp_path / "rejects",
        latest_run_only=False,
    )
    after_hash = hashlib.sha256(TEMPLATE_PATH.read_bytes()).hexdigest()

    assert before_hash == after_hash
    assert result.exported_records == 1
    workbook = load_workbook(result.output)
    assert workbook.sheetnames == [SHEET_NAME, META_SHEET_NAME]
    assert workbook[META_SHEET_NAME].sheet_state == "hidden"
    sheet = workbook[SHEET_NAME]
    assert tuple(sheet.cell(10, column).value for column in range(1, 19)) == TBMT_COLUMNS
    assert "Cung cấp cáp & thiết bị mạng" in sheet["E11"].value
    assert "Số thông báo: IB260001" in sheet["E11"].value
    assert sheet["F11"].value == "Cáp quang và router"
    assert sheet["H11"].value == 1_250_000_000
    assert isinstance(sheet["K11"].value, datetime)
    assert isinstance(sheet["P11"].value, datetime)
    assert sheet.freeze_panes == "A11"
    assert sheet.auto_filter.ref == "A10:R11"
    assert sheet["O11"].hyperlink.target == "https://example.test/tender/IB260001"
    assert workbook[META_SHEET_NAME]["B3"].value == __version__
    _assert_reconciled(result)


def test_tbmt_export_logs_completion_stages(tmp_path, caplog) -> None:
    db = _database(tmp_path)
    caplog.set_level("INFO", logger="qi_crawler.export.tbmt_excel")

    export_tbmt(
        db,
        output=tmp_path / "TBMT-stage-log.xlsx",
        rejects_dir=tmp_path / "rejects",
    )

    messages = "\n".join(caplog.messages)
    for stage in (
        "EXPORT_START",
        "DB_QUERY_DONE",
        "MAPPING_DONE",
        "VALIDATION_DONE",
        "WORKBOOK_SAVE_START",
        "WORKBOOK_SAVE_DONE",
        "EXPORT_DONE",
    ):
        assert stage in messages


def test_date_only_deadline_is_exported_without_a_fake_midnight_time(tmp_path) -> None:
    db = _database(tmp_path)
    with db.session() as session:
        session.add(
            Notice(
                source_url="https://example.test/tender/IB260002",
                url_hash="d" * 64,
                notice_code="IB260002",
                title="Gói chỉ công bố ngày đóng thầu",
                closing_at="14/08/2026",
                closing_at_dt=datetime(2026, 8, 14, 0, 0, tzinfo=UTC),
            )
        )

    result = export_tbmt(
        db,
        output=tmp_path / "TBMT-date-only.xlsx",
        rejects_dir=tmp_path / "rejects",
        latest_run_only=False,
    )

    sheet = load_workbook(result.output)[SHEET_NAME]
    assert sheet["P11"].number_format == "dd/mm/yyyy"
    assert sheet["P11"].value.date().isoformat() == "2026-08-14"


def test_description_is_preferred_over_package_name() -> None:
    notice = Notice(
        source_url="https://example.test/tender/IB260003",
        url_hash="e" * 64,
        notice_code="IB260003",
        title="Tên gói ngắn",
        package_description="Phạm vi công việc: cung cấp và lắp đặt thiết bị.",
    )

    mapped = TBMTExcelMapper().map(TBMTExcelMapper().normalize(notice), index=1)

    assert mapped.values["NỘI DUNG CHÍNH CỦA GÓI THẦU"] == (
        "Phạm vi công việc: cung cấp và lắp đặt thiết bị."
    )


def test_tbmt_export_rejects_invalid_records_and_replaces_explicit_output(tmp_path) -> None:
    db = _database(tmp_path)
    with db.session() as session:
        session.add(
            Notice(
                source_url="https://example.test/invalid",
                url_hash="c" * 64,
                title="Thiếu mã thông báo",
            )
        )

    first = export_tbmt(
        db,
        output=tmp_path / "TBMT.xlsx",
        rejects_dir=tmp_path / "rejects",
        latest_run_only=False,
    )
    second = export_tbmt(
        db,
        output=tmp_path / "TBMT.xlsx",
        rejects_dir=tmp_path / "rejects",
        latest_run_only=False,
    )

    assert first.output.name == "TBMT.xlsx"
    assert second.output.name == "TBMT.xlsx"
    assert first.rejected_records == 1
    assert first.reject_output is not None
    assert first.reject_output.exists()
    assert first.mapped_records == 1
    assert first.exported_records == 0
    assert first.rejected_records == 1
    assert first.deduplicated_records == 0
    assert first.skipped_records == 0
    _assert_reconciled(first)
    workbook = load_workbook(first.output)
    assert workbook[SHEET_NAME].max_row == 10


def test_default_export_replaces_latest_without_versioned_clutter(tmp_path) -> None:
    db = _database(tmp_path)
    legacy = tmp_path / "reports" / "TBMT_1_2_2026.xlsx"
    legacy.parent.mkdir(parents=True)
    legacy.write_bytes(b"legacy report")
    with db.session() as session:
        session.add(
            Notice(
                source_url="https://example.test/latest",
                url_hash="l" * 64,
                notice_code="IB-LATEST",
                title="Bản đầu tiên",
            )
        )

    first = export_tbmt(db, report_dir=legacy.parent, rejects_dir=tmp_path / "rejects")
    with db.session() as session:
        session.get(Notice, 1).title = "Bản đã cập nhật"
    second = export_tbmt(db, report_dir=legacy.parent, rejects_dir=tmp_path / "rejects")

    assert first.output == legacy.parent / LATEST_REPORT_NAME
    assert second.output == first.output
    assert not list(legacy.parent.glob("TBMT_Latest_v*.xlsx"))
    assert legacy.read_bytes() == b"legacy report"
    meta = load_workbook(second.output, read_only=True)[META_SHEET_NAME]
    assert meta["B3"].value == __version__
    assert "current_day_only=False" in meta["B7"].value


def test_latest_output_preserves_existing_report_when_atomic_replace_fails(
    tmp_path,
    monkeypatch,
) -> None:
    db = _database(tmp_path)
    reports = tmp_path / "reports"
    reports.mkdir()
    latest = reports / LATEST_REPORT_NAME
    latest.write_bytes(b"keep previous latest")

    monkeypatch.setattr(tbmt_excel.os, "replace", lambda *_args: (_ for _ in ()).throw(OSError("locked")))

    with pytest.raises(ReportOutputError, match="đang được mở"):
        export_tbmt(db, report_dir=reports, rejects_dir=tmp_path / "rejects")

    assert latest.read_bytes() == b"keep previous latest"
    assert not list(reports.glob(f".{latest.stem}.*.xlsx"))


def test_snapshot_exports_to_archive_with_collision_safe_name(tmp_path) -> None:
    db = _database(tmp_path)
    generated_at = datetime(2026, 8, 16, 10, 30, tzinfo=UTC)

    first = export_tbmt(
        db,
        report_dir=tmp_path / "reports",
        rejects_dir=tmp_path / "rejects",
        snapshot=True,
        generated_at=generated_at,
    )
    second = export_tbmt(
        db,
        report_dir=tmp_path / "reports",
        rejects_dir=tmp_path / "rejects",
        snapshot=True,
        generated_at=generated_at,
    )

    archive = tmp_path / "reports" / "archive" / "2026" / "2026-08"
    assert first.output.parent == archive
    assert second.output.parent == archive
    assert first.output.name.startswith("TBMT_2026-08-16_1030_run_unassigned_")
    assert first.output != second.output
    assert load_workbook(first.output, read_only=True)[META_SHEET_NAME]["B2"].value.startswith(
        "2026-08-16T10:30"
    )


def test_duplicate_record_is_explicitly_accounted(tmp_path, caplog) -> None:
    db = _database(tmp_path)
    with db.session() as session:
        for index in (1, 2):
            session.add(
                Notice(
                    source_url=f"https://example.test/tender/duplicate-{index}",
                    url_hash=str(index) * 64,
                    notice_code="IB-DUPLICATE",
                    title=f"Gói trùng phiên bản {index}",
                    last_seen_at=datetime(2026, 8, 13, index, tzinfo=UTC),
                )
            )
    caplog.set_level("INFO", logger="qi_crawler.export.tbmt_excel")

    result = export_tbmt(
        db,
        output=tmp_path / "duplicate.xlsx",
        rejects_dir=tmp_path / "rejects",
    )

    assert result.queried_records == 2
    assert result.mapped_records == 2
    assert result.exported_records == 1
    assert result.deduplicated_records == 1
    assert result.rejected_records == 0
    assert result.skipped_records == 0
    assert "EXPORT_RECORD_DEDUPLICATED" in "\n".join(caplog.messages)
    assert "IB-DUPLICATE" in "\n".join(caplog.messages)
    _assert_reconciled(result)


def test_filtered_record_is_skipped_with_explicit_reason(tmp_path, caplog) -> None:
    db = _database(tmp_path)
    yesterday = datetime.now(UTC) - timedelta(days=1)
    with db.session() as session:
        session.add(
            Notice(
                source_url="https://ebidding.coteccons.vn/Index/ChiTiet/2604255",
                url_hash="f" * 64,
                source_name="coteccons",
                source_notice_id="2604255",
                title="Gói từ ngày hôm trước",
                last_seen_at=yesterday,
            )
        )
    caplog.set_level("INFO", logger="qi_crawler.export.tbmt_excel")

    result = export_tbmt(
        db,
        output=tmp_path / "today-only.xlsx",
        rejects_dir=tmp_path / "rejects",
        current_day_only=True,
    )

    assert result.mapped_records == 1
    assert result.exported_records == 0
    assert result.skipped_records == 1
    assert result.deduplicated_records == 0
    assert result.rejected_records == 0
    messages = "\n".join(caplog.messages)
    assert "EXPORT_RECORD_SKIPPED" in messages
    assert "source_notice_id=2604255" in messages
    assert "reason=crawl_date_not_equal:" in messages
    _assert_reconciled(result)


def test_three_separate_runs_export_three_rows_by_default(tmp_path) -> None:
    db = _database(tmp_path)
    with db.session() as session:
        for index, run_id in enumerate((11, 12, 13), start=1):
            session.add(
                Notice(
                    source_url=f"https://ebidding.coteccons.vn/Index/ChiTiet/260730{index}",
                    url_hash=str(index) * 64,
                    source_name="coteccons",
                    source_notice_id=f"260730{index}",
                    title=f"Goi thau {index}",
                    crawl_run_id=run_id,
                )
            )

    result = export_tbmt(db, output=tmp_path / "today.xlsx", rejects_dir=tmp_path / "rejects")
    workbook = load_workbook(result.output)

    assert result.exported_records == 3
    assert workbook[SHEET_NAME].max_row == 13

    one_run = export_tbmt(
        db,
        output=tmp_path / "run-12.xlsx",
        rejects_dir=tmp_path / "rejects",
        crawl_run_id=12,
    )
    assert one_run.exported_records == 1


def test_duplicate_rerun_still_exports_three_rows(tmp_path) -> None:
    config = AppConfig()
    config.storage.database_url = f"sqlite:///{tmp_path / 'rerun.db'}"
    service = CrawlerService(config)
    try:
        for index in range(1, 4):
            parsed = ParsedNotice(
                source_url=f"https://ebidding.coteccons.vn/Index/ChiTiet/260730{index}",
                source_name="coteccons",
                source_notice_id=f"260730{index}",
                title=f"Goi thau {index}",
            )
            service.upsert_parsed_notice(parsed, crawl_run_id=index)
            service.upsert_parsed_notice(parsed, crawl_run_id=index + 10)

        result = export_tbmt(
            service.db,
            output=tmp_path / "rerun.xlsx",
            rejects_dir=tmp_path / "rejects",
        )
        assert result.exported_records == 3
        with service.db.session() as session:
            assert session.query(Notice).count() == 3
    finally:
        asyncio.run(service.close())
