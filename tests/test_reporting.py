from datetime import UTC, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from qi_crawler.db import Database
from qi_crawler.models import Notice
from qi_crawler.reporting import build_daily_report


def test_build_daily_report(tmp_path: Path):
    db = Database(f"sqlite:///{tmp_path / 'report.db'}")
    db.create_all()
    with db.session() as session:
        session.add(
            Notice(
                source_url="import://sample#1",
                url_hash="a" * 64,
                notice_code="IB260001",
                title="Mua thiet bi mang",
                closing_at=(datetime.now(UTC).date() + timedelta(days=2)).isoformat(),
            )
        )
    output = build_daily_report(db, tmp_path / "daily.xlsx", days_ahead=7)
    workbook = load_workbook(output, read_only=True)
    try:
        assert workbook.sheetnames == [
            "Goi thau moi",
            "Sap dong thau",
            "Tat ca goi thau",
            "Tep tai loi",
            "Chat luong du lieu",
        ]
        assert workbook["Sap dong thau"].max_row == 2
    finally:
        workbook.close()
