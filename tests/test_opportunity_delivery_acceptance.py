from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from qi_crawler.config import AppConfig
from qi_crawler.db import Database
from qi_crawler.gui_services import (
    run_bid_radar_export,
    run_bid_radar_import_search,
    run_bid_radar_legal_docx,
    run_bid_radar_review,
)
from qi_crawler.market_intelligence.confirmed_opportunity_export import (
    DEFAULT_CONFIRMED_OPPORTUNITY_EXPORT_FILENAME,
)
from qi_crawler.market_intelligence.khmt_contract import OBSERVED_KHMT_HEADERS
from qi_crawler.market_intelligence.opportunity_contract import OpportunitySourceType
from qi_crawler.market_intelligence.opportunity_intelligence import (
    OpportunityIntelligenceService,
)
from qi_crawler.market_intelligence.opportunity_review import OpportunityReviewService
from qi_crawler.market_intelligence.search import TargetedSearchRequest
from qi_crawler.market_intelligence.source_detection import SourceType
from qi_crawler.market_intelligence.source_integrity import OpportunitySourceIntegrityError
from qi_crawler.market_intelligence.tbmt_schema import OBSERVED_TBMT_HEADERS
from qi_crawler.opportunity_review_persistence import SqlAlchemyOpportunityReviewRepository


def _config(tmp_path: Path) -> AppConfig:
    config = AppConfig()
    config.storage.database_url = f"sqlite:///{tmp_path / 'acceptance.db'}"
    config.storage.report_dir = tmp_path / "reports"
    return config


def _write_khmt(path: Path, *, plan_id: str = "PL2600000001-00") -> Path:
    values = {
        "GÓI TIN": "Synthetic KHMT bulletin",
        "SỐ KẾ HOẠCH": plan_id,
        "TÊN DỰ ÁN": "Synthetic KHMT project",
        "TÊN CHỦ ĐẦU TƯ": "Synthetic investor",
        "TỔNG MỨC ĐẦU TƯ": "900.000.000",
        "NỘI DUNG PHÊ DUYỆT": "Synthetic approved scope",
        "TÊN GÓI THẦU": "Synthetic KHMT package",
        "NGUỒN VỐN": "Synthetic fund",
        "GIÁ GÓI THẦU": "477.850.000",
        "HÌNH THỨC LỰA CHỌN": "Đấu thầu rộng rãi",
        "HÌNH THỨC HỢP ĐỒNG": "Trọn gói",
        "THỜI GIAN LỰA CHỌN": "Q4/2026",
        "THỜI GIAN THỰC HIỆN": "45 ngày",
    }
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "KHMT"
    sheet.append(OBSERVED_KHMT_HEADERS)
    sheet.append([values.get(header) for header in OBSERVED_KHMT_HEADERS])
    workbook.save(path)
    return path


def _write_tbmt(path: Path, rows: tuple[dict[str, object], ...]) -> Path:
    common: dict[str, object] = {
        "BÊN MỜI THẦU": "Synthetic procuring entity",
        "ĐỊA CHỈ BÊN MỜI THẦU": "Synthetic address",
        "DỰ ÁN": "Synthetic project",
        "NỘI DUNG CHÍNH CỦA GÓI THẦU": "Synthetic scope",
        "NGUỒN VỐN": "Synthetic fund",
        "PHƯƠNG THỨC LỰA CHỌN NHÀ THẦU": "Một giai đoạn",
        "HÌNH THỨC LỰA CHỌN NHÀ THẦU": "Đấu thầu rộng rãi",
        "THỜI GIAN PHÁT HÀNH HSMT": "01/08/2026",
        "GIÁ BÁN 1 BỘ HSMT": 23,
        "BẢO ĐẢM DỰ THẦU": "23",
        "HÌNH THỨC BẢO ĐẢM DỰ THẦU": "Tiền mặt",
        "ĐỊA ĐIỂM PHÁT HÀNH": "Synthetic location",
        "THỜI GIAN ĐÓNG THẦU(HẠN CUỐI TIẾP NHẬN BG)": "10/08/2026",
        "THỜI GIAN MỞ THẦU": "10/08/2026",
        "THỜI GIAN THỰC HIỆN HỢP ĐỒNG": "120 ngày",
    }
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TBMT"
    sheet.append(OBSERVED_TBMT_HEADERS)
    for row in rows:
        values = {
            **common,
            "GÓI THẦU": row["package"],
            "GIÁ GÓI THẦU": row.get("price"),
        }
        sheet.append([values.get(header) for header in OBSERVED_TBMT_HEADERS])
    workbook.save(path)
    return path


def _restart_service(config: AppConfig) -> OpportunityIntelligenceService:
    database = Database(config.storage.database_url)
    repository = SqlAlchemyOpportunityReviewRepository(database)
    return OpportunityIntelligenceService(OpportunityReviewService(repository))


def _export_rows(path: Path) -> tuple[dict[str, object], ...]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    headers = tuple(rows[0])
    return tuple(dict(zip(headers, row, strict=True)) for row in rows[1:])


def test_khmt_vertical_review_restart_confirmed_export_and_legal_docx(
    tmp_path: Path,
) -> None:
    config = _config(tmp_path)
    source = _write_khmt(tmp_path / "KHMT-synthetic.xlsx")
    loaded = run_bid_radar_import_search(
        config,
        source,
        TargetedSearchRequest(include_keywords=("Synthetic",)),
        source_type=SourceType.KHMT,
    )

    assert loaded.source_type is OpportunitySourceType.KHMT
    assert loaded.items[0].identity.namespace.value == "PL"
    assert loaded.matched_count == 1
    assert loaded.rows[0].disposition.value == "MATCH"
    assert loaded.rows[0].review_state == "UNREVIEWED"

    assert run_bid_radar_review(config, loaded.items[0], "CONFIRMED", "Team Bid") == "CONFIRMED"
    restarted = _restart_service(config)
    reloaded = restarted.load_workbook(source, OpportunitySourceType.KHMT)
    confirmed = restarted.current_confirmed(reloaded.items)
    assert len(confirmed) == 1
    assert confirmed[0].decision.value == "CONFIRMED"

    export = run_bid_radar_export(
        config,
        loaded.load_result,
        source_path=source,
        expected_source_sha256=loaded.source_sha256,
    )
    assert export.exported_rows == 1
    row = _export_rows(export.output)[0]
    assert row["NAMESPACE"] == "PL"
    assert row["MÃ CƠ HỘI"] == "PL2600000001-00"
    assert row["FILE NGUỒN"] == source.name
    assert row["SHA-256 NGUỒN"] == loaded.source_sha256
    assert row["SHEET NGUỒN"] == "KHMT"
    assert row["DÒNG NGUỒN"] == 2

    legal = run_bid_radar_legal_docx(
        config,
        loaded.load_result,
        source_path=source,
        expected_source_sha256=loaded.source_sha256,
    )
    assert len(legal) == 1
    assert legal[0].output.name == "ThongTin_PL2600000001.docx"

    assert run_bid_radar_review(config, loaded.items[0], "REJECTED", "Team Bid") == "REJECTED"
    latest_export = run_bid_radar_export(
        config,
        loaded.load_result,
        source_path=source,
        expected_source_sha256=loaded.source_sha256,
    )
    assert latest_export.exported_rows == 0
    assert _export_rows(latest_export.output) == ()


def test_tbmt_vertical_tri_state_revision_review_restart_export_and_sha_guard(
    tmp_path: Path,
) -> None:
    config = _config(tmp_path)
    source = _write_tbmt(
        tmp_path / "TBMT-synthetic.xlsx",
        (
            {"package": "Target package IB2600463290-00", "price": "1000"},
            {"package": "Other package IB2600463290-01", "price": "1000"},
            {"package": "Target unknown IB2600463290-02", "price": None},
        ),
    )
    loaded = run_bid_radar_import_search(
        config,
        source,
        TargetedSearchRequest(min_budget=Decimal(1), include_keywords=("Target",)),
        source_type=SourceType.TBMT,
    )

    assert loaded.source_type is OpportunitySourceType.TBMT
    assert {item.identity.namespace.value for item in loaded.items} == {"IB"}
    assert all(not hasattr(item, "plan") for item in loaded.items)
    assert loaded.matched_count == 1
    assert loaded.nonmatched_count == 1
    assert loaded.indeterminate_count == 1
    dispositions = {row.item.identity.revision: row.disposition.value for row in loaded.rows}
    assert dispositions == {"00": "MATCH", "01": "NO_MATCH", "02": "INDETERMINATE"}
    assert loaded.rows[0].review_state == "UNREVIEWED"

    assert run_bid_radar_review(config, loaded.items[0], "CONFIRMED", "Team Bid") == "CONFIRMED"
    restarted = _restart_service(config)
    reloaded = restarted.load_workbook(source, OpportunitySourceType.TBMT)
    confirmed = restarted.current_confirmed(reloaded.items)
    assert [record.identity.identity_revision for record in confirmed] == ["00"]
    assert len(restarted.review_service.list_history(reloaded.items[0])) == 1

    export = run_bid_radar_export(
        config,
        loaded.load_result,
        source_path=source,
        expected_source_sha256=loaded.source_sha256,
    )
    assert export.exported_rows == 1
    rows = _export_rows(export.output)
    assert rows[0]["LOẠI NGUỒN"] == "TBMT"
    assert rows[0]["NAMESPACE"] == "IB"
    assert rows[0]["MÃ CƠ HỘI"] == "IB2600463290-00"
    assert rows[0]["MÃ CƠ HỘI"] != "IB2600463290-01"

    with pytest.raises(ValueError, match="chỉ hỗ trợ nguồn KHMT"):
        run_bid_radar_legal_docx(
            config,
            loaded.load_result,
            source_path=source,
            expected_source_sha256=loaded.source_sha256,
        )
    assert not tuple(config.storage.report_dir.glob("ThongTin_*.docx"))

    output_before = export.output.read_bytes()
    source.write_bytes(b"mutated after import")
    with pytest.raises(OpportunitySourceIntegrityError):
        run_bid_radar_export(
            config,
            loaded.load_result,
            source_path=source,
            expected_source_sha256=loaded.source_sha256,
        )
    assert export.output.read_bytes() == output_before
    assert export.output.name == DEFAULT_CONFIRMED_OPPORTUNITY_EXPORT_FILENAME
