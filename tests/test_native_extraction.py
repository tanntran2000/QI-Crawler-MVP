from __future__ import annotations

import json
import zipfile
from hashlib import sha256
from pathlib import Path

import pytest
from openpyxl import Workbook
from pypdf import PdfWriter
from pypdf.generic import DecodedStreamObject, DictionaryObject, NameObject, NumberObject
from sqlalchemy import select

from qi_crawler.db import Database
from qi_crawler.document_intake import DocumentIntakeService
from qi_crawler.models import DocumentEvidence, DocumentExtraction, Notice
from qi_crawler.native_extraction import (
    EMPTY_PAGE,
    NATIVE_OK,
    NEEDS_OCR,
    NEEDS_REVIEW,
    TABLE_STRUCTURE_UNCERTAIN,
    TEXT_ENCODING_SUSPECT,
    NativeExtractionError,
    NativeHSMTExtractionService,
)


@pytest.fixture
def services(tmp_path: Path) -> tuple[DocumentIntakeService, NativeHSMTExtractionService, Database]:
    database = Database(f"sqlite:///{tmp_path / 'extraction.db'}")
    return (
        DocumentIntakeService(database, tmp_path / "documents"),
        NativeHSMTExtractionService(database),
        database,
    )


def _pdf(path: Path) -> Path:
    writer = PdfWriter()
    writer.add_blank_page(width=144, height=144)
    with path.open("wb") as stream:
        writer.write(stream)
    return path


def _text_pdf(path: Path, text: str = "Tender technical requirement") -> Path:
    writer = PdfWriter()
    page = writer.add_blank_page(width=144, height=144)
    font = DictionaryObject(
        {
            NameObject("/Type"): NameObject("/Font"),
            NameObject("/Subtype"): NameObject("/Type1"),
            NameObject("/BaseFont"): NameObject("/Helvetica"),
        }
    )
    font_ref = writer._add_object(font)
    page[NameObject("/Resources")] = DictionaryObject(
        {
            NameObject("/Font"): DictionaryObject({NameObject("/F1"): font_ref}),
        }
    )
    contents = DecodedStreamObject()
    contents.set_data(f"BT /F1 12 Tf 12 72 Td ({text}) Tj ET".encode("latin-1"))
    page[NameObject("/Contents")] = writer._add_object(contents)
    with path.open("wb") as stream:
        writer.write(stream)
    return path


def _image_pdf(path: Path, *, hybrid: bool = False) -> Path:
    writer = PdfWriter()
    page = writer.add_blank_page(width=144, height=144)
    image = DecodedStreamObject()
    image.set_data(b"\x00")
    image.update(
        {
            NameObject("/Type"): NameObject("/XObject"),
            NameObject("/Subtype"): NameObject("/Image"),
            NameObject("/Width"): NumberObject(1),
            NameObject("/Height"): NumberObject(1),
            NameObject("/ColorSpace"): NameObject("/DeviceGray"),
            NameObject("/BitsPerComponent"): NumberObject(8),
        }
    )
    image_ref = writer._add_object(image)
    resources = DictionaryObject(
        {
            NameObject("/XObject"): DictionaryObject({NameObject("/Im1"): image_ref}),
        }
    )
    if hybrid:
        font = DictionaryObject(
            {
                NameObject("/Type"): NameObject("/Font"),
                NameObject("/Subtype"): NameObject("/Type1"),
                NameObject("/BaseFont"): NameObject("/Helvetica"),
            }
        )
        resources[NameObject("/Font")] = DictionaryObject(
            {NameObject("/F1"): writer._add_object(font)}
        )
        contents = DecodedStreamObject()
        contents.set_data(b"BT /F1 12 Tf 12 72 Td (Readable hybrid text) Tj ET")
        page[NameObject("/Contents")] = writer._add_object(contents)
    page[NameObject("/Resources")] = resources
    with path.open("wb") as stream:
        writer.write(stream)
    return path


def _docx(path: Path) -> Path:
    document = """<?xml version="1.0" encoding="UTF-8"?>
    <w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"><w:body>
      <w:p><w:pPr><w:pStyle w:val="Heading1"/></w:pPr><w:r><w:t>Yêu cầu kỹ thuật</w:t></w:r></w:p>
      <w:p><w:r><w:t>Cáp quang</w:t></w:r></w:p>
      <w:tbl><w:tr><w:tc><w:p><w:r><w:t>Hạng mục</w:t></w:r></w:p></w:tc>
      <w:tc><w:p><w:r><w:t>Số lượng</w:t></w:r></w:p></w:tc></w:tr>
      <w:tr><w:tc><w:p><w:r><w:t>Router</w:t></w:r></w:p></w:tc>
      <w:tc><w:p><w:r><w:t>2</w:t></w:r></w:p></w:tc></w:tr></w:tbl>
    </w:body></w:document>"""
    styles = """<?xml version="1.0" encoding="UTF-8"?>
    <w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
      <w:style w:type="paragraph" w:styleId="Heading1"><w:name w:val="heading 1"/></w:style>
    </w:styles>"""
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr("word/document.xml", document)
        archive.writestr("word/styles.xml", styles)
    return path


def _xlsx(path: Path) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "BOQ"
    worksheet.append(["Hạng mục", "Số lượng"])
    worksheet.append(["Switch", 3])
    workbook.save(path)
    return path


def _merged_docx(path: Path, *, vertical: bool) -> Path:
    merge = "<w:vMerge w:val=\"restart\"/>" if vertical else "<w:gridSpan w:val=\"2\"/>"
    continuation = "<w:vMerge/>" if vertical else ""
    document = f"""<?xml version="1.0" encoding="UTF-8"?>
    <w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"><w:body>
      <w:tbl><w:tr><w:tc><w:tcPr>{merge}</w:tcPr><w:p><w:r><w:t>Gộp</w:t></w:r></w:p></w:tc>
      <w:tc><w:p><w:r><w:t>Giá trị</w:t></w:r></w:p></w:tc></w:tr>
      <w:tr><w:tc><w:tcPr>{continuation}</w:tcPr><w:p/></w:tc>
      <w:tc><w:p><w:r><w:t>2</w:t></w:r></w:p></w:tc></w:tr></w:tbl>
    </w:body></w:document>"""
    styles = """<?xml version="1.0" encoding="UTF-8"?>
    <w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"/>"""
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr("word/document.xml", document)
        archive.writestr("word/styles.xml", styles)
    return path


def _merged_multisheet_xlsx(path: Path) -> Path:
    workbook = Workbook()
    boq = workbook.active
    boq.title = "BOQ"
    boq["A1"] = "Phạm vi"
    boq.merge_cells("A1:B1")
    boq["A2"] = "Switch"
    boq["B2"] = 3
    pricing = workbook.create_sheet("Bảng giá")
    pricing.append(["Hạng mục", "Đơn giá"])
    pricing.append(["Switch", 10])
    workbook.save(path)
    return path


def _add_tender(database: Database, code: str) -> None:
    source_url = f"https://example.test/tender/{code}"
    with database.session() as session:
        session.add(
            Notice(
                source_url=source_url,
                url_hash=sha256(source_url.encode()).hexdigest(),
                notice_code=code,
                source_name="test",
                source_notice_id=code,
                title=f"Tender {code}",
            )
        )


@pytest.mark.parametrize(
    ("filename", "create", "expected_locator"),
    [
        ("hsmt.pdf", _pdf, "page:1"),
        ("hsmt.docx", _docx, "word/document.xml:paragraph:1"),
        ("boq.xlsx", _xlsx, "sheet:BOQ!A1:B1"),
    ],
)
def test_native_extraction_preserves_source_trace(
    services: tuple[DocumentIntakeService, NativeHSMTExtractionService, Database],
    tmp_path: Path,
    filename: str,
    create,
    expected_locator: str,
) -> None:
    intake, extractor, database = services
    source = create(tmp_path / filename)
    original_bytes = source.read_bytes()
    document = intake.intake_file(source)

    result = extractor.extract_document(document.document_id)

    assert result.outcome == "EXTRACTED"
    assert result.evidence_count >= 1
    assert source.read_bytes() == original_bytes
    with database.session() as session:
        extraction = session.get(DocumentExtraction, result.extraction_id)
        evidence = list(
            session.scalars(
                select(DocumentEvidence)
                .where(DocumentEvidence.extraction_id == result.extraction_id)
                .order_by(DocumentEvidence.ordinal)
            )
        )
    assert extraction is not None
    metadata = json.loads(extraction.metadata_json or "{}")
    assert metadata["file_format"] == result.file_format
    assert metadata["source_sha256"] == document.sha256
    assert metadata["document_version"] == 1
    assert metadata["status"] == result.status
    assert evidence[0].source_locator == expected_locator
    assert json.loads(evidence[0].metadata_json or "{}")


def test_docx_heading_and_table_cells_are_preserved(
    services: tuple[DocumentIntakeService, NativeHSMTExtractionService, Database], tmp_path: Path
) -> None:
    intake, extractor, database = services
    document = intake.intake_file(_docx(tmp_path / "hsmt.docx"))
    result = extractor.extract_document(document.document_id)

    with database.session() as session:
        evidence = list(
            session.scalars(
                select(DocumentEvidence)
                .where(DocumentEvidence.extraction_id == result.extraction_id)
                .order_by(DocumentEvidence.ordinal)
            )
        )
    table = next(item for item in evidence if item.content_type == "TABLE")
    assert evidence[0].section_heading == "Yêu cầu kỹ thuật"
    assert json.loads(table.table_json or "[]") == [
        ["Hạng mục", "Số lượng"],
        ["Router", "2"],
    ]


def test_repeat_extraction_is_idempotent(
    services: tuple[DocumentIntakeService, NativeHSMTExtractionService, Database], tmp_path: Path
) -> None:
    intake, extractor, _database = services
    document = intake.intake_file(_xlsx(tmp_path / "boq.xlsx"))

    first = extractor.extract_document(document.document_id)
    second = extractor.extract_document(document.document_id)

    assert second.outcome == "ALREADY_EXTRACTED"
    assert second.extraction_id == first.extraction_id
    assert second.evidence_count == first.evidence_count
    assert second.status == first.status


def test_unsupported_format_is_rejected_without_creating_extraction(
    services: tuple[DocumentIntakeService, NativeHSMTExtractionService, Database], tmp_path: Path
) -> None:
    intake, extractor, database = services
    source = tmp_path / "archive.zip"
    with zipfile.ZipFile(source, "w") as archive:
        archive.writestr("hsmt.pdf", b"not extracted in WP2")
    document = intake.intake_file(source)

    with pytest.raises(NativeExtractionError, match="PDF, DOCX va XLSX"):
        extractor.extract_document(document.document_id)
    with database.session() as session:
        assert session.scalar(select(DocumentExtraction)) is None


@pytest.mark.parametrize(
    ("name", "factory", "expected_flag", "expected_status"),
    [
        ("text.pdf", _text_pdf, NATIVE_OK, NATIVE_OK),
        ("hybrid.pdf", lambda path: _image_pdf(path, hybrid=True), NATIVE_OK, NATIVE_OK),
        ("image.pdf", _image_pdf, NEEDS_OCR, NEEDS_REVIEW),
        ("empty.pdf", _pdf, EMPTY_PAGE, NEEDS_REVIEW),
    ],
)
def test_pdf_flags_fail_closed_with_source_trace(
    services: tuple[DocumentIntakeService, NativeHSMTExtractionService, Database],
    tmp_path: Path,
    name: str,
    factory,
    expected_flag: str,
    expected_status: str,
) -> None:
    intake, extractor, database = services
    document = intake.intake_file(factory(tmp_path / name))

    result = extractor.extract_document(document.document_id)

    with database.session() as session:
        extraction = session.get(DocumentExtraction, result.extraction_id)
        evidence = session.scalar(
            select(DocumentEvidence).where(DocumentEvidence.extraction_id == result.extraction_id)
        )
    assert extraction is not None
    assert evidence is not None
    assert result.status == expected_status
    assert extraction.status == expected_status
    assert expected_flag in json.loads(evidence.metadata_json or "{}")["flags"]
    warnings = json.loads(extraction.metadata_json or "{}")["warnings"]
    if expected_flag == NATIVE_OK:
        assert warnings == []
    else:
        assert warnings == [
            {
                "flag": expected_flag,
                "source_locator": "page:1",
                "page_number": 1,
                "sheet_name": None,
                "section_heading": None,
            }
        ]


def test_suspicious_pdf_text_is_needs_review_with_page_trace(
    services: tuple[DocumentIntakeService, NativeHSMTExtractionService, Database],
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class SuspectPage:
        def extract_text(self) -> str:
            return "Untrusted\ufffdtext"

        def get(self, _key: str) -> None:
            return None

    class SuspectReader:
        def __init__(self, _path: str) -> None:
            self.pages = [SuspectPage()]

    monkeypatch.setattr("qi_crawler.native_extraction.PdfReader", SuspectReader)
    intake, extractor, database = services
    document = intake.intake_file(_text_pdf(tmp_path / "suspect.pdf"))

    result = extractor.extract_document(document.document_id)

    with database.session() as session:
        extraction = session.get(DocumentExtraction, result.extraction_id)
        evidence = session.scalar(
            select(DocumentEvidence).where(DocumentEvidence.extraction_id == result.extraction_id)
        )
    assert extraction is not None
    assert evidence is not None
    assert result.status == NEEDS_REVIEW
    assert TEXT_ENCODING_SUSPECT in json.loads(evidence.metadata_json or "{}")["flags"]
    assert json.loads(extraction.metadata_json or "{}")["warnings"][0]["source_locator"] == "page:1"


@pytest.mark.parametrize("vertical", [False, True])
def test_docx_merged_table_is_preserved_and_marked_uncertain(
    services: tuple[DocumentIntakeService, NativeHSMTExtractionService, Database],
    tmp_path: Path,
    vertical: bool,
) -> None:
    intake, extractor, database = services
    document = intake.intake_file(_merged_docx(tmp_path / f"merged-{vertical}.docx", vertical=vertical))

    result = extractor.extract_document(document.document_id)

    with database.session() as session:
        table = session.scalar(
            select(DocumentEvidence).where(
                DocumentEvidence.extraction_id == result.extraction_id,
                DocumentEvidence.content_type == "TABLE",
            )
        )
    assert table is not None
    assert result.status == NEEDS_REVIEW
    assert TABLE_STRUCTURE_UNCERTAIN in json.loads(table.metadata_json or "{}")["flags"]
    stored = json.loads(table.table_json or "{}")
    assert stored["merged_cells"]
    assert stored["merged_cells"][0]["vertical_merge"] == ("restart" if vertical else None)


def test_xlsx_merged_ranges_and_multiple_sheets_are_traceable(
    services: tuple[DocumentIntakeService, NativeHSMTExtractionService, Database],
    tmp_path: Path,
) -> None:
    intake, extractor, database = services
    document = intake.intake_file(_merged_multisheet_xlsx(tmp_path / "merged.xlsx"))

    result = extractor.extract_document(document.document_id)

    with database.session() as session:
        evidence = list(
            session.scalars(
                select(DocumentEvidence)
                .where(DocumentEvidence.extraction_id == result.extraction_id)
                .order_by(DocumentEvidence.ordinal)
            )
        )
    assert result.status == NEEDS_REVIEW
    assert {item.sheet_name for item in evidence} == {"BOQ", "Bảng giá"}
    merged = next(
        item
        for item in evidence
        if TABLE_STRUCTURE_UNCERTAIN in json.loads(item.metadata_json or "{}")["flags"]
    )
    assert json.loads(merged.table_json or "{}")["merged_ranges"] == ["A1:B1"]


def test_extraction_keeps_tender_and_document_version_separate(
    services: tuple[DocumentIntakeService, NativeHSMTExtractionService, Database],
    tmp_path: Path,
) -> None:
    intake, extractor, database = services
    _add_tender(database, "IB-A")
    _add_tender(database, "IB-B")
    first = intake.intake_file(_text_pdf(tmp_path / "first.pdf", "Tender A version one"), tender_reference="IB-A")
    second = intake.intake_file(_text_pdf(tmp_path / "second.pdf", "Tender A version two"), tender_reference="IB-A")
    other = intake.intake_file(_text_pdf(tmp_path / "other.pdf", "Tender B version one"), tender_reference="IB-B")

    results = [extractor.extract_document(document.document_id) for document in (first, second, other)]

    with database.session() as session:
        extractions = list(
            session.scalars(
                select(DocumentExtraction)
                .where(DocumentExtraction.id.in_([result.extraction_id for result in results]))
                .order_by(DocumentExtraction.document_id)
            )
        )
    assert [item.document_id for item in extractions] == sorted(
        [first.document_id, second.document_id, other.document_id]
    )
    metadata = [json.loads(item.metadata_json or "{}") for item in extractions]
    assert {item["source_sha256"] for item in metadata} == {
        first.sha256,
        second.sha256,
        other.sha256,
    }
    assert sorted(item["document_version"] for item in metadata) == [1, 1, 2]
