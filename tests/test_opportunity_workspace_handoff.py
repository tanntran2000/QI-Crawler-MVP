from __future__ import annotations

from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from qi_crawler.db import Database
from qi_crawler.market_intelligence.opportunity_contract import (
    OpportunityIdentity,
    OpportunityIdentityNamespace,
    OpportunityImportBatch,
    OpportunitySourceType,
)
from qi_crawler.market_intelligence.opportunity_radar import (
    OpportunityRadarItem,
    build_observation_key,
)
from qi_crawler.market_intelligence.opportunity_review import (
    OpportunityReviewDecision,
    OpportunityReviewService,
)
from qi_crawler.market_intelligence.opportunity_intelligence import OpportunityIntelligenceService
from qi_crawler.market_intelligence.khmt_contract import OBSERVED_KHMT_HEADERS
from qi_crawler.market_intelligence.tbmt_schema import OBSERVED_TBMT_HEADERS
from qi_crawler.migrations import upgrade_database
from qi_crawler.opportunity_review_persistence import SqlAlchemyOpportunityReviewRepository
from qi_crawler.tender_case import PlanContext
from qi_crawler.tender_workspace import TenderWorkspaceService


def _item(
    *,
    source_type: OpportunitySourceType = OpportunitySourceType.TBMT,
    raw_id: str = "IB2600463290-00",
    source_sha256: str = "a" * 64,
    source_row: int = 7,
) -> OpportunityRadarItem:
    identity = OpportunityIdentity.from_raw(raw_id)
    filename = f"{source_type.value.lower()}-handoff.xlsx"
    batch = OpportunityImportBatch(
        source_filename=filename,
        source_sha256=source_sha256,
        sheet=source_type.value,
        imported_at=datetime(2026, 1, 1, tzinfo=UTC),
        schema_version=f"{source_type.value.lower()}-v1",
        source_type=source_type,
    )
    provenance = {
        "source_filename": filename,
        "source_sha256": source_sha256,
        "sheet": batch.sheet,
        "source_row": source_row,
        "source_locator": f"{batch.sheet}!A{source_row}",
    }
    return OpportunityRadarItem(
        source_type=source_type,
        identity=identity,
        observation_key=build_observation_key(
            source_type=source_type,
            identity=identity,
            source_sha256=source_sha256,
            sheet=batch.sheet,
            source_row=source_row,
        ),
        source_filename=filename,
        source_sha256=source_sha256,
        sheet=batch.sheet,
        source_row=source_row,
        schema_version=batch.schema_version,
        package_name="Gói chuyển workspace",
        project="Dự án thử nghiệm",
        package_price_raw="1000",
        package_price=Decimal("1000"),
        funding_source="Ngân sách",
        source_fields={},
        raw_fields={},
        provenance=provenance,
    )


def _services(tmp_path: Path):
    database = Database(f"sqlite:///{tmp_path / 'handoff.db'}")
    upgrade_database(database.url, backup_dir=tmp_path / "backups")
    review = OpportunityReviewService(SqlAlchemyOpportunityReviewRepository(database))
    workspace = TenderWorkspaceService(database, tmp_path / "managed")
    from qi_crawler.opportunity_workspace_handoff import OpportunityWorkspaceHandoffService

    return database, review, workspace, OpportunityWorkspaceHandoffService(review, workspace)


def _confirm(review: OpportunityReviewService, item: OpportunityRadarItem) -> None:
    review.record_decision(item, decision=OpportunityReviewDecision.CONFIRMED, reviewer="Team Bid")


def _write_tbmt(path: Path) -> Path:
    common = {
        "BÊN MỜI THẦU": "Synthetic buyer",
        "ĐỊA CHỈ BÊN MỜI THẦU": "Synthetic address",
        "DỰ ÁN": "Synthetic project",
        "NỘI DUNG CHÍNH CỦA GÓI THẦU": "Synthetic scope",
        "NGUỒN VỐN": "Synthetic fund",
        "PHƯƠNG THỨC LỰA CHỌN NHÀ THẦU": "Một giai đoạn",
        "HÌNH THỨC LỰA CHỌN NHÀ THẦU": "Đấu thầu rộng rãi",
        "THỜI GIAN PHÁT HÀNH HSMT": "01/08/2026",
        "GIÁ BÁN 1 BỘ HSMT": 23,
        "BẢO ĐẢM DỰ THẦU": "23",
        "HÌNH THỨC BẢO ĐẢM DỰ THẦU": "Tiền mặt",
        "ĐỊA ĐIỂM PHÁT HÀNH": "Synthetic location",
        "THỜI GIAN ĐÓNG THẦU(HẠN CUỐI TIẾP NHẬN BG)": "10/08/2026",
        "THỜI GIAN MỞ THẦU": "10/08/2026",
        "THỜI GIAN THỰC HIỆN HỢP ĐỒNG": "120 ngày",
    }
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TBMT"
    sheet.append(OBSERVED_TBMT_HEADERS)
    values = {**common, "GÓI THẦU": "Gói nguồn IB2600463290-00", "GIÁ GÓI THẦU": "1000"}
    sheet.append([values.get(header) for header in OBSERVED_TBMT_HEADERS])
    workbook.save(path)
    return path


def _write_khmt(path: Path) -> Path:
    values = {
        "GÓI TIN": "Synthetic KHMT bulletin",
        "SỐ KẾ HOẠCH": "PL2600000001-00",
        "TÊN DỰ ÁN": "Synthetic project",
        "TÊN CHỦ ĐẦU TƯ": "Synthetic investor",
        "TỔNG MỨC ĐẦU TƯ": "900.000.000",
        "NỘI DUNG PHÊ DUYỆT": "Synthetic scope",
        "TÊN GÓI THẦU": "Synthetic package",
        "NGUỒN VỐN": "Synthetic fund",
        "GIÁ GÓI THẦU": "477.850.000",
        "HÌNH THỨC LỰA CHỌN": "Đấu thầu rộng rãi",
        "HÌNH THỨC HỢP ĐỒNG": "Trọn gói",
        "THỜI GIAN LỰA CHỌN": "Q4/2026",
        "THỜI GIAN THỰC HIỆN": "45 ngày",
    }
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "KHMT"
    sheet.append(OBSERVED_KHMT_HEADERS)
    sheet.append([values.get(header) for header in OBSERVED_KHMT_HEADERS])
    workbook.save(path)
    return path


@pytest.mark.parametrize(
    "decision",
    [None, OpportunityReviewDecision.REJECTED, OpportunityReviewDecision.NEEDS_REVIEW],
)
def test_non_confirmed_opportunity_cannot_handoff(tmp_path: Path, decision) -> None:
    _database, review, _workspace, handoff = _services(tmp_path)
    item = _item()
    if decision is not None:
        review.record_decision(item, decision=decision, reviewer="Team Bid")

    from qi_crawler.opportunity_workspace_handoff import OpportunityWorkspaceHandoffError

    with pytest.raises(OpportunityWorkspaceHandoffError):
        handoff.handoff(item)


def test_confirmed_then_rejected_cannot_handoff(tmp_path: Path) -> None:
    _database, review, workspace, handoff = _services(tmp_path)
    item = _item()
    _confirm(review, item)
    review.record_decision(item, decision=OpportunityReviewDecision.REJECTED, reviewer="Checker")

    from qi_crawler.opportunity_workspace_handoff import OpportunityWorkspaceHandoffError

    with pytest.raises(OpportunityWorkspaceHandoffError):
        handoff.handoff(item)
    assert workspace.search_cases(item.identity.base_id) == ()


def test_confirmed_tbmt_creates_case_and_exact_release(tmp_path: Path) -> None:
    database, review, workspace, handoff = _services(tmp_path)
    item = _item()
    _confirm(review, item)

    result = handoff.handoff(item)

    assert result.disposition.value == "CREATED_EXACT_RELEASE"
    assert result.case_id == "IB2600463290"
    assert result.release_raw_id == "IB2600463290-00"
    assert result.release_id is not None
    assert not result.human_link_required
    assert result.created_case
    assert result.created_release
    assert workspace.open_case(result.case_id).releases[0].identity.raw_id == item.identity.raw_id
    database.engine.dispose()


def test_repeating_tbmt_handoff_reopens_without_duplicate(tmp_path: Path) -> None:
    _database, review, workspace, handoff = _services(tmp_path)
    item = _item()
    _confirm(review, item)

    first = handoff.handoff(item)
    second = handoff.handoff(item)

    assert first.release_id == second.release_id
    assert second.disposition.value == "OPENED_EXISTING"
    assert len(workspace.open_case(item.identity.base_id).releases) == 1


def test_tbmt_same_lineage_adds_new_exact_revision(tmp_path: Path) -> None:
    _database, review, workspace, handoff = _services(tmp_path)
    first = _item(raw_id="IB2600463290-00")
    second = _item(raw_id="IB2600463290-01", source_row=8)
    _confirm(review, second)
    workspace.create_case("IB2600463290")
    workspace.add_release("IB2600463290", first.identity)

    result = handoff.handoff(second)

    assert result.disposition.value == "ADDED_EXACT_REVISION"
    assert [release.identity.raw_id for release in workspace.open_case("IB2600463290").releases] == [
        "IB2600463290-00",
        "IB2600463290-01",
    ]


def test_tbmt_without_exact_revision_rejects_without_mutation(tmp_path: Path) -> None:
    _database, review, workspace, handoff = _services(tmp_path)
    item = _item(raw_id="IB2600463290")
    _confirm(review, item)

    from qi_crawler.opportunity_workspace_handoff import OpportunityWorkspaceHandoffError

    with pytest.raises(OpportunityWorkspaceHandoffError):
        handoff.handoff(item)
    assert workspace.search_cases("IB2600463290") == ()


def test_ambiguous_tbmt_lineage_fails_closed(tmp_path: Path) -> None:
    _database, review, workspace, handoff = _services(tmp_path)
    item = _item(raw_id="IB2600463290-02")
    _confirm(review, item)
    workspace.create_case("case-a")
    workspace.add_release("case-a", "IB2600463290-00")
    workspace.create_case("case-b")
    workspace.add_release("case-b", "IB2600463290-01")

    from qi_crawler.opportunity_workspace_handoff import OpportunityWorkspaceHandoffError

    with pytest.raises(OpportunityWorkspaceHandoffError):
        handoff.handoff(item)
    assert workspace.search_cases("IB2600463290-02") == ()


def test_khmt_creates_provisional_plan_context_without_ib(tmp_path: Path) -> None:
    _database, review, workspace, handoff = _services(tmp_path)
    item = _item(
        source_type=OpportunitySourceType.KHMT,
        raw_id="PL2600000001-00",
    )
    _confirm(review, item)

    result = handoff.handoff(item)
    case = workspace.open_case(result.case_id)

    assert result.disposition.value == "CREATED_PROVISIONAL_CASE"
    assert result.release_id is None
    assert result.release_raw_id is None
    assert result.human_link_required
    assert case.plan_context is not None
    assert case.plan_context.identity.raw_id == item.identity.raw_id
    assert case.releases == ()


def test_repeating_exact_khmt_handoff_reopens_same_provisional_case(tmp_path: Path) -> None:
    _database, review, workspace, handoff = _services(tmp_path)
    item = _item(source_type=OpportunitySourceType.KHMT, raw_id="PL2600000001-00")
    _confirm(review, item)

    first = handoff.handoff(item)
    second = handoff.handoff(item)

    assert first.case_id == second.case_id == item.identity.raw_id
    assert second.disposition.value == "OPENED_EXISTING"
    assert workspace.open_case(item.identity.raw_id).releases == ()


def test_different_khmt_revision_requires_human_resolution(tmp_path: Path) -> None:
    _database, review, workspace, handoff = _services(tmp_path)
    first = _item(source_type=OpportunitySourceType.KHMT, raw_id="PL2600000001-00")
    second = _item(source_type=OpportunitySourceType.KHMT, raw_id="PL2600000001-01", source_row=8)
    _confirm(review, second)
    workspace.create_case(first.identity.raw_id, plan_context=PlanContext(first.identity))

    from qi_crawler.opportunity_workspace_handoff import OpportunityWorkspaceHandoffError

    with pytest.raises(OpportunityWorkspaceHandoffError):
        handoff.handoff(second)
    assert workspace.search_cases(second.identity.raw_id) == ()


def test_ambiguous_khmt_exact_plan_fails_closed(tmp_path: Path) -> None:
    _database, review, workspace, handoff = _services(tmp_path)
    item = _item(source_type=OpportunitySourceType.KHMT, raw_id="PL2600000001-00")
    _confirm(review, item)
    workspace.create_case("case-a", plan_context=PlanContext(item.identity))
    workspace.create_case("case-b", plan_context=PlanContext(item.identity))

    from qi_crawler.opportunity_workspace_handoff import OpportunityWorkspaceHandoffError

    with pytest.raises(OpportunityWorkspaceHandoffError):
        handoff.handoff(item)
    assert {result.case_id for result in workspace.search_cases(item.identity.raw_id)} == {
        "case-a",
        "case-b",
    }


def test_restart_reopens_tbmt_exact_case_and_release(tmp_path: Path) -> None:
    database, review, workspace, handoff = _services(tmp_path)
    item = _item()
    _confirm(review, item)
    first = handoff.handoff(item)
    database.engine.dispose()

    database2 = Database(f"sqlite:///{tmp_path / 'handoff.db'}")
    review2 = OpportunityReviewService(SqlAlchemyOpportunityReviewRepository(database2))
    workspace2 = TenderWorkspaceService(database2, tmp_path / "managed")
    from qi_crawler.opportunity_workspace_handoff import OpportunityWorkspaceHandoffService

    second = OpportunityWorkspaceHandoffService(review2, workspace2).handoff(item)

    assert second.disposition.value == "OPENED_EXISTING"
    assert second.release_id == first.release_id


def test_restart_reopens_khmt_provisional_case(tmp_path: Path) -> None:
    database, review, workspace, handoff = _services(tmp_path)
    item = _item(source_type=OpportunitySourceType.KHMT, raw_id="PL2600000001-00")
    _confirm(review, item)
    first = handoff.handoff(item)
    database.engine.dispose()

    database2 = Database(f"sqlite:///{tmp_path / 'handoff.db'}")
    review2 = OpportunityReviewService(SqlAlchemyOpportunityReviewRepository(database2))
    workspace2 = TenderWorkspaceService(database2, tmp_path / "managed")
    from qi_crawler.opportunity_workspace_handoff import OpportunityWorkspaceHandoffService

    second = OpportunityWorkspaceHandoffService(review2, workspace2).handoff(item)

    assert second.disposition.value == "OPENED_EXISTING"
    assert second.case_id == first.case_id
    assert workspace2.open_case(first.case_id).releases == ()


def test_tbmt_import_confirm_handoff_restart_preserves_exact_revision(tmp_path: Path) -> None:
    database, review, workspace, handoff = _services(tmp_path)
    source = _write_tbmt(tmp_path / "TBMT-source.xlsx")
    loaded = OpportunityIntelligenceService(review).load_workbook(
        source, OpportunitySourceType.TBMT
    )
    assert loaded.items[0].identity.raw_id == "IB2600463290-00"
    _confirm(review, loaded.items[0])

    first = handoff.handoff(loaded.items[0])
    database.engine.dispose()
    database2 = Database(f"sqlite:///{tmp_path / 'handoff.db'}")
    review2 = OpportunityReviewService(SqlAlchemyOpportunityReviewRepository(database2))
    workspace2 = TenderWorkspaceService(database2, tmp_path / "managed")
    from qi_crawler.opportunity_workspace_handoff import OpportunityWorkspaceHandoffService

    second = OpportunityWorkspaceHandoffService(review2, workspace2).handoff(loaded.items[0])
    assert first.release_raw_id == second.release_raw_id == "IB2600463290-00"
    assert second.disposition.value == "OPENED_EXISTING"


def test_khmt_import_confirm_handoff_restart_stays_provisional(tmp_path: Path) -> None:
    database, review, workspace, handoff = _services(tmp_path)
    source = _write_khmt(tmp_path / "KHMT-source.xlsx")
    loaded = OpportunityIntelligenceService(review).load_workbook(
        source, OpportunitySourceType.KHMT
    )
    assert loaded.items[0].identity.raw_id == "PL2600000001-00"
    _confirm(review, loaded.items[0])

    first = handoff.handoff(loaded.items[0])
    database.engine.dispose()
    database2 = Database(f"sqlite:///{tmp_path / 'handoff.db'}")
    review2 = OpportunityReviewService(SqlAlchemyOpportunityReviewRepository(database2))
    workspace2 = TenderWorkspaceService(database2, tmp_path / "managed")
    from qi_crawler.opportunity_workspace_handoff import OpportunityWorkspaceHandoffService

    second = OpportunityWorkspaceHandoffService(review2, workspace2).handoff(loaded.items[0])
    case = workspace2.open_case(first.case_id)
    assert first.human_link_required and second.human_link_required
    assert second.disposition.value == "OPENED_EXISTING"
    assert case.plan_context.identity.raw_id == "PL2600000001-00"
    assert case.releases == ()
