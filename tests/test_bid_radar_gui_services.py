from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook

from qi_crawler import gui_services
from qi_crawler.config import AppConfig
from qi_crawler.gui_services import (
    run_bid_radar_import_search,
    run_bid_radar_legal_docx,
    run_bid_radar_review,
)
from qi_crawler.market_intelligence.active_tender_context import (
    ActiveTenderContext,
    exact_identity_matches,
)
from qi_crawler.market_intelligence.khmt_contract import OBSERVED_KHMT_HEADERS
from qi_crawler.market_intelligence.opportunity_contract import OpportunitySourceType
from qi_crawler.market_intelligence.search import TargetedSearchRequest
from qi_crawler.market_intelligence.source_detection import SourceType
from qi_crawler.market_intelligence.tbmt_schema import OBSERVED_TBMT_HEADERS


def _context_item(raw_id: str, *, source_sha256: str = "a" * 64) -> SimpleNamespace:
    base_id, revision = raw_id.rsplit("-", 1)
    return SimpleNamespace(
        source_type=SimpleNamespace(value="TBMT"),
        identity=SimpleNamespace(raw_id=raw_id, base_id=base_id, revision=revision),
        source_sha256=source_sha256,
        observation_key=f"{source_sha256}:{raw_id}",
        package_name="Gói mẫu",
    )


def test_active_context_keeps_exact_revision_and_provenance() -> None:
    active = ActiveTenderContext.from_item(_context_item("IB12345678-00"))

    assert active.exact_identity == ("TBMT", "IB12345678-00", "IB12345678", "00")
    assert exact_identity_matches(active, _context_item("IB12345678-00"))
    assert not exact_identity_matches(active, _context_item("IB12345678-01"))
    assert not exact_identity_matches(active, _context_item("IB12345678-00", source_sha256="b" * 64))


def test_active_context_accepts_source_raw_id_spacing_with_normalized_revision() -> None:
    item = SimpleNamespace(
        source_type=SimpleNamespace(value="TBMT"),
        identity=SimpleNamespace(
            raw_id="IB2600488839- 00",
            base_id="IB2600488839",
            revision="00",
        ),
        source_sha256="a" * 64,
        observation_key="source:a:IB2600488839- 00",
        package_name="Gói nguồn thật",
    )

    active = ActiveTenderContext.from_item(item)

    assert active.raw_id == "IB2600488839- 00"
    assert active.base_id == "IB2600488839"
    assert active.revision == "00"
    assert exact_identity_matches(active, item)


def _write_tbmt(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TBMT"
    values = {
        "GÓI THẦU": "Gói máy chủ IB2600463290-00",
        "BÊN MỜI THẦU": "Đơn vị mời thầu",
        "DỰ ÁN": "Dự án tổng hợp",
        "GIÁ GÓI THẦU": "1000000",
        "HÌNH THỨC LỰA CHỌN NHÀ THẦU": "Đấu thầu rộng rãi",
    }
    for column, header in enumerate(OBSERVED_TBMT_HEADERS, start=1):
        sheet.cell(row=1, column=column, value=header)
        sheet.cell(row=2, column=column, value=values.get(header))
    workbook.save(path)
    return path


def _write_khmt(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "KHMT"
    values = {
        "GÓI TIN": "Thông báo tổng hợp",
        "SỐ KẾ HOẠCH": "PL2600265077-00",
        "TÊN DỰ ÁN": "Dự án mẫu",
        "TÊN CHỦ ĐẦU TƯ": "Chủ đầu tư mẫu",
        "TỔNG MỨC ĐẦU TƯ": "1000000",
        "NỘI DUNG PHÊ DUYỆT": "Nội dung mẫu",
        "TÊN GÓI THẦU": "Gói mẫu",
        "NGUỒN VỐN": "Nguồn vốn mẫu",
        "GIÁ GÓI THẦU": "1000000",
        "HÌNH THỨC LỰA CHỌN": "Đấu thầu rộng rãi",
        "HÌNH THỨC HỢP ĐỒNG": "Trọn gói",
        "THỜI GIAN LỰA CHỌN": "Q4/2026",
        "THỜI GIAN THỰC HIỆN": "45 ngày",
    }
    for column, header in enumerate(OBSERVED_KHMT_HEADERS, start=1):
        sheet.cell(row=1, column=column, value=header)
        sheet.cell(row=2, column=column, value=values.get(header))
    workbook.save(path)
    return path


def test_tbmt_adapter_routes_real_workbook_through_source_neutral_backend(tmp_path: Path) -> None:
    source = _write_tbmt(tmp_path / "TBMT.xlsx")
    config = AppConfig()
    config.storage.database_url = f"sqlite:///{tmp_path / 'bid-radar.db'}"

    result = run_bid_radar_import_search(
        config,
        source,
        TargetedSearchRequest(),
        source_type=SourceType.TBMT,
    )

    assert result.source_type is OpportunitySourceType.TBMT
    assert result.items[0].identity.namespace.value == "IB"
    assert result.items[0].selection_method_raw == "Đấu thầu rộng rãi"
    assert result.items[0].selection_method == "DAU_THAU_RONG_RAI"
    assert result.rows[0].item is result.items[0]
    assert result.rows[0].disposition == "UNFILTERED"
    assert result.unfiltered_count == 1


def test_bid_radar_row_propagates_structured_filter_evidence(tmp_path: Path) -> None:
    source = _write_tbmt(tmp_path / "TBMT-evidence.xlsx")
    config = AppConfig()
    config.storage.database_url = f"sqlite:///{tmp_path / 'bid-radar-evidence.db'}"

    result = run_bid_radar_import_search(
        config,
        source,
        TargetedSearchRequest(include_keywords=("đơn vị",)),
        source_type=SourceType.TBMT,
    )

    assert result.rows[0].criteria[0].evidence[0].field == "procuring_entity"
    assert result.rows[0].criteria[0].evidence[0].matched_terms == ("don vi",)


def test_filter_evidence_is_independent_of_human_review_state(tmp_path: Path) -> None:
    source = _write_tbmt(tmp_path / "TBMT-review-evidence.xlsx")
    config = AppConfig()
    config.storage.database_url = f"sqlite:///{tmp_path / 'bid-radar-review-evidence.db'}"
    request = TargetedSearchRequest(include_keywords=("đơn vị",))

    before = run_bid_radar_import_search(config, source, request, source_type=SourceType.TBMT)
    run_bid_radar_review(config, before.items[0], "CONFIRMED", "Team Bid")
    after = run_bid_radar_import_search(config, source, request, source_type=SourceType.TBMT)

    assert before.rows[0].criteria == after.rows[0].criteria
    assert before.rows[0].review_state == "UNREVIEWED"
    assert after.rows[0].review_state == "CONFIRMED"


def test_khmt_review_and_legal_docx_keep_existing_output_capability(tmp_path: Path) -> None:
    source = _write_khmt(tmp_path / "KHMT.xlsx")
    config = AppConfig()
    config.storage.database_url = f"sqlite:///{tmp_path / 'bid-radar-khmt.db'}"
    config.storage.report_dir = tmp_path / "reports"

    loaded = run_bid_radar_import_search(
        config,
        source,
        TargetedSearchRequest(),
        source_type=SourceType.KHMT,
    )
    assert run_bid_radar_review(config, loaded.items[0], "CONFIRMED", "Team Bid") == "CONFIRMED"
    results = run_bid_radar_legal_docx(
        config,
        loaded.load_result,
        source_path=source,
        expected_source_sha256=loaded.source_sha256,
    )

    assert len(results) == 1
    assert results[0].output.name == "ThongTin_PL2600265077.docx"


def test_tbmt_legal_docx_has_explicit_bounded_message(tmp_path: Path) -> None:
    source = _write_tbmt(tmp_path / "TBMT.xlsx")
    config = AppConfig()
    config.storage.database_url = f"sqlite:///{tmp_path / 'bid-radar.db'}"
    loaded = run_bid_radar_import_search(
        config,
        source,
        TargetedSearchRequest(),
        source_type=SourceType.TBMT,
    )

    with pytest.raises(ValueError, match="chỉ hỗ trợ nguồn KHMT"):
        run_bid_radar_legal_docx(
            config,
            loaded.load_result,
            source_path=source,
            expected_source_sha256=loaded.source_sha256,
        )


def test_database_upgrade_adapter_uses_explicit_migration_and_verifies_schema(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    config = AppConfig()
    config.storage.database_url = f"sqlite:///{tmp_path / 'unready.db'}"
    config.storage.report_dir = tmp_path / "reports"
    calls: list[tuple[str, Path]] = []

    class FakeDatabase:
        def __init__(self, url: str) -> None:
            calls.append((url, Path("database-constructed")))

        def require_current_schema(self) -> None:
            calls.append(("verified", Path("database-verified")))

    monkeypatch.setattr(gui_services, "Database", FakeDatabase)
    monkeypatch.setattr(
        gui_services,
        "upgrade_database",
        lambda url, backup_dir: SimpleNamespace(
            revision="0020_add_tender_operational_revision_events",
            backup_path=tmp_path / "backups" / "egp.db",
        ),
        raising=False,
    )

    result = gui_services.run_database_upgrade(config)

    assert result.revision == "0020_add_tender_operational_revision_events"
    assert result.database_path.name == "unready.db"
    assert result.backup_path == tmp_path / "backups" / "egp.db"
    assert calls[-1][0] == "verified"
